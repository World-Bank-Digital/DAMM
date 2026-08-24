#!/usr/bin/env python3
"""DAMM v1.7 scoring workbook builder — three files (blank template, Egypt, Nigeria).

Changes from v1.6:
  * Source URL + Source tier entry columns (Source-Tier Protocol made part of the record)
  * New "Tiers" reference sheet (tier definitions + starter domain lookup)
  * New "Issues" sheet (QC Protocol issues log)
  * Gate sign-off block on Read Me (QC Protocol G1/G2/G3)
  * Matrix encodes the two loop-1 rulings: UC:AI (7.12) binds AGI; a universal
    prerequisite at "Present (narrow)" caps every column at Partial
  * Loop-1 clean-slate data for both worked examples

Column map (Scoring sheet)
  A ID · B Indicator · C Pillar · D Layer · E Use cases · F Prereq · G Meth · H Dir · I–L thresholds   [model]
  M Value · N Source · O Source URL · P Tier · Q Year · R Assessor level                               [entry]
  S Class · T Level · U Stale · V–AA use-case flags                                                    [derived]
"""
import json, os, sys
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.chart import RadarChart, BarChart, Reference

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
from engine_v17 import MODEL

OUTDIR = os.path.join(HERE, "workbooks-v1.7")
os.makedirs(OUTDIR, exist_ok=True)

# ---------- house style ----------
TEAL   = "FF1F5F5B"
PALE   = "FFEAF1F0"
YELLOW = "FFFFF9C4"
GREY   = "FF898781"
F_TITLE = Font(name="Arial", size=14, bold=True, color=TEAL)
F_HEAD  = Font(name="Arial", size=10, bold=True, color="FFFFFFFF")
F_SECT  = Font(name="Arial", size=11, bold=True, color=TEAL)
F_BODY  = Font(name="Arial", size=10)
F_BOLD  = Font(name="Arial", size=10, bold=True)
F_ENTRY = Font(name="Arial", size=10, color="FF0000FF")
F_NOTE  = Font(name="Arial", size=9, color=GREY)
FILL_HEAD  = PatternFill("solid", fgColor=TEAL)
FILL_SECT  = PatternFill("solid", fgColor=PALE)
FILL_ENTRY = PatternFill("solid", fgColor=YELLOW)
WRAP = Alignment(wrap_text=True, vertical="top")
TOP  = Alignment(vertical="top")
THIN = Side(style="thin", color="FFD8DDDA")
BOX  = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

R0 = 5                      # first data row
IDS = list(MODEL.keys())
RN = R0 + len(IDS) - 1      # last data row (61)
PILLARS = ["A1", "C1", "C2", "C3", "C4", "E1", "O1"]
PILLAR_LABEL = {"A1": "A1 Agriculture & need", "C1": "C1 Connectivity & access", "C2": "C2 Data & DPI",
                "C3": "C3 Policy & safeguards", "C4": "C4 People & institutions",
                "E1": "E1 Innovation & emerging tech", "O1": "O1 Outcomes & inclusion"}
LAYERS = ["Foundation", "Enablers", "Transformation", "Outcomes"]

# dashboard geometry
P0 = 66                                   # pillar rows 66..72
L0 = 76                                   # layer rows 76..79
GAPROW = 81                               # leapfrog
PQ0 = 86                                  # prerequisite rows 86..97
PREREQ_ORDER = ["2.1", "2.9", "4.1", "3.3", "3.11", "4.5", "4.7", "4.9", "5.5", "5.7", "6.14", "7.12"]
PRQ = {pid: PQ0 + i for i, pid in enumerate(PREREQ_ORDER)}
PREREQ_SCOPE = {"2.1": "Universal", "2.9": "Universal", "4.1": "Universal",
                "3.3": "Use case: FIN, AGI", "3.11": "Use case: AGI", "4.5": "Use case: AGI",
                "4.7": "Use case: FIN", "4.9": "Delivery flag", "5.5": "Use case: ADV",
                "5.7": "Delivery flag", "6.14": "Use case: FIN", "7.12": "Use case: AI-enabled (binds AGI)"}
UNIVERSAL = ["2.1", "2.9", "4.1"]
UC_ROWS = 102                             # matrix rows 102..107
UCS = [("ADV", "Advisory & extension", []),
       ("SMF", "Smart farming", []),
       ("MKT", "Market linkage", []),
       ("SCM", "Supply chain", []),
       ("FIN", "Financial services", ["3.3", "4.7", "6.14"]),
       ("AGI", "Ag intelligence", ["3.3", "3.11", "4.5", "7.12"])]
UC_SPECIFIC = {"ADV": ["5.5"], "SMF": [], "MKT": [], "SCM": [], "FIN": ["3.3", "4.7", "6.14"],
               "AGI": ["3.3", "3.11", "4.5", "7.12"]}

TIER_ROWS = [
    ("T1", "Official statistics & international databases",
     "FAOSTAT · AQUASTAT · World Bank WDI/ID4D/Findex · ITU · UN DESA · national statistical offices",
     "Any claim. The default for every number."),
    ("T2", "Peer-reviewed literature & IO flagship reports",
     "Journals · systematic reviews · World Bank / FAO / IFPRI / CGIAR flagships · evaluated impact studies",
     "Any claim. Required for impact and effectiveness claims."),
    ("T3", "Government legal, policy & budget artefacts",
     "Gazetted laws · adopted strategies · budgets · regulator reports · official programme pages",
     "Presence facts, policy intent, programme design — not outcomes."),
    ("T4", "Reputable grey literature",
     "GSMA · GIZ · USAID · AGRA · AfDB · donor evaluations · established industry analyses",
     "Context and sector narrative, flagged as T4; presence facts where T3 is silent."),
    ("T5", "News, vendor & market material",
     "Press · company sites · market-research releases · blogs",
     "Initiative register only, existence facts only. A T5-only citation derives the class Judged, never Documented."),
]


def col(i):
    return get_column_letter(i)


def put(ws, coord, value, font=F_BODY, fill=None, align=TOP, fmt=None, border=False):
    c = ws[coord]
    c.value = value
    c.font = font
    if fill: c.fill = fill
    c.alignment = align
    if fmt: c.number_format = fmt
    if border: c.border = BOX
    return c



def setup_page(ws, landscape=True, print_area=None, titles=None):
    """Every sheet must print legibly — a reviewer who never opens Excel reads the PDF."""
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.page_margins.left = ws.page_margins.right = 0.4
    ws.page_margins.top = ws.page_margins.bottom = 0.5
    if print_area: ws.print_area = print_area
    if titles: ws.print_title_rows = titles


# ---------------------------------------------------------------- sheets
def sheet_readme(wb, country, is_template):
    ws = wb.create_sheet("Read Me")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 3
    ws.column_dimensions["C"].width = 26
    ws.column_dimensions["D"].width = 30
    for c in "EFG": ws.column_dimensions[c].width = 17
    ws.column_dimensions["H"].width = 34

    put(ws, "C2", "DAMM v1.7 — Scoring Workbook", F_TITLE)
    put(ws, "C3", "Digital Agriculture Maturity Model · draft for review · August 2026", F_NOTE)
    subject = ("This file is the blank template. Copy it once per country and rename the scoring sheet."
               if is_template else
               f"This file is the {country} worked example, compiled 22 August 2026.")
    put(ws, "C5", subject, F_BODY)

    put(ws, "C7", "The four steps", F_SECT)
    steps = [
        ("1 · ASSEMBLE", "Fill the six yellow columns on the scoring sheet: Value, Source, Source URL, Tier, Year, "
                         "Assessor level. Every indicator gets a value, a source and a year — or a recorded gap. "
                         "No level may exist without a recorded value."),
        ("2 · DERIVE", "Automatic. Evidence class, level, staleness, pillar bands, the layer profile, prerequisite "
                       "status and the use-case matrix all compute by formula. Nothing here is typed."),
        ("3 · REPORT", "The diagnostic report renders from this workbook. It is the deliverable; this file is the "
                       "instrument behind it."),
        ("4 · USE", "The task team leader takes the report into roadmap preparation, closes the verify list on "
                    "mission, and executes the refresh list."),
    ]
    r = 8
    for a, b in steps:
        put(ws, f"C{r}", a, F_BOLD)
        put(ws, f"D{r}", b, F_BODY, align=WRAP)
        ws.merge_cells(f"D{r}:H{r}")
        ws.row_dimensions[r].height = 40
        r += 1

    r += 1
    put(ws, f"C{r}", "What to type, and where", F_SECT); r += 1
    legend = [
        ("Yellow cells, blue text", "The only cells an assessor edits: Value, Source, Source URL, Tier, Year, "
                                    "Assessor level (columns M–R of the scoring sheet)."),
        ("Everything else", "Model reference or derived by formula. Do not overwrite: a typed level that "
                            "contradicts its value is the defect this model exists to prevent."),
        ("Value", "A number, a citation (\"NDPA 2023 in force; NDPC enforcing\"), or a recorded gap beginning "
                  "DATA GAP — followed by the sources actually searched."),
        ("Tier", "T1–T5 per the Tiers sheet. Reported, never weighted. Machine-fetched rows arrive with the tier "
                 "already filled by the pipeline."),
        ("Assessor level", "1–5, for ladder indicators only. Threshold indicators score themselves from the value."),
    ]
    for a, b in legend:
        put(ws, f"C{r}", a, F_BOLD)
        put(ws, f"D{r}", b, F_BODY, align=WRAP)
        ws.merge_cells(f"D{r}:H{r}")
        ws.row_dimensions[r].height = 30
        r += 1

    r += 1
    put(ws, f"C{r}", "Quality control — gate sign-off", F_SECT); r += 1
    put(ws, f"D{r}", "Automated checks run at render time and block the report on failure. The three human gates "
                     "are recorded here (QC Protocol).", F_NOTE, align=WRAP)
    ws.merge_cells(f"D{r}:H{r}"); ws.row_dimensions[r].height = 26; r += 1
    hdr = ["Gate", "What it certifies", "Name", "Date", "Note"]
    for i, h in enumerate(hdr):
        put(ws, f"{col(3+i)}{r}", h, F_HEAD, FILL_HEAD, border=True)
    r += 1
    gates = [
        ("G1", "Assessor has confirmed every machine-filled row: value, source, year, tier; every recorded gap "
               "accepted or re-searched."),
        ("G2", "A second reviewer has re-checked 100% of prerequisite rows and 100% of judged rows, plus a 15% "
               "sample of the remainder."),
        ("G3", "Task team leader sign-off against the four prohibitions, before anything leaves the team."),
    ]
    for g, what in gates:
        put(ws, f"C{r}", g, F_BOLD, border=True)
        put(ws, f"D{r}", what, F_BODY, align=WRAP, border=True)
        for i in (5, 6, 7):
            put(ws, f"{col(i)}{r}", None, F_ENTRY, FILL_ENTRY, border=True)
        ws.row_dimensions[r].height = 58
        r += 1

    r += 1
    put(ws, f"C{r}", "The four prohibitions", F_SECT); r += 1
    for t in ["No cross-country ranking.",
              "No band used as a project development objective, disbursement-linked indicator, or disbursement condition.",
              "No automatic financing decisions.",
              "No public claim before human review."]:
        put(ws, f"D{r}", "· " + t, F_BODY); r += 1
    setup_page(ws, print_area=f"B1:H{r}")
    return ws


def sheet_config(wb):
    ws = wb.create_sheet("Config")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 16
    ws.column_dimensions["D"].width = 78
    put(ws, "B1", "Config — every constant with its meaning", F_TITLE)
    rows = [
        ("Assessment year", 2026, "ENTRY (yellow). Staleness is measured against this year."),
        ("Staleness limit (years)", 3, "A reading older than this is flagged STALE. One rule for all indicators."),
        ("Readiness threshold", 2.6, "A use-case column whose bearing indicators average below this reads Partial."),
        ("Leapfrog threshold", 1.5, "|mean(Foundation) − mean(Transformation)| above this raises a structural flag."),
        ("Band: Nascent", "1.0 – <1.5", "Half-open bands; a pillar band in parentheses means judged, gap and held rows "
                                        "outnumber the levelled measured and documented ones."),
        ("Band: Emerging", "1.5 – <2.5", ""),
        ("Band: Established", "2.5 – <3.5", ""),
        ("Band: Advanced", "3.5 – <4.5", ""),
        ("Band: Transformative", "4.5 – 5.0", ""),
        ("Universal prerequisites", "2.1, 2.9, 4.1", "Absence blocks every use-case column. A universal prerequisite "
                                                     "at Present (narrow) caps every column at Partial."),
        ("AI-enabled services", "7.12 binds AGI", "Ruling pending ratification: the consent-and-rights prerequisite "
                                                  "binds the agricultural-intelligence column."),
    ]
    r = 2
    for a, b, c in rows:
        put(ws, f"B{r}", a, F_BOLD)
        cell = put(ws, f"C{r}", b, F_ENTRY if r in (2, 3) else F_BODY,
                   FILL_ENTRY if r in (2, 3) else None)
        put(ws, f"D{r}", c, F_NOTE, align=WRAP)
        ws.row_dimensions[r].height = 26 if c else 14
        r += 1
    setup_page(ws, print_area=f"B1:D{r}")
    return ws


def sheet_ladder(wb):
    ws = wb.create_sheet("Ladder")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 30
    ws.column_dimensions["D"].width = 84
    put(ws, "B1", "The shared ladder — scoring an indicator that is not a measurement", F_TITLE)
    put(ws, "B2", "Ask in stages: presence first, then quality, then scale. Stop at the first rung the evidence "
                  "does not support.", F_NOTE)
    hdr = ["Level", "Condition", "Evidence required"]
    for i, h in enumerate(hdr):
        put(ws, f"{col(2+i)}4", h, F_HEAD, FILL_HEAD, border=True)
    rungs = [
        (1, "Absent", "The search that established absence: sources consulted, in order."),
        (2, "Announced", "The announcement, draft or commitment, cited."),
        (3, "Adopted, or operating narrowly", "The instrument or system itself, with its adoption date or its "
                                              "operating footprint."),
        (4, "Operating with quality", "Governance, funding and institutionalisation — the evidence that it is run, "
                                      "not just live."),
        (5, "Operating at scale, with evidence of use", "Coverage or reach in the indicator's own unit, from a "
                                                        "source that is not the operator's own claim."),
    ]
    r = 5
    for lv, cond, ev in rungs:
        put(ws, f"B{r}", lv, F_BOLD, border=True)
        put(ws, f"C{r}", cond, F_BODY, align=WRAP, border=True)
        put(ws, f"D{r}", ev, F_BODY, align=WRAP, border=True)
        ws.row_dimensions[r].height = 30
        r += 1
    r += 1
    put(ws, f"B{r}", "A scale figure reported only by the operator of the service does not carry rung 5.", F_NOTE)
    setup_page(ws, print_area=f"B1:D{r}")
    return ws


def sheet_tiers(wb):
    ws = wb.create_sheet("Tiers")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["B"].width = 8
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 62
    ws.column_dimensions["E"].width = 52
    put(ws, "B1", "Source tiers — what may carry what", F_TITLE)
    put(ws, "B2", "Tiers are reported beside the evidence class and never weighted. Search in tier order and stop "
                  "when found; a dead end is recorded as a gap with its trail.", F_NOTE)
    for i, h in enumerate(["Tier", "What it is", "Typical sources", "Admissible for"]):
        put(ws, f"{col(2+i)}4", h, F_HEAD, FILL_HEAD, border=True)
    r = 5
    for t, what, typical, adm in TIER_ROWS:
        put(ws, f"B{r}", t, F_BOLD, border=True)
        put(ws, f"C{r}", what, F_BODY, align=WRAP, border=True)
        put(ws, f"D{r}", typical, F_BODY, align=WRAP, border=True)
        put(ws, f"E{r}", adm, F_BODY, align=WRAP, border=True)
        ws.row_dimensions[r].height = 34
        r += 1
    r += 1
    for t in ["Every citation is a deep link to the document, with an access date. A link that resolves only to a "
              "domain root is not a citation.",
              "A load-bearing number needs two independent T1–T2 confirmations, or it ships flagged single-source.",
              "Vintage beats tier at the margin for narrative; the scored row still records the official figure and "
              "its staleness honestly."]:
        put(ws, f"C{r}", "· " + t, F_NOTE, align=WRAP)
        ws.merge_cells(f"C{r}:E{r}")
        ws.row_dimensions[r].height = 26
        r += 1
    setup_page(ws, print_area=f"B1:E{r}")
    return ws


def sheet_issues(wb):
    ws = wb.create_sheet("Issues")
    ws.sheet_view.showGridLines = False
    widths = {"B": 6, "C": 12, "D": 10, "E": 56, "F": 12, "G": 46, "H": 16, "I": 14}
    for k, v in widths.items(): ws.column_dimensions[k].width = v
    put(ws, "B1", "Issues log — quality control findings", F_TITLE)
    put(ws, "B2", "Every automated failure, peer-review disagreement and sign-off exception is recorded here. "
                  "Fixes change entries on the scoring sheet; derived fields are never edited.", F_NOTE)
    hdr = ["#", "Date", "Class", "Finding", "Row / ID", "Action", "Resolver", "Status"]
    for i, h in enumerate(hdr):
        put(ws, f"{col(2+i)}4", h, F_HEAD, FILL_HEAD, border=True)
    for r in range(5, 45):
        for i in range(len(hdr)):
            put(ws, f"{col(2+i)}{r}", None, F_ENTRY, FILL_ENTRY, align=WRAP, border=True)
    dv = DataValidation(type="list", formula1='"D1 data,D2 protocol,D3 instrument,D4 engine,D5 presentation,D6 spec"',
                        allow_blank=True)
    ws.add_data_validation(dv); dv.add(f"D5:D44")
    dv2 = DataValidation(type="list", formula1='"Open,Fixed,Accepted,Superseded"', allow_blank=True)
    ws.add_data_validation(dv2); dv2.add(f"I5:I44")
    ws.freeze_panes = "B5"
    setup_page(ws, print_area="B1:I44", titles="4:4")
    return ws


def sheet_scoring(wb, sheet_name, country, rows):
    ws = wb.create_sheet(sheet_name)
    ws.sheet_view.showGridLines = False
    widths = {"A": 7, "B": 44, "C": 7, "D": 15, "E": 14, "F": 10, "G": 6, "H": 5,
              "I": 7, "J": 7, "K": 7, "L": 7, "M": 46, "N": 34, "O": 30, "P": 6,
              "Q": 7, "R": 9, "S": 12, "T": 7, "U": 8}
    for k, v in widths.items(): ws.column_dimensions[k].width = v
    for i in range(22, 28): ws.column_dimensions[col(i)].width = 5
    ws.column_dimensions[col(28)].width = 78
    ws.column_dimensions[col(29)].width = 16

    put(ws, "A1", f"{country} — DAMM v1.7 assessment", F_TITLE)
    put(ws, "A2", "Edit only the yellow columns (M–R). Everything else is model reference or derived by formula.",
        F_NOTE)

    headers = ["ID", "Indicator", "Pillar", "Layer", "Use cases", "Prereq", "Meth", "Dir",
               "L2", "L3", "L4", "L5", "Value", "Source", "Source URL", "Tier", "Year",
               "Assessor level", "Class", "Level", "Stale",
               "b_ADV", "b_SMF", "b_MKT", "b_SCM", "b_FIN", "b_AGI",
               "Open definition question (for ratification)", "Ratification hold"]
    for i, h in enumerate(headers, 1):
        put(ws, f"{col(i)}4", h, F_HEAD, FILL_HEAD, align=WRAP, border=True)
    ws.row_dimensions[4].height = 30

    for n, iid in enumerate(IDS):
        r = R0 + n
        m = MODEL[iid]
        put(ws, f"A{r}", iid, F_BODY, border=True)
        put(ws, f"B{r}", m["name"], F_BODY, align=WRAP, border=True)
        put(ws, f"C{r}", m["pillar"], F_BODY, border=True)
        put(ws, f"D{r}", m["layer"], F_BODY, border=True)
        put(ws, f"E{r}", ",".join(m["uc"]), F_BODY, border=True)
        put(ws, f"F{r}", m["prereq"], F_BODY, border=True)
        put(ws, f"G{r}", "T" if m["kind"] == "t" else "L", F_BODY, border=True)
        put(ws, f"H{r}", m["dir"], F_BODY, border=True)
        for j, th in enumerate(m["th"]):
            put(ws, f"{col(9+j)}{r}", th, F_BODY, border=True)

        d = (rows or {}).get(iid, {})
        val, src, url = d.get("value"), d.get("src"), d.get("url")
        tier, yr = d.get("tier"), d.get("year")
        alevel = d.get("level") if (m["kind"] == "l" and d.get("cls") not in (None, "Gap")) else None
        # threshold rows with a non-numeric value still need the assessor level recorded
        if m["kind"] == "t" and d.get("cls") in ("Documented", "Judged") and d.get("level") is not None:
            alevel = d.get("level")
        for cl, v in [("M", val), ("N", src), ("O", url), ("P", tier), ("Q", yr), ("R", alevel)]:
            put(ws, f"{cl}{r}", v, F_ENTRY, FILL_ENTRY, align=WRAP if cl in "MNO" else TOP, border=True)

        put(ws, f"S{r}", f'=IF($M{r}="","",IF(ISNUMBER($M{r}),"Measured",'
                         f'IF(ISNUMBER(SEARCH("DATA GAP",UPPER($M{r}))),"Gap",'
                         f'IF(AND($N{r}<>"",$P{r}<>"T5"),"Documented","Judged"))))', F_BODY, border=True)
        put(ws, f"T{r}", f'=IF($AC{r}<>"","",IF(OR($S{r}="",$S{r}="Gap"),"",IF($S{r}="Measured",'
                         f'1+IF($H{r}="H",IF($M{r}>=$I{r},1,0)+IF($M{r}>=$J{r},1,0)+IF($M{r}>=$K{r},1,0)+IF($M{r}>=$L{r},1,0),'
                         f'IF($M{r}<=$I{r},1,0)+IF($M{r}<=$J{r},1,0)+IF($M{r}<=$K{r},1,0)+IF($M{r}<=$L{r},1,0)),'
                         f'IF($R{r}<>"",$R{r},""))))', F_BODY, border=True)
        put(ws, f"U{r}", f'=IF(OR($S{r}="",$S{r}="Gap",$Q{r}=""),"",'
                         f'IF($Q{r}<Config!$C$2-Config!$C$3,"STALE",""))', F_BODY, border=True)
        for j, uc in enumerate(["ADV", "SMF", "MKT", "SCM", "FIN", "AGI"]):
            put(ws, f"{col(22+j)}{r}",
                f'=IF(OR(ISNUMBER(SEARCH("ALL",$E{r})),ISNUMBER(SEARCH("{uc}",$E{r}))),1,0)', F_BODY, border=True)
        if d and d.get("level") is None and d.get("cls") != "Gap":
            put(ws, f"{col(29)}{r}", "HOLD", F_ENTRY, FILL_ENTRY, border=True)
        note = (d.get("defnote") or "")
        if note:
            mark = "\u26a0 " if d.get("defsev") == "asserts-falsehood" else ""
            put(ws, f"{col(28)}{r}", mark + note, F_NOTE, align=WRAP, border=True)

    cands = {k: v for k, v in (rows or {}).items() if k.startswith("A1-CAND-")}
    if cands:
        cr = UC_ROWS + 8
        put(ws, f"A{cr}", "Proposed additions — recorded for context, outside every aggregate (ratification item)",
            F_SECT, FILL_SECT)
        ws.merge_cells(f"A{cr}:H{cr}")
        cr += 1
        for k, v in cands.items():
            nm = ("Cereal import dependency ratio (%)" if "IMP" in k else
                  "Irrigation coverage (% of cultivated/agricultural land equipped)")
            put(ws, f"A{cr}", k, F_BODY, border=True)
            put(ws, f"B{cr}", nm, F_BODY, align=WRAP, border=True)
            put(ws, f"C{cr}", "A1?", F_BODY, border=True)
            put(ws, f"M{cr}", v.get("value"), F_ENTRY, FILL_ENTRY, align=WRAP, border=True)
            put(ws, f"N{cr}", v.get("src"), F_ENTRY, FILL_ENTRY, align=WRAP, border=True)
            put(ws, f"O{cr}", v.get("url"), F_ENTRY, FILL_ENTRY, align=WRAP, border=True)
            put(ws, f"P{cr}", v.get("tier"), F_ENTRY, FILL_ENTRY, border=True)
            put(ws, f"Q{cr}", v.get("year"), F_ENTRY, FILL_ENTRY, border=True)
            put(ws, f"S{cr}", "not scored", F_NOTE, border=True)
            if v.get("defnote"): put(ws, f"{col(28)}{cr}", v["defnote"], F_NOTE, align=WRAP, border=True)
            cr += 1

    dvt = DataValidation(type="list", formula1='"T1,T2,T3,T4,T5"', allow_blank=True)
    ws.add_data_validation(dvt); dvt.add(f"P{R0}:P{RN}")
    dvl = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    ws.add_data_validation(dvl); dvl.add(f"R{R0}:R{RN}")

    # ---------------- dashboard ----------------
    put(ws, f"A{P0-2}", "Pillar profile  ·  n = indicators in the pillar; Rated = those that produced a level, "
                        "which is what the mean averages; Held = level withheld pending ratification. "
                        "A band in parentheses = judged rows, recorded gaps and held rows outnumber the "
                        "levelled measured and documented ones", F_SECT, FILL_SECT)
    for i, h in enumerate(["Pillar", "n", "Rated", "Mean", "Band", "Measured", "Documented",
                           "Judged", "Gap", "Held", "Stale"]):
        put(ws, f"{col(1+i)}{P0-1}", h, F_HEAD, FILL_HEAD, border=True)
    for i, p in enumerate(PILLARS):
        r = P0 + i
        put(ws, f"A{r}", p, F_BOLD, border=True)
        put(ws, f"B{r}", f'=COUNTIF($C${R0}:$C${RN},$A{r})', F_BODY, border=True)
        # Rated is the mean's real denominator: levels are numbers, so ">0" counts exactly the
        # rows AVERAGEIF averages, and excludes gaps and ratification holds.
        put(ws, f"C{r}", f'=COUNTIFS($C${R0}:$C${RN},$A{r},$T${R0}:$T${RN},">0")', F_BODY, border=True)
        put(ws, f"D{r}", f'=IFERROR(ROUND(AVERAGEIF($C${R0}:$C${RN},$A{r},$T${R0}:$T${RN}),2),"")',
            F_BODY, border=True)
        # Ruling 13.1: the band is the level the pillar mean rounds to. These four cuts
        # are the same numbers as engine_v17.BANDS; verifier stage 6 holds them together.
        band = ('IF($D{r}<1.5,"Nascent",IF($D{r}<2.5,"Emerging",IF($D{r}<3.5,"Established",'
                'IF($D{r}<4.5,"Advanced","Transformative"))))').format(r=r)
        # Judged rows carrying a level, counted separately: a Judged row sits on the weak side of
        # the test and must not also be counted among the levelled measured/documented rows.
        jr = f'COUNTIFS($C${R0}:$C${RN},$A{r},$S${R0}:$S${RN},"Judged",$T${R0}:$T${RN},">0")'
        put(ws, f"E{r}", f'=IF($D{r}="","Not rated",IF(({jr}+$I{r}+$J{r})>($C{r}-{jr}),'
                         f'"("&{band}&")",{band}))', F_BOLD, border=True)
        for j, cls in enumerate(["Measured", "Documented", "Judged", "Gap"]):
            put(ws, f"{col(6+j)}{r}", f'=COUNTIFS($C${R0}:$C${RN},$A{r},$S${R0}:$S${RN},"{cls}")',
                F_BODY, border=True)
        put(ws, f"J{r}", f'=COUNTIFS($C${R0}:$C${RN},$A{r},$AC${R0}:$AC${RN},"<>")', F_BODY, border=True)
        put(ws, f"K{r}", f'=COUNTIFS($C${R0}:$C${RN},$A{r},$U${R0}:$U${RN},"STALE")', F_BODY, border=True)

    put(ws, f"A{L0-1}", "Layer profile and structural gap", F_SECT, FILL_SECT)
    for i, L in enumerate(LAYERS):
        r = L0 + i
        put(ws, f"A{r}", L, F_BOLD, border=True)
        put(ws, f"B{r}", f'=IFERROR(ROUND(AVERAGEIF($D${R0}:$D${RN},$A{r},$T${R0}:$T${RN}),2),"")',
            F_BODY, border=True)
    put(ws, f"A{GAPROW}", "Gap (Foundation − Transformation)", F_BOLD)
    put(ws, f"B{GAPROW}", f'=IF(OR($B${L0}="",$B${L0+2}=""),"",ROUND($B${L0}-$B${L0+2},2))', F_BOLD, border=True)
    put(ws, f"C{GAPROW}", f'=IF($B${GAPROW}="","",IF(ABS($B${GAPROW})>Config!$C$5,'
                          f'IF($B${GAPROW}>0,"FLAG — unrealized foundations","FLAG — services ahead of their rails"),'
                          f'"No structural flag"))', F_BODY, border=True)

    put(ws, f"A{PQ0-2}", "Prerequisites  (presence only; Unverified = the presence could not be evidenced)",
        F_SECT, FILL_SECT)
    for i, h in enumerate(["ID", "Prerequisite", "Scope", "Status"]):
        put(ws, f"{col(1+i)}{PQ0-1}", h, F_HEAD, FILL_HEAD, border=True)
    for pid, r in PRQ.items():
        put(ws, f"A{r}", pid, F_BODY, border=True)
        put(ws, f"B{r}", f'=INDEX($B${R0}:$B${RN},MATCH($A{r},$A${R0}:$A${RN},0))', F_BODY, align=WRAP, border=True)
        put(ws, f"C{r}", PREREQ_SCOPE[pid], F_BODY, border=True)
        cls_ = f'INDEX($S${R0}:$S${RN},MATCH($A{r},$A${R0}:$A${RN},0))'
        lvl_ = f'INDEX($T${R0}:$T${RN},MATCH($A{r},$A${R0}:$A${RN},0))'
        put(ws, f"D{r}", f'=IF(OR({cls_}="Gap",{cls_}="",{lvl_}=""),"Unverified",'
                         f'IF({lvl_}>=3,"Present",IF({lvl_}=2,"Present (narrow)","Absent")))', F_BOLD, border=True)

    put(ws, f"A{UC_ROWS-2}", "Use-case readiness  (feeds Playbook Step 2B — Digital Readiness)", F_SECT, FILL_SECT)
    for i, h in enumerate(["Use case area", "Mean of bearing indicators", "Status", "Blocking / limiting factor"]):
        put(ws, f"{col(1+i)}{UC_ROWS-1}", h, F_HEAD, FILL_HEAD, border=True)

    def anyeq(pids, state):
        return ",".join(f'$D${PRQ[p]}="{state}"' for p in pids)

    for i, (code, label, _) in enumerate(UCS):
        r = UC_ROWS + i
        spec = UC_SPECIFIC[code]
        put(ws, f"A{r}", label, F_BOLD, border=True)
        put(ws, f"B{r}", f'=IFERROR(ROUND(AVERAGEIF({col(22+i)}${R0}:{col(22+i)}${RN},1,$T${R0}:$T${RN}),2),"")',
            F_BODY, border=True)
        # precedence mirrors the engine: a KNOWN block outranks an UNKNOWN prerequisite
        blocked = anyeq(UNIVERSAL + spec, "Absent")
        unver = anyeq(UNIVERSAL + spec, "Unverified")
        narrow_col = anyeq(spec, "Present (narrow)") if spec else None
        narrow_uni = anyeq(UNIVERSAL, "Present (narrow)")
        partial_terms = []
        if narrow_col: partial_terms.append(f'OR({narrow_col})')
        partial_terms.append(f'$B{r}<Config!$C$4')
        partial_terms.append(f'OR({narrow_uni})')
        status = (f'=IF(OR({blocked}),"Blocked",'
                  f'IF(OR({unver}),"Unverified",'
                  f'IF(OR({",".join(partial_terms)}),"Partial","Ready")))')
        put(ws, f"C{r}", status, F_BOLD, border=True)
        # why: name every prerequisite for this column that is not Present, else thin enablers
        parts = []
        for p in UNIVERSAL + spec:
            pr = PRQ[p]
            parts.append(f'IF($D${pr}<>"Present",$A${pr}&" ","")')
        why = ("&".join(parts)) if parts else '""'
        put(ws, f"D{r}", f'=IF($C{r}="Ready","",IF(TRIM({why})<>"",TRIM({why}),'
                         f'IF($B{r}<Config!$C$4,"thin enablers (mean below "&TEXT(Config!$C$4,"0.0")&")","")))',
            F_BODY, align=WRAP, border=True)

    for i in range(22, 28):                     # hide the derived use-case flag helpers
        ws.column_dimensions[col(i)].hidden = True
    ws.freeze_panes = "C5"
    setup_page(ws, print_area=f"A1:U{UC_ROWS+5}", titles="4:4")  # note column AB is read on screen, not in print
    return ws


def sheet_visuals(wb, sheet_name, years):
    ws = wb.create_sheet("Visuals")
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 30
    for c in "BCDEF": ws.column_dimensions[c].width = 13
    q = f"'{sheet_name}'" if " " in sheet_name else sheet_name
    put(ws, "A1", "Visuals — charts read live from the scoring sheet", F_TITLE)
    put(ws, "A2", "Read a level chart together with the evidence-mix chart: a strong bar standing on a mostly grey "
                  "evidence bar is a hypothesis, not a finding.", F_NOTE)

    for i, h in enumerate(["Pillar", "Mean", "Measured", "Documented", "Judged", "Gap"]):
        put(ws, f"{col(1+i)}3", h, F_BOLD)
    for i, p in enumerate(PILLARS):
        r = 4 + i
        put(ws, f"A{r}", PILLAR_LABEL[p], F_BODY)
        put(ws, f"B{r}", f'={q}!$D${P0+i}', F_BODY)
        for j, cl in enumerate("FGHI"):
            put(ws, f"{col(3+j)}{r}", f'={q}!${cl}${P0+i}', F_BODY)

    put(ws, "A12", "Layer", F_BOLD); put(ws, "B12", "Mean", F_BOLD)
    for i, L in enumerate(LAYERS):
        put(ws, f"A{13+i}", L, F_BODY)
        put(ws, f"B{13+i}", f'={q}!$B${L0+i}', F_BODY)

    put(ws, "A18", "Level", F_BOLD); put(ws, "B18", "Indicators", F_BOLD)
    for lv in range(1, 6):
        put(ws, f"A{18+lv}", f"Level {lv}", F_BODY)
        put(ws, f"B{18+lv}", f'=COUNTIF({q}!$T${R0}:$T${RN},{lv})', F_BODY)

    put(ws, "A25", "Year", F_BOLD); put(ws, "B25", "Readings", F_BOLD)
    for i, y in enumerate(years):
        put(ws, f"A{26+i}", str(y), F_BODY)
        put(ws, f"B{26+i}", f'=SUMPRODUCT(({q}!$Q${R0}:$Q${RN}={y})*({q}!$S${R0}:$S${RN}<>"Gap"))', F_BODY)
    last_y = 26 + len(years) - 1
    put(ws, f"A{last_y+2}", "Note: readings more than three years old are flagged STALE on the scoring sheet.",
        F_NOTE)

    def place(chart, anchor, w=13, h=8, title=None):
        chart.width, chart.height = w, h
        if title: chart.title = title
        chart.style = 2
        ws.add_chart(chart, anchor)

    cats = Reference(ws, min_col=1, min_row=4, max_row=10)
    radar = RadarChart(); radar.type = "marker"
    radar.add_data(Reference(ws, min_col=2, min_row=3, max_row=10), titles_from_data=True)
    radar.set_categories(cats); radar.y_axis.scaling.max = 5; radar.y_axis.scaling.min = 0
    place(radar, "H3", 11, 9, "Pillar profile (1–5)")

    bar = BarChart(); bar.type = "bar"
    bar.add_data(Reference(ws, min_col=2, min_row=3, max_row=10), titles_from_data=True)
    bar.set_categories(cats); bar.x_axis.scaling.max = 5
    place(bar, "H22", 11, 9, "Pillar mean level")

    mix = BarChart(); mix.type = "bar"; mix.grouping = "stacked"; mix.overlap = 100
    mix.add_data(Reference(ws, min_col=3, max_col=6, min_row=3, max_row=10), titles_from_data=True)
    mix.set_categories(cats)
    place(mix, "H41", 11, 9, "Evidence mix per pillar")

    lay = BarChart(); lay.type = "col"
    lay.add_data(Reference(ws, min_col=2, min_row=12, max_row=16), titles_from_data=True)
    lay.set_categories(Reference(ws, min_col=1, min_row=13, max_row=16)); lay.y_axis.scaling.max = 5
    place(lay, "S3", 10, 8, "Layer profile")

    lvl = BarChart(); lvl.type = "col"
    lvl.add_data(Reference(ws, min_col=2, min_row=18, max_row=23), titles_from_data=True)
    lvl.set_categories(Reference(ws, min_col=1, min_row=19, max_row=23))
    place(lvl, "S20", 10, 8, "Indicators by level")

    vin = BarChart(); vin.type = "col"
    vin.add_data(Reference(ws, min_col=2, min_row=25, max_row=last_y), titles_from_data=True)
    vin.set_categories(Reference(ws, min_col=1, min_row=26, max_row=last_y))
    place(vin, "S37", 10, 8, "Evidence vintage")
    setup_page(ws, print_area=f"A1:AB{max(last_y + 2, 56)}")
    return ws


# ---------------------------------------------------------------- build
def build(country, rows, outfile, is_template=False):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    sheet_readme(wb, country, is_template)
    sheet_config(wb)
    sheet_ladder(wb)
    sheet_tiers(wb)
    sheet_name = "Scoring" if is_template else country
    sheet_scoring(wb, sheet_name, country, rows)
    sheet_issues(wb)
    years = list(range(2010, 2027)) if is_template else sorted(
        {r["year"] for r in rows.values() if r.get("year") and r.get("cls") != "Gap"} | {2026})
    sheet_visuals(wb, sheet_name, years)
    wb.save(outfile)
    return outfile


def load_rows(iso):
    return json.load(open(os.path.join(HERE, f"{iso}_v17_input.json")))


if __name__ == "__main__":
    outs = []
    outs.append(build("Country", None, os.path.join(OUTDIR, "DAMM-v1.7-Scoring-Workbook-Blank-Template.xlsx"),
                      is_template=True))
    outs.append(build("Egypt", load_rows("EGY"), os.path.join(OUTDIR, "DAMM-v1.7-Scoring-Workbook-Egypt.xlsx")))
    outs.append(build("Nigeria", load_rows("NGA"), os.path.join(OUTDIR, "DAMM-v1.7-Scoring-Workbook-Nigeria.xlsx")))
    for o in outs:
        print("built", os.path.basename(o))
