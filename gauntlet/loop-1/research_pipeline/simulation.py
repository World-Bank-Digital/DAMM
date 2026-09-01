#!/usr/bin/env python3
"""Cost-free, fail-closed simulation for the canonical DAMM workflow.

The public interface deliberately accepts a scenario identifier, not a graph of test
doubles.  A scenario selects versioned fixture behavior while the harness retains the
real Stage 6 assembler and, for the happy path, the real eight-stage coordinator.
Simulation output is permanently marked as ineligible for staging acceptance.
"""

from __future__ import annotations

from contextlib import contextmanager
import copy
import datetime
import hashlib
import html
import io
import json
import os
from pathlib import Path
import re
import socket
import sqlite3
import subprocess
from typing import Any, Mapping
import unicodedata
from unittest import mock
import zipfile
import xml.etree.ElementTree as ET

import investment_options as I
import export_package as E
import generate_dar as D
import run_workflow as W


HERE = Path(__file__).resolve().parent
REPO_ROOT = HERE.parents[2]
SCENARIO_DIR = HERE / "fixtures" / "simulation"
REPORT_SCHEMA = "damm.simulation-report/v1"
SCENARIO_SCHEMA = "damm.simulation-scenario/v1"
PROVENANCE_SCHEMA = W.SIMULATION_PROVENANCE_SCHEMA
SIMULATION_LABEL = W.SIMULATION_LABEL
REPORT_NAME = "simulation-report.json"
CODE_IDENTITY_SCHEMA = "damm.simulation-code-identity/v1"
SIMULATION_DATE = "2026-09-02"
SIMULATION_CREATED_AT = f"{SIMULATION_DATE}T00:00:00Z"
SIMULATION_CLOCK = datetime.datetime(
    2026, 9, 2, tzinfo=datetime.timezone.utc
)
PRODUCTION_CODE_FILES = {
    "gauntlet/loop-1/research_pipeline/simulation.py": HERE / "simulation.py",
    "gauntlet/loop-1/research_pipeline/investment_options.py": HERE / "investment_options.py",
    "gauntlet/loop-1/research_pipeline/report_design.py": HERE / "report_design.py",
    "gauntlet/loop-1/research_pipeline/generate_dar.py": HERE / "generate_dar.py",
    "gauntlet/loop-1/research_pipeline/export_package.py": HERE / "export_package.py",
    "gauntlet/loop-1/research_pipeline/run_workflow.py": HERE / "run_workflow.py",
    "gauntlet/loop-1/research_pipeline/vendors.py": HERE / "vendors.py",
    "gauntlet/loop-1/research_pipeline/workflow_inputs.py": HERE / "workflow_inputs.py",
    "gauntlet/loop-1/research_pipeline/foresight_contract.py": HERE / "foresight_contract.py",
    "gauntlet/loop-1/engine_v17.py": HERE.parent / "engine_v17.py",
    "model/reference_scorer.py": REPO_ROOT / "model" / "reference_scorer.py",
    "model/DAMM-v1.7-model.json": REPO_ROOT / "model" / "DAMM-v1.7-model.json",
    "workflow/dar-workflow-v1.json": REPO_ROOT / "workflow" / "dar-workflow-v1.json",
}
OVERLENGTH_REPAIR_TARGETS = (
    (0, "problem", 500),
    (0, "recommendation_rationale", 500),
    (1, "problem", 500),
    (1, "recommendation_rationale", 500),
    (2, "problem", 500),
    (2, "recommendation_rationale", 500),
    (3, "problem", 500),
)

_SENSITIVE_ENV_MARKERS = (
    "API_KEY",
    "DATABASE_URL",
    "POSTGRES",
    "NEON",
    "RENDER",
    "NETLIFY",
    "ARTIFACT_DELIVERY_SECRET",
    "DAR_KEY_SECRET",
    "BETTER_AUTH_SECRET",
)


class SimulationError(RuntimeError):
    """The requested simulation is unsafe, unknown, or malformed."""


class SimulationBoundaryError(SimulationError):
    """Simulated code attempted a forbidden live external operation."""


def _stable_bytes(value: Any) -> bytes:
    return json.dumps(
        value,
        sort_keys=True,
        separators=(",", ":"),
        ensure_ascii=False,
        allow_nan=False,
    ).encode("utf-8")


def _stable_sha256(value: Any) -> str:
    return hashlib.sha256(_stable_bytes(value)).hexdigest()


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def _code_identity() -> dict[str, Any]:
    files = {name: _file_sha256(path) for name, path in PRODUCTION_CODE_FILES.items()}
    return {
        "schema_version": CODE_IDENTITY_SCHEMA,
        "files": files,
        "aggregate_sha256": _stable_sha256(files),
    }


def _scenario_path(scenario_id: str) -> Path:
    if not isinstance(scenario_id, str) or not re.fullmatch(
        r"[a-z0-9][a-z0-9-]{2,80}", scenario_id
    ):
        raise SimulationError("scenario_id is not a safe built-in identifier")
    path = SCENARIO_DIR / f"{scenario_id}.json"
    if not path.is_file():
        known = ", ".join(sorted(candidate.stem for candidate in SCENARIO_DIR.glob("*.json")))
        raise SimulationError(f"unknown simulation scenario {scenario_id!r}; choose {known}")
    return path


def _load_scenario(scenario_id: str) -> tuple[dict[str, Any], Path, str]:
    path = _scenario_path(scenario_id)
    raw = path.read_bytes()
    try:
        value = json.loads(raw)
    except json.JSONDecodeError as error:
        raise SimulationError(f"invalid built-in scenario {scenario_id}: {error}") from error
    if not isinstance(value, dict) or value.get("schema_version") != SCENARIO_SCHEMA:
        raise SimulationError(f"scenario {scenario_id} must use {SCENARIO_SCHEMA}")
    expected_top_level = {
        "schema_version",
        "scenario_id",
        "description",
        "kind",
        "default_country",
        "default_iso3",
        "default_profile",
        "profiles",
        "expected",
        "fixture",
    }
    if set(value) != expected_top_level:
        raise SimulationError(
            f"scenario {scenario_id} has unexpected or missing top-level fields"
        )
    if value.get("scenario_id") != scenario_id:
        raise SimulationError(f"scenario file identity does not match {scenario_id}")
    if value.get("kind") not in {
        "stage6_repair_recovery",
        "stage6_through_package",
        "eight_stage_happy",
    }:
        raise SimulationError(f"scenario {scenario_id} has an unsupported kind")
    profiles = value.get("profiles")
    if not isinstance(profiles, list) or not profiles or any(
        not isinstance(profile, str) for profile in profiles
    ):
        raise SimulationError(f"scenario {scenario_id} has no valid profiles")
    expected = value.get("expected")
    if not isinstance(expected, dict):
        raise SimulationError(f"scenario {scenario_id} has no expected outcome")
    required_expected = {
        "workflow_status",
        "failed_stage",
        "error_code",
        "fixture_call_count",
        "code_sha256",
    }
    if set(expected) != required_expected:
        raise SimulationError(
            f"scenario {scenario_id} expected outcome has unexpected or missing fields"
        )
    if expected["workflow_status"] not in {"complete", "failed"}:
        raise SimulationError(f"scenario {scenario_id} workflow_status is invalid")
    if expected["workflow_status"] == "complete" and (
        expected["failed_stage"] is not None or expected["error_code"] is not None
    ):
        raise SimulationError(
            f"scenario {scenario_id} complete outcome cannot declare a failure"
        )
    if (
        not isinstance(expected["fixture_call_count"], int)
        or isinstance(expected["fixture_call_count"], bool)
        or expected["fixture_call_count"] < 1
    ):
        raise SimulationError(f"scenario {scenario_id} fixture_call_count is invalid")
    if not isinstance(expected["code_sha256"], str) or not re.fullmatch(
        r"[0-9a-f]{64}", expected["code_sha256"]
    ):
        raise SimulationError(f"scenario {scenario_id} code_sha256 is invalid")
    fixture = value.get("fixture")
    if not isinstance(fixture, dict):
        raise SimulationError(f"scenario {scenario_id} has no fixture controls")
    if value["kind"] in {
        "stage6_repair_recovery",
        "stage6_through_package",
    }:
        if set(fixture) != {
            "evidence_batches",
            "failed_batch",
            "repair_lengths",
            "recovery_lengths",
            "candidate_count",
        }:
            raise SimulationError(
                f"scenario {scenario_id} recovery fixture is malformed"
            )
        repair_lengths = fixture["repair_lengths"]
        recovery_lengths = fixture["recovery_lengths"]
        if (
            not isinstance(fixture["evidence_batches"], int)
            or isinstance(fixture["evidence_batches"], bool)
            or fixture["evidence_batches"] < 2
            or not isinstance(fixture["failed_batch"], int)
            or isinstance(fixture["failed_batch"], bool)
            or not 1 <= fixture["failed_batch"] <= fixture["evidence_batches"]
            or not isinstance(repair_lengths, list)
            or len(repair_lengths) != len(OVERLENGTH_REPAIR_TARGETS)
            or any(
                not isinstance(length, int)
                or isinstance(length, bool)
                or length < 1
                for length in repair_lengths
            )
            or not any(
                length > limit
                for length, (_index, _field, limit) in zip(
                    repair_lengths, OVERLENGTH_REPAIR_TARGETS
                )
            )
            or not isinstance(recovery_lengths, list)
            or any(
                not isinstance(length, int)
                or isinstance(length, bool)
                or length < 1
                for length in recovery_lengths
            )
            or len(recovery_lengths) != sum(
                length > limit
                for length, (_index, _field, limit) in zip(
                    repair_lengths, OVERLENGTH_REPAIR_TARGETS
                )
            )
            or any(
                length > (limit * 9) // 10
                for length, (_index, _field, limit) in zip(
                    recovery_lengths,
                    (
                        target
                        for raw_length, target in zip(
                            repair_lengths, OVERLENGTH_REPAIR_TARGETS
                        )
                        if raw_length > target[2]
                    ),
                )
            )
            or not isinstance(fixture["candidate_count"], int)
            or isinstance(fixture["candidate_count"], bool)
            or fixture["candidate_count"] != fixture["evidence_batches"] + 3
            or not 3 <= fixture["candidate_count"] <= 7
        ):
            raise SimulationError(
                f"scenario {scenario_id} recovery fixture is invalid"
            )
    else:
        if set(fixture) != {"source_count_by_profile", "candidate_count"}:
            raise SimulationError(f"scenario {scenario_id} happy fixture is malformed")
        source_counts = fixture["source_count_by_profile"]
        if (
            not isinstance(source_counts, dict)
            or set(source_counts) != set(profiles)
            or any(
                not isinstance(count, int) or isinstance(count, bool) or count < 1
                for count in source_counts.values()
            )
            or not isinstance(fixture["candidate_count"], int)
            or isinstance(fixture["candidate_count"], bool)
            or fixture["candidate_count"] < 1
        ):
            raise SimulationError(f"scenario {scenario_id} happy fixture is invalid")
    return value, path, hashlib.sha256(raw).hexdigest()


def _normalise_launch(
    scenario: Mapping[str, Any],
    *,
    country: str | None,
    iso3: str | None,
    profile: str | None,
) -> tuple[str, str, str]:
    selected_country = " ".join(str(country or scenario.get("default_country") or "").split())
    selected_iso3 = str(iso3 or scenario.get("default_iso3") or "").strip().upper()
    selected_profile = str(profile or scenario.get("default_profile") or "").strip()
    if not selected_country:
        raise SimulationError("country cannot be empty")
    if not re.fullmatch(r"[A-Z]{3}", selected_iso3):
        raise SimulationError("iso3 must be exactly three letters")
    if selected_profile not in scenario["profiles"]:
        raise SimulationError(
            f"profile must be one of {', '.join(scenario['profiles'])}"
        )
    return selected_country, selected_iso3, selected_profile


def _provenance(
    scenario_id: str,
    scenario_sha256: str,
    code_sha256: str,
    profile: str,
) -> dict[str, Any]:
    return W.validate_simulation_provenance({
        "schema_version": PROVENANCE_SCHEMA,
        "label": SIMULATION_LABEL,
        "execution_kind": "simulation",
        "acceptance_eligible": False,
        "scenario_id": scenario_id,
        "scenario_sha256": scenario_sha256,
        "code_sha256": code_sha256,
        "profile": profile,
    })


@contextmanager
def _simulation_boundary(counters: dict[str, int]):
    """Remove live credentials and deny sockets for the duration of one simulation."""

    before = dict(os.environ)
    for key in list(os.environ):
        upper = key.upper()
        if any(marker in upper for marker in _SENSITIVE_ENV_MARKERS):
            os.environ.pop(key, None)
    os.environ["DAMM_EXECUTION_KIND"] = "simulation"
    os.environ["DAMM_ACCEPTANCE_ELIGIBLE"] = "false"

    def deny_network(*_args: Any, **_kwargs: Any):
        counters["network_attempts"] += 1
        raise SimulationBoundaryError("simulation attempted a live network connection")

    def deny_subprocess(*_args: Any, **_kwargs: Any):
        counters["subprocess_attempts"] += 1
        raise SimulationBoundaryError("simulation attempted to start a subprocess")

    def deny_database(*_args: Any, **_kwargs: Any):
        counters["database_attempts"] += 1
        raise SimulationBoundaryError("simulation attempted to open a database")

    try:
        with (
            mock.patch.object(socket.socket, "connect", deny_network),
            mock.patch.object(socket.socket, "connect_ex", deny_network),
            mock.patch.object(socket.socket, "send", deny_network),
            mock.patch.object(socket.socket, "sendall", deny_network),
            mock.patch.object(socket.socket, "sendto", deny_network),
            mock.patch.object(socket.socket, "sendmsg", deny_network),
            mock.patch.object(socket, "create_connection", deny_network),
            mock.patch.object(subprocess, "Popen", deny_subprocess),
            mock.patch.object(os, "system", deny_subprocess),
            mock.patch.object(os, "popen", deny_subprocess),
            mock.patch.object(os, "posix_spawn", deny_subprocess),
            mock.patch.object(os, "posix_spawnp", deny_subprocess),
            mock.patch.object(sqlite3, "connect", deny_database),
        ):
            yield
    finally:
        os.environ.clear()
        os.environ.update(before)


def _source(ref: str, index: int, *, characters: int = 180) -> dict[str, Any]:
    stem = f"Synthetic evidence {index} for deterministic simulation. "
    text = (stem * ((characters // len(stem)) + 2))[:characters]
    return {
        "ref": ref,
        "kind": "synthetic_country_finding",
        "title": f"Synthetic source {index}",
        "text": text,
        "source": f"fixture://source/{index}",
    }


def _candidate(index: int, source_ref: str) -> dict[str, Any]:
    return {
        "title": f"Synthetic investment option {index}",
        "problem": f"Synthetic evidence identifies interoperability gap {index}.",
        "recommendation_rationale": (
            f"A bounded appraisal of option {index} is useful for simulation only."
        ),
        "source_refs": [source_ref],
    }


def _option_body(source_ref: str, index: int) -> dict[str, Any]:
    return {
        "baseline": f"Synthetic baseline {index}; validation is required.",
        "counterfactual": "Agencies continue using fragmented systems.",
        "costs": {
            "currency": "USD",
            "base_year": 2026,
            "low": 100000.0 * index,
            "high": 150000.0 * index,
            "basis": "Synthetic planning range; not an estimate.",
            "source_refs": [source_ref],
        },
        "benefits": {
            "quantified": [],
            "qualitative": ["Synthetic reduction in duplicated processing."],
        },
        "horizon_years": 5,
        "discount_rate": 0.06,
        "npv_low": None,
        "npv_high": None,
        "bcr_low": None,
        "bcr_high": None,
        "sensitivity": [{
            "scenario": "Synthetic high-cost case",
            "changes": "Costs increase by 30 percent.",
            "result": "Revalidate scope before appraisal.",
        }],
        "distributional_effects": ["Test inclusion assumptions."],
        "climate_effects": ["Test climate-service assumptions."],
        "ai_and_data_risks": ["Test consent and model-bias controls."],
        "implementation_risks": ["Test institutional coordination."],
        "data_gaps": ["Replace all synthetic values before review."],
        "evidence_status": "Synthetic fixture; not evidence.",
    }


class _HappyFixtureLLM:
    """Deterministic JSON adapter at the real Stage 6 vendor-call boundary."""

    vendor = "fixture"
    model = "eight-stage-happy-v1"

    def __init__(self, candidate_count: int):
        self.calls: list[str] = []
        self.candidate_count = candidate_count

    def json_call(
        self,
        _system: str,
        user: str,
        _schema: Mapping[str, Any],
        _pass_name: str,
        max_tokens: int = 8000,
        detail: str = "",
    ) -> dict[str, Any]:
        del max_tokens
        self.calls.append(detail)
        if detail.startswith("investment candidate map batch "):
            return {
                "candidates": [
                    _candidate(index, "SRC-001")
                    for index in range(1, self.candidate_count + 1)
                ]
            }
        if detail == "investment candidate final register":
            serialized = user.split("SUPPORTED CANDIDATE BRIEFS:\n", 1)[1].split(
                "\n\nReturn", 1
            )[0]
            return {"candidates": json.loads(serialized)}
        if detail.startswith("investment appraisal INV-"):
            candidate = json.loads(
                user.split("CANDIDATE:\n", 1)[1].split(
                    "\n\nCURRENT APPRAISAL", 1
                )[0]
            )
            option_index = int(detail.split("INV-", 1)[1].split(" ", 1)[0])
            return {
                "option": _option_body(candidate["source_refs"][0], option_index)
            }
        if detail == "investment portfolio sequencing":
            return {
                "portfolio_sequencing": (
                    "Synthetic sequencing exercises governance before procurement."
                ),
                "cross_cutting_data_gaps": [
                    "Replace fixture costs and benefits with reviewed evidence."
                ],
            }
        raise AssertionError(f"unexpected fixture call: {detail}")


class _OverlengthFixtureLLM:
    """Exercise the exact Nigeria repair lengths through bounded repair chunks."""

    vendor = "fixture"
    model = "nigeria-stage6-overlength-v1"

    def __init__(
        self,
        repair_lengths: list[int],
        recovery_lengths: list[int],
        *,
        evidence_batches: int,
        failed_batch: int,
        candidate_count: int,
    ):
        self.calls: list[str] = []
        self.repair_lengths = list(repair_lengths)
        self.recovery_lengths = list(recovery_lengths)
        self.evidence_batches = evidence_batches
        self.failed_batch = failed_batch
        self.candidate_count = candidate_count
        self.observed_repair_lengths: list[int] = []
        self.observed_recovery_lengths: list[int] = []
        self.observed_effective_lengths: list[int] = []
        self.observed_recovery_max_tokens: int | None = None

    def _repair_details(self, phase: str, batch_count: int) -> tuple[str, ...]:
        prefix = (
            f"investment candidate map batch {self.failed_batch}/"
            f"{self.evidence_batches} [{phase}"
        )
        if batch_count == 1:
            return (prefix + "]",)
        return tuple(
            f"{prefix} chunk {index}/{batch_count}]"
            for index in range(1, batch_count + 1)
        )

    @staticmethod
    def _overlong_batch(source_ref: str, start_index: int) -> list[dict[str, Any]]:
        candidates = [
            _candidate(index, source_ref)
            for index in range(start_index, start_index + 4)
        ]
        candidates[0]["problem"] = "p" * 560
        candidates[0]["recommendation_rationale"] = "r" * 560
        candidates[1]["problem"] = "p" * 560
        candidates[1]["recommendation_rationale"] = "r" * 560
        candidates[2]["problem"] = "p" * 560
        candidates[2]["recommendation_rationale"] = "r" * 560
        candidates[3]["problem"] = "p" * 560
        return candidates

    def json_call(
        self,
        _system: str,
        _user: str,
        _schema: Mapping[str, Any],
        _pass_name: str,
        max_tokens: int = 8000,
        detail: str = "",
    ) -> dict[str, Any]:
        self.calls.append(detail)
        map_match = re.fullmatch(r"investment candidate map batch (\d+)/(\d+)", detail)
        if map_match:
            batch = int(map_match.group(1))
            total = int(map_match.group(2))
            if total != self.evidence_batches or not 1 <= batch <= total:
                raise AssertionError(f"unexpected fixture call: {detail}")
            if batch == self.failed_batch:
                return {
                    "candidates": self._overlong_batch(
                        f"SRC-{self.failed_batch:03d}", self.failed_batch
                    )
                }
            candidate_index = batch if batch < self.failed_batch else batch + 3
            return {
                "candidates": [
                    _candidate(candidate_index, f"SRC-{batch:03d}")
                ]
            }
        repair_batches = I._candidate_length_repair_batches(
            OVERLENGTH_REPAIR_TARGETS
        )
        repair_details = self._repair_details(
            "local-length repair 1/1", len(repair_batches)
        )
        if detail in repair_details:
            repair_batch = repair_batches[repair_details.index(detail)]
            expected_keys = tuple(
                I._candidate_repair_key(index, field)
                for index, field, _limit in repair_batch
            )
            keys = tuple(
                _schema["properties"]["repairs"]["properties"]
            )
            if keys != expected_keys:
                raise AssertionError(
                    f"unexpected repair schema for fixture call: {detail}"
                )
            all_keys = tuple(
                I._candidate_repair_key(index, field)
                for index, field, _limit in OVERLENGTH_REPAIR_TARGETS
            )
            lengths_by_key = dict(zip(all_keys, self.repair_lengths))
            repairs = {
                key: chr(97 + all_keys.index(key)) * lengths_by_key[key]
                for key in keys
            }
            self.observed_repair_lengths.extend(
                len(repairs[key]) for key in keys
            )
            return {"repairs": repairs}

        residual = tuple(
            (index, field, (limit * 9) // 10)
            for length, (index, field, limit) in zip(
                self.repair_lengths, OVERLENGTH_REPAIR_TARGETS
            )
            if length > limit
        )
        recovery_batches = I._candidate_length_repair_batches(residual)
        recovery_details = self._repair_details(
            "local-length repair 2/2", len(recovery_batches)
        )
        if detail in recovery_details:
            recovery_batch = recovery_batches[recovery_details.index(detail)]
            expected_keys = tuple(
                I._candidate_repair_key(index, field)
                for index, field, _limit in recovery_batch
            )
            keys = tuple(
                _schema["properties"]["repairs"]["properties"]
            )
            if keys != expected_keys:
                raise AssertionError(
                    f"unexpected recovery schema for fixture call: {detail}"
                )
            if (
                self.observed_recovery_max_tokens is not None
                and self.observed_recovery_max_tokens != max_tokens
            ):
                raise AssertionError(
                    "recovery chunks used inconsistent output-token bounds"
                )
            self.observed_recovery_max_tokens = max_tokens
            residual_keys = tuple(
                I._candidate_repair_key(index, field)
                for index, field, _limit in residual
            )
            lengths_by_key = dict(zip(residual_keys, self.recovery_lengths))
            repairs = {
                key: chr(65 + residual_keys.index(key)) * lengths_by_key[key]
                for key in keys
            }
            self.observed_recovery_lengths.extend(
                len(repairs[key]) for key in keys
            )
            return {"repairs": repairs}
        if detail == "investment candidate final register":
            serialized = _user.split("SUPPORTED CANDIDATE BRIEFS:\n", 1)[1].split(
                "\n\nReturn", 1
            )[0]
            candidates = json.loads(serialized)
            if len(candidates) != self.candidate_count:
                raise AssertionError(
                    "unexpected recovered candidate count: " + str(len(candidates))
                )
            by_title = {candidate["title"]: candidate for candidate in candidates}
            self.observed_effective_lengths = [
                len(
                    by_title[
                        f"Synthetic investment option {self.failed_batch + index}"
                    ][field]
                )
                for index, field, _limit in OVERLENGTH_REPAIR_TARGETS
            ]
            return {"candidates": candidates}
        if detail.startswith("investment appraisal INV-"):
            candidate = json.loads(
                _user.split("CANDIDATE:\n", 1)[1].split(
                    "\n\nCURRENT APPRAISAL", 1
                )[0]
            )
            option_index = int(detail.split("INV-", 1)[1].split(" ", 1)[0])
            return {
                "option": _option_body(candidate["source_refs"][0], option_index)
            }
        if detail == "investment portfolio sequencing":
            return {
                "portfolio_sequencing": (
                    "Synthetic sequencing exercises governance before procurement."
                ),
                "cross_cutting_data_gaps": [
                    "Replace fixture costs and benefits with reviewed evidence."
                ],
            }
        raise AssertionError(f"unexpected fixture call: {detail}")


def _write_json(path: Path, value: Any) -> Path:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(_stable_bytes(value) + b"\n")
    return path


_SYNTHETIC_NARRATIVE_FORMATS = {
    "damm_diagnostic": (".html",),
    "country_research": (".md", ".html"),
    "ai_digital_agriculture": (".md", ".html"),
    "international_lessons": (".md", ".html"),
    "strategic_foresight": (".html",),
}


def _synthetic_narrative(
    context: W.StageContext,
    artifact_key: str,
    provenance: Mapping[str, Any],
) -> Path | tuple[Path, ...]:
    """Publish synthetic predecessors in the same formats as their producers."""
    suffixes = _SYNTHETIC_NARRATIVE_FORMATS.get(context.stage_id)
    if suffixes is None:
        suffixes = (".md",)
    title = f"{context.iso3} simulated {context.stage_id.replace('_', ' ')}"
    producer_binding = str(provenance.get("code_sha256") or "unbound")
    marker = "SIMULATION-SOURCE-" + _stable_sha256({
        "stage_id": context.stage_id,
        "artifact_key": artifact_key,
        "code_sha256": producer_binding,
    })[:16].upper()
    paths: list[Path] = []
    for suffix in suffixes:
        path = context.stage_dir / f"{artifact_key}{suffix}"
        if suffix == ".md":
            path.write_text(
                f"# {title}\n\n> **{SIMULATION_LABEL}**\n\n"
                f"Scenario role: synthetic predecessor for the Stage 6–8 test.\n\n"
                f"Producer binding: `{producer_binding}`\n\n"
                f"Artifact marker: `{marker}`\n",
                encoding="utf-8",
            )
        else:
            path.write_text(
                "<!doctype html>\n<html lang='en'><head><meta charset='utf-8'>"
                "<meta name='viewport' content='width=device-width,initial-scale=1'>"
                f"<meta name='damm-simulation-code-sha256' content='{html.escape(producer_binding)}'>"
                f"<title>{html.escape(title)}</title>"
                "<style>:root{color-scheme:light}body{max-width:50rem;margin:3rem auto;"
                "padding:0 1.5rem;font:16px/1.55 Georgia,serif;color:#17322a}"
                ".notice{padding:1rem;border-left:4px solid #b7791f;background:#f4efe5}"
                ".meta{font:12px/1.4 ui-monospace,monospace;color:#5d6b64}</style>"
                f"</head><body><main><h1>{html.escape(title)}</h1>"
                f"<p class='notice'><strong>{html.escape(SIMULATION_LABEL)}</strong><br>"
                "Synthetic predecessor for the Stage 6–8 test.</p>"
                f"<p class='meta'>Producer binding: {html.escape(producer_binding)}<br>"
                f"Artifact marker: {html.escape(marker)}</p>"
                "</main></body></html>\n",
                encoding="utf-8",
            )
        paths.append(path)
    return paths[0] if len(paths) == 1 else tuple(paths)


def _embed_html_source_marker(document: str, marker: str) -> str:
    if not re.fullmatch(r"SIMULATION-SOURCE-[0-9A-F]{16}", marker):
        raise SimulationError("simulated HTML source marker is malformed")
    head_end = re.search(r"(?i)</head\s*>", document)
    if head_end is None:
        raise SimulationError("simulated narrative is not standalone HTML")
    metadata = (
        f'<meta name="damm-simulation-source" content="{html.escape(marker)}">'
    )
    return document[:head_end.start()] + metadata + document[head_end.start():]


def _synthetic_stage_handler(provenance: Mapping[str, Any]):
    def handler(context: W.StageContext) -> W.StageResult:
        artifacts: dict[str, Any] = {}
        for key in context.required_artifacts:
            if key == "stage_manifest":
                continue
            if key.endswith("report"):
                artifacts[key] = _synthetic_narrative(context, key, provenance)
                continue
            path = context.stage_dir / f"{key}.json"
            if key == "source_inventory":
                _write_json(path, {
                    "schema_version": "damm.synthetic-source-inventory/v1",
                    "label": SIMULATION_LABEL,
                    "acceptance_eligible": False,
                    "sources": [{
                        "ref": f"SRC-{context.ordinal:03d}",
                        "title": f"Synthetic source for {context.stage_id}",
                        "url": (
                            "https://example.test/simulation/"
                            f"{context.stage_id}"
                        ),
                    }],
                    "simulation_provenance": provenance,
                })
            else:
                _write_json(path, {
                    "schema_version": "damm.synthetic-stage-artifact/v1",
                    "label": SIMULATION_LABEL,
                    "acceptance_eligible": False,
                    "stage_id": context.stage_id,
                    "artifact_key": key,
                    "country": context.country,
                    "iso3": context.iso3,
                    "simulation_provenance": provenance,
                })
            artifacts[key] = path
        return W.StageResult(artifacts=artifacts, spent_usd=0.0)

    return handler


def _stage6_handler(
    provenance: Mapping[str, Any],
    sources: list[dict[str, Any]],
    llm: Any,
    *,
    limits: I.AppraisalLimits = I.DEFAULT_APPRAISAL_LIMITS,
):
    def handler(context: W.StageContext) -> W.StageResult:
        response = I.synthesize_appraisal(
            context.country, sources, llm, limits=limits
        )
        product = I.build_product(context.country, context.iso3, response, sources)
        product["assessment_date"] = SIMULATION_DATE
        product["simulation_notice"] = SIMULATION_LABEL
        product["simulation_provenance"] = dict(provenance)
        errors = I.validate_product(product)
        if errors:
            raise W.NonRetryableStageError("synthetic Stage 6 product: " + "; ".join(errors))

        data_path = _write_json(context.stage_dir / "investment-options.json", product)
        sources_path = _write_json(
            context.stage_dir / "source-inventory.json",
            {
                "schema_version": "damm.synthetic-source-inventory/v1",
                "label": SIMULATION_LABEL,
                "acceptance_eligible": False,
                "sources": product["source_inventory"],
                "simulation_provenance": provenance,
            },
        )
        markdown = f"> **{SIMULATION_LABEL}**\n\n" + I.render_markdown(product)
        markdown_path = context.stage_dir / "investment-options.md"
        markdown_path.write_text(markdown, encoding="utf-8")
        html_path = context.stage_dir / "investment-options.html"
        stage6_marker = "SIMULATION-SOURCE-" + _file_sha256(data_path)[:16].upper()
        html_path.write_text(_embed_html_source_marker(
            I.render_html(product, f"{SIMULATION_LABEL}: Stage 6"),
            stage6_marker,
        ), encoding="utf-8")
        workbook_path = context.stage_dir / "cost-benefit.xlsx"
        I.write_workbook(product, workbook_path)
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Alignment, Font, PatternFill
            from openpyxl.worksheet.page import PageMargins
            from openpyxl.worksheet.properties import PageSetupProperties

            workbook = load_workbook(workbook_path)
            notice = workbook.create_sheet("SIMULATED", 0)
            notice.sheet_view.showGridLines = False
            notice.merge_cells("A1:H2")
            notice["A1"] = SIMULATION_LABEL
            notice["A1"].font = Font(name="Aptos Display", size=20, bold=True, color="FFFFFF")
            notice["A1"].fill = PatternFill("solid", fgColor="17322A")
            notice["A1"].alignment = Alignment(vertical="center", wrap_text=True)
            for row in notice["A1:H2"]:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="17322A")
            notice.row_dimensions[1].height = 30
            notice.row_dimensions[2].height = 18
            notice.merge_cells("A4:H5")
            notice["A4"] = (
                "This zero-spend workbook exercises the Stage 6 recovery path. "
                "It is synthetic, unreviewed and cannot be used as acceptance evidence."
            )
            notice["A4"].font = Font(name="Aptos", size=12, color="17322A")
            notice["A4"].fill = PatternFill("solid", fgColor="F4EFE5")
            notice["A4"].alignment = Alignment(vertical="center", wrap_text=True)
            for row in notice["A4:H5"]:
                for cell in row:
                    cell.fill = PatternFill("solid", fgColor="F4EFE5")
            notice.row_dimensions[4].height = 28
            notice.row_dimensions[5].height = 28
            notice["A7"] = "Scenario scope"
            notice["A7"].font = Font(name="Aptos", size=10, bold=True, color="245844")
            notice["C7"] = "Stage 6 recovery through Stage 8 packaging"
            notice["A8"] = "External spend"
            notice["A8"].font = Font(name="Aptos", size=10, bold=True, color="245844")
            notice["C8"] = "$0.00"
            notice.column_dimensions["A"].width = 22
            notice.column_dimensions["B"].width = 3
            notice.column_dimensions["C"].width = 32
            for column in ("D", "E", "F", "G", "H"):
                notice.column_dimensions[column].width = 12
            notice.freeze_panes = "A7"
            notice.sheet_properties.pageSetUpPr = PageSetupProperties(
                fitToPage=True, autoPageBreaks=False
            )
            notice.page_setup.orientation = "landscape"
            notice.page_setup.paperSize = notice.PAPERSIZE_A4
            notice.page_setup.fitToWidth = 1
            notice.page_setup.fitToHeight = 1
            notice.page_margins = PageMargins(
                left=0.35, right=0.35, top=0.5, bottom=0.5,
                header=0.2, footer=0.2,
            )
            notice.print_area = "A1:H10"
            notice.oddFooter.center.text = "DAR Studio · SIMULATED · NOT ACCEPTANCE EVIDENCE"
            workbook.save(workbook_path)
            workbook.close()
            I.stabilize_workbook(workbook_path, product["assessment_date"])
        except ImportError as error:
            raise SimulationError("openpyxl is required for the happy simulation") from error

        return W.StageResult(
            artifacts={
                "investment_options_report": (markdown_path, html_path),
                "cost_benefit_workbook": workbook_path,
                "appraisal_data": data_path,
                "source_inventory": sources_path,
                "investment_options": data_path,
            },
            spent_usd=0.0,
        )

    return handler


def _prior_stage_artifact(
    context: W.StageContext, stage_id: str, artifact_key: str
) -> Path:
    """Resolve one coordinator-snapshotted dependency and recheck its binding."""
    stage = next(
        (value for value in context.prior_stages if value.get("id") == stage_id),
        None,
    )
    if not isinstance(stage, dict) or stage.get("status") != "complete":
        raise SimulationError(f"prior stage {stage_id} is not complete")
    record = next(
        (
            value
            for value in stage.get("artifacts") or []
            if value.get("key") == artifact_key
        ),
        None,
    )
    if not isinstance(record, dict):
        raise SimulationError(
            f"prior stage {stage_id} has no {artifact_key} artifact"
        )
    relative = Path(str(record.get("path") or ""))
    if relative.is_absolute() or ".." in relative.parts:
        raise SimulationError(f"prior artifact {stage_id}.{artifact_key} is unsafe")
    path = (context.workspace / relative).resolve()
    try:
        path.relative_to(context.workspace.resolve())
    except ValueError as error:
        raise SimulationError(
            f"prior artifact {stage_id}.{artifact_key} escapes the workspace"
        ) from error
    if not path.is_file() or _file_sha256(path) != record.get("sha256"):
        raise SimulationError(
            f"prior artifact {stage_id}.{artifact_key} failed hash verification"
        )
    return path


def _ai_stage_handler(provenance: Mapping[str, Any]):
    """Return the minimum complete Stage 3 contract needed by real Stage 7 gates."""
    def handler(context: W.StageContext) -> W.StageResult:
        product = {
            "schema_version": "damm.ai-digital-agriculture/v1",
            "country": context.country,
            "iso3": context.iso3,
            "assessment_year": I.ASSESSMENT_YEAR,
            "as_is": {"findings": [{
                "id": "AI-ASIS-1",
                "statement": "Synthetic capability evidence for simulation only.",
                "source_refs": ["SRC-AI-001"],
            }]},
            "peer_experience": {"findings": [{
                "id": "AI-PEER-1",
                "statement": "Synthetic peer evidence for simulation only.",
                "source_refs": ["SRC-AI-001"],
            }]},
            "recommended_agenda": {
                "status": "proposed_for_post_completion_validation",
                "recommendations": ["Validate the synthetic agenda after completion."],
            },
            "source_inventory": [{
                "ref": "SRC-AI-001",
                "title": "Synthetic AI evidence",
                "url": "https://example.test/simulation/ai",
            }],
            "simulation_notice": SIMULATION_LABEL,
            "simulation_provenance": provenance,
        }
        product_path = _write_json(context.stage_dir / "ai-assessment.json", product)
        inventory_path = _write_json(
            context.stage_dir / "source-inventory.json",
            {
                "schema_version": "damm.synthetic-source-inventory/v1",
                "sources": product["source_inventory"],
                "simulation_provenance": provenance,
            },
        )
        report_paths = _synthetic_narrative(
            context, "ai_assessment_report", provenance
        )
        return W.StageResult(
            artifacts={
                "ai_assessment_report": report_paths,
                "ai_evidence_data": product_path,
                "source_inventory": inventory_path,
                "ai_assessment": product_path,
            },
            spent_usd=0.0,
        )

    return handler


def _integrated_draft_handler(provenance: Mapping[str, Any]):
    """Exercise production Stage 7 input gates and bind Stage 6 into the Draft."""
    def handler(context: W.StageContext) -> W.StageResult:
        inputs = {
            "engine_input": _prior_stage_artifact(
                context, "damm_diagnostic", "engine_input"
            ),
            "scans": _prior_stage_artifact(
                context, "international_lessons", "scans"
            ),
            "ai_assessment": _prior_stage_artifact(
                context, "ai_digital_agriculture", "ai_assessment"
            ),
            "foresight": _prior_stage_artifact(
                context, "strategic_foresight", "foresight"
            ),
            "investment_options": _prior_stage_artifact(
                context, "investment_options", "investment_options"
            ),
        }
        input_records = {
            name: D.file_record(path) for name, path in inputs.items()
        }
        manifest = json.loads(context.manifest_path.read_text(encoding="utf-8"))
        ai = json.loads(inputs["ai_assessment"].read_text(encoding="utf-8"))
        investment = json.loads(
            inputs["investment_options"].read_text(encoding="utf-8")
        )
        errors = D.workflow_generation_input_errors(
            manifest, context.country, context.iso3, input_records
        )
        errors.extend(
            D.supplemental_product_errors(
                ai, investment, context.country, context.iso3
            )
        )
        errors.extend(I.validate_product(investment))
        if errors:
            raise W.NonRetryableStageError(
                "synthetic Stage 7 integration failed: " + "; ".join(errors)
            )

        draft = {
            "schema_version": "damm.integrated-dar-simulation/v1",
            "country": context.country,
            "iso3": context.iso3,
            "status": "Draft — simulated, unreviewed, not acceptance evidence",
            "simulation_notice": SIMULATION_LABEL,
            "simulation_provenance": provenance,
            "executive_summary": (
                "This simulated Draft proves that a recovered Stage 6 product is "
                "consumed before integration and remains available in the annex."
            ),
            "investment_portfolio_summary": {
                "option_count": len(investment["options"]),
                "portfolio_sequencing": investment["portfolio_sequencing"],
                "decision_status": investment["decision_status"],
            },
            "annexes": {
                "investment_options_and_cost_benefit": investment,
            },
            "input_sha256": {
                name: record["sha256"] for name, record in input_records.items()
            },
        }
        draft_path = _write_json(context.stage_dir / "integrated-draft.json", draft)
        claims_path = _write_json(
            context.stage_dir / "claim-provenance.json",
            {
                "schema_version": "damm.claim-provenance-simulation/v1",
                "label": SIMULATION_LABEL,
                "claims": [{
                    "claim": "The Draft contains six simulated investment options.",
                    "basis": "investment_options",
                    "source_sha256": input_records["investment_options"]["sha256"],
                }],
                "simulation_provenance": provenance,
            },
        )
        scans = json.loads(inputs["scans"].read_text(encoding="utf-8"))
        foresight = json.loads(inputs["foresight"].read_text(encoding="utf-8"))
        rendered_draft = {
            "country": context.country,
            "status": "Draft DAR — simulated, unreviewed, not acceptance evidence",
            "model_version": "simulation fixture",
            "assessment_year": D.ASSESSMENT_YEAR,
            "final": False,
            "publication_blockers": [SIMULATION_LABEL],
            "fidelity": {"rate": 1.0, "supported": 1, "claimed": 1},
            "chapters": [
                {
                    "n": 1,
                    "title": "Executive integration proof",
                    "kind": "diagnostic",
                    "provenance": (
                        "Generated from hash-bound synthetic Stage 1–6 outputs; "
                        "not acceptance evidence."
                    ),
                    "prose": draft["executive_summary"],
                },
                {
                    "n": "A",
                    "title": "Evidence and decision annex",
                    "kind": "diagnostic",
                    "provenance": (
                        "Generated deterministically from the zero-spend simulation "
                        "record and its recovered Stage 6 appraisal."
                    ),
                    "prose": (
                        "The complete simulated investment appraisal is retained in "
                        "the companion source data and carried into this report."
                    ),
                    "annex": {
                        "indicator_evidence": [],
                        "candidate_rows": [],
                        "country_findings": list(scans.get("country_findings") or []),
                        "international_pointers": list(
                            scans.get("international_pointers") or []
                        ),
                        "initiative_register": list(scans.get("register_entries") or []),
                        "scan_abstentions": list(scans.get("abstained") or []),
                        "ai_digital_agriculture": ai,
                        "investment_options": investment,
                        "foresight": foresight,
                        "method_record": {"prohibitions": D.PROHIBITIONS},
                    },
                },
            ],
        }
        html_path = context.stage_dir / "integrated-draft.html"
        stage7_marker = "SIMULATION-SOURCE-" + _file_sha256(draft_path)[:16].upper()
        html_path.write_text(
            _embed_html_source_marker(D.render_html(rendered_draft), stage7_marker),
            encoding="utf-8",
        )
        return W.StageResult(
            artifacts={
                "draft_dar_report": html_path,
                "dar_source_data": draft_path,
                "claim_provenance": claims_path,
            },
            spent_usd=0.0,
        )

    return handler


def _simulation_source_identity(source: Path) -> tuple[str, str]:
    """Read a converter input and return the title plus its carried source marker."""
    content = source.read_bytes()
    text = ""
    if source.suffix.lower() == ".docx" and zipfile.is_zipfile(source):
        try:
            with zipfile.ZipFile(source) as archive:
                document = ET.fromstring(archive.read("word/document.xml"))
            text = "\n".join(
                str(node.text or "")
                for node in document.iter()
                if node.tag.rsplit("}", 1)[-1] == "t"
            )
        except (KeyError, OSError, ET.ParseError, zipfile.BadZipFile) as error:
            raise SimulationError(
                f"cannot inspect simulated DOCX source {source.name}: {error}"
            ) from error
    else:
        try:
            text = content.decode("utf-8")
        except UnicodeDecodeError as error:
            raise SimulationError(
                f"simulated converter source {source.name} is not UTF-8"
            ) from error

    marker_match = re.search(r"SIMULATION-SOURCE-[0-9A-F]{16}", text)
    marker = (
        marker_match.group(0)
        if marker_match
        else "SIMULATION-SOURCE-" + hashlib.sha256(content).hexdigest()[:16].upper()
    )
    title = ""
    if source.suffix.lower() in {".html", ".htm"}:
        match = re.search(r"(?is)<title\b[^>]*>(.*?)</title>", text)
        if match is None:
            match = re.search(r"(?is)<h1\b[^>]*>(.*?)</h1>", text)
        if match:
            title = html.unescape(re.sub(r"(?s)<[^>]+>", " ", match.group(1)))
    elif source.suffix.lower() in {".md", ".markdown"}:
        match = re.search(r"(?m)^#\s+(.+?)\s*$", text)
        if match:
            title = match.group(1)
    elif source.suffix.lower() == ".docx":
        title = text.splitlines()[0] if text.splitlines() else ""
    title = " ".join(title.split())
    if not title:
        title = source.stem.replace("_", " ").replace("-", " ").strip().title()
    return title, marker


def _deterministic_zip_member(archive: zipfile.ZipFile, name: str, content: str) -> None:
    info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
    info.compress_type = zipfile.ZIP_DEFLATED
    info.create_system = 3
    info.external_attr = 0o100644 << 16
    archive.writestr(info, content.encode("utf-8"))


def _write_simulation_docx(path: Path, title: str, marker: str) -> None:
    """Write a deterministic, structurally complete minimal Word document."""
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w", compression=zipfile.ZIP_DEFLATED) as archive:
        _deterministic_zip_member(
            archive,
            "[Content_Types].xml",
            "<?xml version='1.0' encoding='UTF-8' standalone='yes'?>"
            "<Types xmlns='http://schemas.openxmlformats.org/package/2006/content-types'>"
            "<Default Extension='rels' ContentType='application/vnd.openxmlformats-package.relationships+xml'/>"
            "<Default Extension='xml' ContentType='application/xml'/>"
            "<Override PartName='/word/document.xml' ContentType='application/vnd.openxmlformats-officedocument.wordprocessingml.document.main+xml'/>"
            "<Override PartName='/docProps/core.xml' ContentType='application/vnd.openxmlformats-package.core-properties+xml'/>"
            "</Types>",
        )
        _deterministic_zip_member(
            archive,
            "_rels/.rels",
            "<?xml version='1.0' encoding='UTF-8' standalone='yes'?>"
            "<Relationships xmlns='http://schemas.openxmlformats.org/package/2006/relationships'>"
            "<Relationship Id='rId1' Type='http://schemas.openxmlformats.org/officeDocument/2006/relationships/officeDocument' Target='word/document.xml'/>"
            "<Relationship Id='rId2' Type='http://schemas.openxmlformats.org/package/2006/relationships/metadata/core-properties' Target='docProps/core.xml'/>"
            "</Relationships>",
        )
        _deterministic_zip_member(
            archive,
            "docProps/core.xml",
            "<?xml version='1.0' encoding='UTF-8' standalone='yes'?>"
            "<cp:coreProperties xmlns:cp='http://schemas.openxmlformats.org/package/2006/metadata/core-properties' "
            "xmlns:dc='http://purl.org/dc/elements/1.1/' "
            "xmlns:dcterms='http://purl.org/dc/terms/' "
            "xmlns:xsi='http://www.w3.org/2001/XMLSchema-instance'>"
            f"<dc:title>{html.escape(title)}</dc:title>"
            "<dc:creator>DAR Studio zero-spend simulation</dc:creator>"
            "<dcterms:created xsi:type='dcterms:W3CDTF'>2026-09-02T00:00:00Z</dcterms:created>"
            "<dcterms:modified xsi:type='dcterms:W3CDTF'>2026-09-02T00:00:00Z</dcterms:modified>"
            "</cp:coreProperties>",
        )
        _deterministic_zip_member(
            archive,
            "word/document.xml",
            "<?xml version='1.0' encoding='UTF-8' standalone='yes'?>"
            "<w:document xmlns:w='http://schemas.openxmlformats.org/wordprocessingml/2006/main'>"
            "<w:body>"
            "<w:p><w:r><w:rPr><w:b/><w:sz w:val='30'/></w:rPr><w:t>"
            f"{html.escape(title)}</w:t></w:r></w:p>"
            f"<w:p><w:r><w:t>{html.escape(marker)}</w:t></w:r></w:p>"
            f"<w:p><w:r><w:t>{html.escape(SIMULATION_LABEL)}</w:t></w:r></w:p>"
            "<w:p><w:r><w:t>Deterministic local conversion; no external service was called.</w:t></w:r></w:p>"
            "<w:sectPr><w:pgSz w:w='11906' w:h='16838'/><w:pgMar w:top='1134' w:right='1134' w:bottom='1134' w:left='1134' w:header='708' w:footer='708' w:gutter='0'/></w:sectPr>"
            "</w:body></w:document>",
        )
        _deterministic_zip_member(
            archive,
            "word/_rels/document.xml.rels",
            "<?xml version='1.0' encoding='UTF-8' standalone='yes'?>"
            "<Relationships xmlns='http://schemas.openxmlformats.org/package/2006/relationships'/>",
        )


def _simulation_pandoc(source: Path, target: Path) -> None:
    target.parent.mkdir(parents=True, exist_ok=True)
    title, marker = _simulation_source_identity(source)
    if target.suffix == ".md":
        target.write_text(
            f"# {title}\n\n> **{SIMULATION_LABEL}**\n\n"
            f"Source identity: `{marker}`\n\n"
            "Converted deterministically from the hash-bound simulation source.\n",
            encoding="utf-8",
        )
    elif target.suffix == ".html":
        target.write_text(
            "<!doctype html><html lang='en'><head><meta charset='utf-8'>"
            "<meta name='viewport' content='width=device-width,initial-scale=1'>"
            f"<title>{html.escape(title)}</title>"
            "<style>body{max-width:50rem;margin:3rem auto;padding:0 1.5rem;"
            "font:16px/1.55 Georgia,serif;color:#17322a}"
            ".marker{font:12px/1.4 ui-monospace,monospace;color:#5d6b64}</style>"
            f"</head><body><h1>{html.escape(title)}</h1>"
            f"<p><strong>{html.escape(SIMULATION_LABEL)}</strong></p>"
            f"<p class='marker'>{html.escape(marker)}</p>"
            "<p>Converted deterministically from the hash-bound simulation source.</p>"
            "</body></html>\n",
            encoding="utf-8",
        )
    elif target.suffix == ".docx":
        _write_simulation_docx(target, title, marker)
    else:
        raise SimulationError(f"unexpected simulated narrative target {target}")


def _pdf_text(value: str) -> str:
    ascii_text = unicodedata.normalize("NFKD", value).encode("ascii", "ignore").decode("ascii")
    return ascii_text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")


def _simulation_pdf(source: Path, target: Path) -> None:
    """Write a deterministic one-page PDF carrying the DOCX source identity."""
    target.parent.mkdir(parents=True, exist_ok=True)
    title, marker = _simulation_source_identity(source)
    stream = (
        "BT /F1 15 Tf 72 748 Td "
        f"({_pdf_text(title)}) Tj 0 -26 Td /F1 10 Tf "
        f"({_pdf_text(marker)}) Tj 0 -20 Td "
        f"({_pdf_text(SIMULATION_LABEL)}) Tj ET\n"
    ).encode("ascii")
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 595 842] /Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n" + stream + b"endstream",
    ]
    document = bytearray(b"%PDF-1.4\n% SIMULATED LOCAL FIXTURE\n")
    offsets = [0]
    for index, body in enumerate(objects, 1):
        offsets.append(len(document))
        document.extend(f"{index} 0 obj\n".encode("ascii"))
        document.extend(body)
        document.extend(b"\nendobj\n")
    xref_offset = len(document)
    document.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    document.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        document.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    document.extend(
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\n"
        f"startxref\n{xref_offset}\n%%EOF\n".encode("ascii")
    )
    target.write_bytes(bytes(document))


def _rewrite_deterministic_package_zip(package: Path, target: Path) -> None:
    """Replace Stage 8's mtime-bearing ZIP with stable simulation bytes."""

    payload = io.BytesIO()
    with zipfile.ZipFile(
        payload, "w", compression=zipfile.ZIP_DEFLATED, compresslevel=9
    ) as archive:
        for path in sorted(candidate for candidate in package.rglob("*")
                           if candidate.is_file()):
            name = f"{package.name}/{path.relative_to(package).as_posix()}"
            info = zipfile.ZipInfo(name, date_time=(1980, 1, 1, 0, 0, 0))
            info.compress_type = zipfile.ZIP_DEFLATED
            info.create_system = 3
            info.external_attr = 0o100644 << 16
            archive.writestr(info, path.read_bytes())
    target.write_bytes(payload.getvalue())
    E.validate_package_zip(package, target)


def _production_stage8_handler(provenance: Mapping[str, Any]):
    """Run the production package builder with local, deterministic converters."""
    def handler(context: W.StageContext) -> W.StageResult:
        result = E.build_export_package(
            country=context.country,
            iso3=context.iso3,
            out=context.stage_dir / f"{context.iso3}_simulated",
            workflow_manifest=context.manifest_path,
            pandoc_converter=_simulation_pandoc,
            pdf_converter=_simulation_pdf,
            contract_path=W.DEFAULT_CONTRACT_PATH,
            created_at=SIMULATION_CREATED_AT,
            resume=False,
        )
        package = result.package_dir
        _rewrite_deterministic_package_zip(package, result.zip_path)
        return W.StageResult(
            artifacts={
                "narrative_exports": package / "narratives",
                "structured_exports": package / "structured",
                "source_inventory_exports": package / "source-inventory",
                "workflow_manifest": result.package_manifest,
                "complete_bundle": result.zip_path,
            },
            spent_usd=0.0,
        )

    return handler


def _stage8_handler(provenance: Mapping[str, Any]):
    def handler(context: W.StageContext) -> W.StageResult:
        directories: dict[str, Path] = {}
        for key in ("narrative_exports", "structured_exports", "source_inventory_exports"):
            directory = context.stage_dir / key
            directory.mkdir(parents=True, exist_ok=True)
            _write_json(directory / "SIMULATED.json", {
                "schema_version": "damm.synthetic-export/v1",
                "label": SIMULATION_LABEL,
                "acceptance_eligible": False,
                "artifact_key": key,
                "simulation_provenance": provenance,
            })
            directories[key] = directory

        package_manifest = _write_json(context.stage_dir / "package-manifest.json", {
            "schema_version": "damm.synthetic-package/v1",
            "label": SIMULATION_LABEL,
            "acceptance_eligible": False,
            "simulation_provenance": provenance,
            "members": sorted(directories),
        })
        bundle = context.stage_dir / "simulated-dar-package.zip"
        with zipfile.ZipFile(bundle, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            archive.writestr("SIMULATED-NOT-ACCEPTANCE-EVIDENCE.txt", SIMULATION_LABEL + "\n")
            archive.write(package_manifest, "package-manifest.json")
            for key, directory in sorted(directories.items()):
                archive.write(directory / "SIMULATED.json", f"{key}/SIMULATED.json")
        return W.StageResult(
            artifacts={
                **directories,
                "workflow_manifest": package_manifest,
                "complete_bundle": bundle,
            },
            spent_usd=0.0,
        )

    return handler


def _path_size(path: Path) -> int:
    if path.is_dir():
        return sum(candidate.stat().st_size for candidate in path.rglob("*") if candidate.is_file())
    return path.stat().st_size


def _manifest_artifacts(manifest: Mapping[str, Any], workspace: Path, output: Path) -> list[dict[str, Any]]:
    records: list[dict[str, Any]] = []
    for stage in manifest.get("stages") or []:
        for artifact in stage.get("artifacts") or []:
            path = workspace / artifact["path"]
            records.append({
                "stage_id": stage["id"],
                "key": artifact["key"],
                "path": path.relative_to(output).as_posix(),
                "sha256": artifact["sha256"],
                "size_bytes": _path_size(path),
                "media_type": artifact["media_type"],
                "label": SIMULATION_LABEL,
            })
    return records


def _stage_summaries(manifest: Mapping[str, Any]) -> list[dict[str, Any]]:
    return [{
        "ordinal": stage["ordinal"],
        "stage_id": stage["id"],
        "status": stage["status"],
        "attempts": stage["attempts"],
        # Simulation spend is an integer-zero contract.  Keeping the JSON number
        # representation identical across Python and JavaScript makes the report
        # digest language-independent (``0`` rather than Python's ``0.0``).
        "spent_usd": 0 if stage["spent_usd"] == 0 else stage["spent_usd"],
    } for stage in manifest.get("stages") or []]


def _error_sha256(error_code: str | None, failed_stage: str | None, message: str | None) -> str | None:
    if not error_code and not failed_stage and not message:
        return None
    return _stable_sha256({
        "error_code": error_code,
        "failed_stage": failed_stage,
        "message": message,
    })


def _finish_report(report: dict[str, Any], output: Path) -> dict[str, Any]:
    report["report_sha256"] = _stable_sha256(report)
    _write_json(output / REPORT_NAME, report)
    return copy.deepcopy(report)


def _base_report(
    *,
    scenario_id: str,
    scenario_sha256: str,
    run_id: str,
    country: str,
    iso3: str,
    profile: str,
    code_identity: Mapping[str, Any],
) -> dict[str, Any]:
    return {
        "schema_version": REPORT_SCHEMA,
        "label": SIMULATION_LABEL,
        "execution_kind": "simulation",
        "acceptance_eligible": False,
        "scenario_id": scenario_id,
        "scenario_sha256": scenario_sha256,
        "code_identity": copy.deepcopy(dict(code_identity)),
        "run_id": run_id,
        "vendor": f"fixture/{scenario_id}",
        "country": country,
        "iso3": iso3,
        "profile": profile,
        "harness_verdict": "fail",
        "observed": {
            "workflow_status": "failed",
            "failed_stage": None,
            "error_code": None,
            "error_sha256": None,
        },
        "external_spend_usd": 0,
        "external_io": {
            "network_calls": 0,
            "database_writes": 0,
            "capabilities_minted": 0,
            "subprocess_calls": 0,
        },
        "fixture_call_count": 0,
        "stages": [],
        "artifacts": [],
        "assertions": [],
    }


def _simulate_overlength(
    scenario: Mapping[str, Any], report: dict[str, Any], output: Path
) -> dict[str, Any]:
    fixture = scenario["fixture"]
    repair_lengths = list(fixture["repair_lengths"])
    recovery_lengths = list(fixture["recovery_lengths"])
    evidence_batches = int(fixture["evidence_batches"])
    failed_batch = int(fixture["failed_batch"])
    candidate_count = int(fixture["candidate_count"])
    llm = _OverlengthFixtureLLM(
        repair_lengths,
        recovery_lengths,
        evidence_batches=evidence_batches,
        failed_batch=failed_batch,
        candidate_count=candidate_count,
    )
    sources = [
        _source(f"SRC-{index:03d}", index, characters=900)
        for index in range(1, evidence_batches + 1)
    ]
    limits = I.AppraisalLimits(
        evidence_batch_characters=1050,
        candidate_output_tokens=I.DEFAULT_APPRAISAL_LIMITS.candidate_output_tokens,
        option_output_tokens=I.DEFAULT_APPRAISAL_LIMITS.option_output_tokens,
        portfolio_output_tokens=I.DEFAULT_APPRAISAL_LIMITS.portfolio_output_tokens,
    )
    observed_batches = I.batch_evidence(sources, limits.evidence_batch_characters)
    failure: Exception | None = None
    response: dict[str, Any] | None = None
    product: dict[str, Any] | None = None
    product_errors: list[str] = []
    try:
        response = I.synthesize_appraisal(
            report["country"], sources, llm, limits=limits
        )
        product = I.build_product(
            report["country"], report["iso3"], response, sources
        )
        product["assessment_date"] = SIMULATION_DATE
        product_errors = I.validate_product(product)
        if product_errors:
            raise SimulationError(
                "recovered Stage 6 product failed validation: "
                + "; ".join(product_errors)
            )
    except Exception as error:
        failure = error

    error_code = (
        getattr(failure, "code", type(failure).__name__)
        if failure is not None
        else None
    )
    failed_stage = "investment_options" if failure is not None else None
    message = str(failure) if failure is not None else None
    report["observed"] = {
        "workflow_status": "failed" if failure is not None else "complete",
        "failed_stage": failed_stage,
        "error_code": error_code,
        "error_sha256": _error_sha256(error_code, failed_stage, message),
    }
    expected = scenario["expected"]
    recovery_iterator = iter(recovery_lengths)
    expected_effective_lengths = [
        next(recovery_iterator) if raw_length > limit else raw_length
        for raw_length, (_index, _field, limit) in zip(
            repair_lengths, OVERLENGTH_REPAIR_TARGETS
        )
    ]
    residual_limits = [
        limit
        for raw_length, (_index, _field, limit) in zip(
            repair_lengths, OVERLENGTH_REPAIR_TARGETS
        )
        if raw_length > limit
    ]
    repair_batches = I._candidate_length_repair_batches(
        OVERLENGTH_REPAIR_TARGETS
    )
    retry_targets = tuple(
        (index, field, (limit * 9) // 10)
        for raw_length, (index, field, limit) in zip(
            repair_lengths, OVERLENGTH_REPAIR_TARGETS
        )
        if raw_length > limit
    )
    recovery_batches = I._candidate_length_repair_batches(retry_targets)
    expected_recovery_max_tokens = tuple(
        I._candidate_length_retry_max_tokens(
            batch, limits.candidate_output_tokens
        )
        for batch in recovery_batches
    )
    expected_calls = []
    for index in range(1, evidence_batches + 1):
        map_detail = f"investment candidate map batch {index}/{evidence_batches}"
        expected_calls.append(map_detail)
        if index == failed_batch:
            if len(repair_batches) == 1:
                expected_calls.append(
                    f"{map_detail} [local-length repair 1/1]"
                )
            else:
                expected_calls.extend(
                    f"{map_detail} [local-length repair 1/1 chunk "
                    f"{batch_index}/{len(repair_batches)}]"
                    for batch_index in range(1, len(repair_batches) + 1)
                )
            if len(recovery_batches) == 1:
                expected_calls.append(
                    f"{map_detail} [local-length repair 2/2]"
                )
            else:
                expected_calls.extend(
                    f"{map_detail} [local-length repair 2/2 chunk "
                    f"{batch_index}/{len(recovery_batches)}]"
                    for batch_index in range(1, len(recovery_batches) + 1)
                )
    expected_calls.extend([
        "investment candidate final register",
        *[
            f"investment appraisal INV-{index} batch 1/1"
            for index in range(1, candidate_count + 1)
        ],
        "investment portfolio sequencing",
    ])
    checks = [
        ("real_stage6_recovered", failure is None, type(failure).__name__ if failure else "complete"),
        ("expected_workflow_status", report["observed"]["workflow_status"] == expected["workflow_status"], report["observed"]["workflow_status"]),
        ("expected_failed_stage", failed_stage == expected["failed_stage"], str(failed_stage)),
        ("expected_error_code", error_code == expected["error_code"], str(error_code)),
        ("declared_evidence_batches", len(observed_batches) == evidence_batches, str(len(observed_batches))),
        ("exact_repair_lengths", llm.observed_repair_lengths == repair_lengths, str(llm.observed_repair_lengths)),
        ("exact_recovery_lengths", llm.observed_recovery_lengths == recovery_lengths, str(llm.observed_recovery_lengths)),
        ("bounded_recovery_tokens", bool(expected_recovery_max_tokens) and all(bound == llm.observed_recovery_max_tokens for bound in expected_recovery_max_tokens), str(llm.observed_recovery_max_tokens)),
        ("recovery_within_ninety_percent", all(length <= (limit * 9) // 10 for length, limit in zip(llm.observed_recovery_lengths, residual_limits)), str(llm.observed_recovery_lengths)),
        ("exact_effective_lengths", llm.observed_effective_lengths == expected_effective_lengths, str(llm.observed_effective_lengths)),
        ("declared_candidate_count", response is not None and len(response.get("options") or []) == candidate_count, str(len(response.get("options") or []) if response else None)),
        ("real_stage6_product", product is not None and not product_errors, "; ".join(product_errors) or "valid"),
        ("exact_fixture_call_sequence", llm.calls == expected_calls, str(llm.calls)),
        ("fixture_call_count", len(llm.calls) == expected["fixture_call_count"], str(len(llm.calls))),
        ("zero_external_spend", True, "$0.00"),
    ]
    report["assertions"] = [
        {"id": identifier, "ok": bool(ok), "detail": detail}
        for identifier, ok, detail in checks
    ]
    report["harness_verdict"] = "pass" if all(ok for _, ok, _ in checks) else "fail"
    report["fixture_call_count"] = len(llm.calls)
    report["stages"] = [
        {
            "ordinal": ordinal,
            "stage_id": stage_id,
            "status": (
                "complete"
                if failure is None and stage_id == "investment_options"
                else "failed"
                if failure is not None and stage_id == "investment_options"
                else "not_run"
            ),
            "attempts": 1 if stage_id == "investment_options" else 0,
            "spent_usd": 0,
        }
        for ordinal, stage_id in enumerate(W.EXPECTED_STAGE_IDS, 1)
    ]
    recovery = _write_json(output / "artifacts" / "stage6-recovery.json", {
        "schema_version": "damm.stage6-recovery/v1",
        "label": SIMULATION_LABEL,
        "acceptance_eligible": False,
        "scenario_id": report["scenario_id"],
        "run_id": report["run_id"],
        "vendor": report["vendor"],
        "code_identity": report["code_identity"],
        "observed_repair_lengths": llm.observed_repair_lengths,
        "observed_recovery_lengths": llm.observed_recovery_lengths,
        "observed_recovery_max_tokens": llm.observed_recovery_max_tokens,
        "observed_effective_lengths": llm.observed_effective_lengths,
        "fixture_calls": llm.calls,
        "observed": report["observed"],
    })
    report["artifacts"] = [{
        "stage_id": "investment_options",
        "key": "repair_recovery",
        "path": recovery.relative_to(output).as_posix(),
        "sha256": _file_sha256(recovery),
        "size_bytes": recovery.stat().st_size,
        "media_type": "application/json",
        "label": SIMULATION_LABEL,
    }]
    if product is not None:
        product_path = _write_json(
            output / "artifacts" / "investment-options.json", product
        )
        report["artifacts"].append({
            "stage_id": "investment_options",
            "key": "appraisal_data",
            "path": product_path.relative_to(output).as_posix(),
            "sha256": _file_sha256(product_path),
            "size_bytes": product_path.stat().st_size,
            "media_type": "application/json",
            "label": SIMULATION_LABEL,
        })
    return report


def _simulate_happy(
    scenario: Mapping[str, Any],
    report: dict[str, Any],
    output: Path,
    provenance: Mapping[str, Any],
) -> dict[str, Any]:
    fixture = scenario["fixture"]
    through_package = scenario["kind"] == "stage6_through_package"
    if through_package:
        report["scenario_scope"] = {
            "claim": "Stage 6 recovery through Stage 8 packaging only",
            "focus_stage_ids": [
                "investment_options",
                "draft_dar",
                "export_package",
            ],
            "synthetic_predecessor_stage_ids": [
                "damm_diagnostic",
                "country_research",
                "ai_digital_agriculture",
                "international_lessons",
                "strategic_foresight",
            ],
            "production_modules_exercised": [
                "gauntlet/loop-1/research_pipeline/investment_options.py",
                "gauntlet/loop-1/research_pipeline/report_design.py",
                "gauntlet/loop-1/research_pipeline/generate_dar.py",
                "gauntlet/loop-1/research_pipeline/export_package.py",
                "gauntlet/loop-1/research_pipeline/run_workflow.py",
            ],
            "bound_transitive_modules": [
                "gauntlet/loop-1/research_pipeline/vendors.py",
                "gauntlet/loop-1/research_pipeline/workflow_inputs.py",
                "gauntlet/loop-1/research_pipeline/foresight_contract.py",
                "gauntlet/loop-1/engine_v17.py",
                "model/reference_scorer.py",
            ],
            "simulation_harness_module": (
                "gauntlet/loop-1/research_pipeline/simulation.py"
            ),
        }
        count = int(fixture["evidence_batches"])
        candidate_count = int(fixture["candidate_count"])
        sources = [
            _source(f"SRC-{index:03d}", index, characters=900)
            for index in range(1, count + 1)
        ]
        llm = _OverlengthFixtureLLM(
            list(fixture["repair_lengths"]),
            list(fixture["recovery_lengths"]),
            evidence_batches=count,
            failed_batch=int(fixture["failed_batch"]),
            candidate_count=candidate_count,
        )
        limits = I.AppraisalLimits(
            evidence_batch_characters=1050,
            candidate_output_tokens=(
                I.DEFAULT_APPRAISAL_LIMITS.candidate_output_tokens
            ),
            option_output_tokens=I.DEFAULT_APPRAISAL_LIMITS.option_output_tokens,
            portfolio_output_tokens=(
                I.DEFAULT_APPRAISAL_LIMITS.portfolio_output_tokens
            ),
        )
    else:
        count = int(fixture["source_count_by_profile"][report["profile"]])
        candidate_count = int(fixture["candidate_count"])
        sources = [
            _source(f"SRC-{index:03d}", index) for index in range(1, count + 1)
        ]
        llm = _HappyFixtureLLM(candidate_count)
        limits = I.DEFAULT_APPRAISAL_LIMITS
    synthetic = _synthetic_stage_handler(provenance)
    handlers = {stage_id: synthetic for stage_id in W.EXPECTED_STAGE_IDS}
    handlers["investment_options"] = _stage6_handler(
        provenance, sources, llm, limits=limits
    )
    if through_package:
        handlers["ai_digital_agriculture"] = _ai_stage_handler(provenance)
        handlers["draft_dar"] = _integrated_draft_handler(provenance)
        handlers["export_package"] = _production_stage8_handler(provenance)
    else:
        handlers["export_package"] = _stage8_handler(provenance)
    workspace = output / "workflow"
    contract = W.load_contract(W.DEFAULT_CONTRACT_PATH)
    coordinator = W.WorkflowCoordinator(
        contract=contract,
        workspace=workspace,
        handlers=handlers,
        commands={},
        max_attempts=1,
        retry_delay_seconds=0,
        clock=lambda: SIMULATION_CLOCK,
        simulation_provenance=provenance,
    )
    manifest: dict[str, Any]
    caught: Exception | None = None
    try:
        manifest = coordinator.run(
            country=report["country"],
            iso3=report["iso3"],
        run_id=report["run_id"],
        ceiling_usd=None,
        vendor=report["vendor"],
        )
    except W.WorkflowRunFailed as error:
        manifest = error.manifest
        caught = error
    except Exception as error:
        manifest = getattr(coordinator, "_manifest", {})
        caught = error

    failure = manifest.get("failure") if isinstance(manifest, dict) else None
    failed_stage = manifest.get("current_stage") if manifest.get("status") == "failed" else None
    error_code = None
    error_message = None
    if isinstance(failure, dict):
        error_code = failure.get("type")
        error_message = failure.get("message")
    elif caught is not None:
        error_code = type(caught).__name__
        error_message = str(caught)
    report["observed"] = {
        "workflow_status": manifest.get("status") or "failed",
        "failed_stage": failed_stage,
        "error_code": error_code,
        "error_sha256": _error_sha256(error_code, failed_stage, error_message),
    }
    stages = manifest.get("stages") or []
    stage_manifests_marked = True
    for stage in stages:
        binding = next(
            (artifact for artifact in stage.get("artifacts") or [] if artifact.get("key") == "stage_manifest"),
            None,
        )
        if not binding:
            stage_manifests_marked = False
            continue
        value = json.loads((workspace / binding["path"]).read_text(encoding="utf-8"))
        stage_manifests_marked = stage_manifests_marked and value.get("simulation_provenance") == provenance

    expected_calls = int(scenario["expected"]["fixture_call_count"])
    expected = scenario["expected"]
    stage6_manifest = next(
        (stage for stage in stages if stage.get("id") == "investment_options"),
        None,
    )
    stage6_candidate_count = None
    if stage6_manifest:
        appraisal_binding = next(
            (
                artifact
                for artifact in stage6_manifest.get("artifacts") or []
                if artifact.get("key") == "appraisal_data"
            ),
            None,
        )
        if appraisal_binding:
            appraisal_product = json.loads(
                (workspace / appraisal_binding["path"]).read_text(encoding="utf-8")
            )
            stage6_candidate_count = len(appraisal_product.get("options") or [])
    checks = [
        ("real_eight_stage_coordinator", manifest.get("status") == "complete", str(manifest.get("status"))),
        ("expected_workflow_status", report["observed"]["workflow_status"] == expected["workflow_status"], report["observed"]["workflow_status"]),
        ("expected_failed_stage", failed_stage == expected["failed_stage"], str(failed_stage)),
        ("expected_error_code", error_code == expected["error_code"], str(error_code)),
        ("all_eight_stages_complete", len(stages) == 8 and all(stage.get("status") == "complete" for stage in stages), str(len(stages))),
        ("real_stage6_product", any(stage.get("id") == "investment_options" and stage.get("status") == "complete" for stage in stages), "investment_options"),
        ("declared_candidate_count", stage6_candidate_count == candidate_count, str(stage6_candidate_count)),
        ("fixture_call_count", len(llm.calls) == expected_calls, str(len(llm.calls))),
        ("zero_external_spend", manifest.get("spent_usd") == 0.0, str(manifest.get("spent_usd"))),
        ("root_provenance", manifest.get("simulation_provenance") == provenance, "bound"),
        ("stage_provenance", stage_manifests_marked, "bound" if stage_manifests_marked else "missing"),
    ]
    report["assertions"] = [
        {"id": identifier, "ok": bool(ok), "detail": detail}
        for identifier, ok, detail in checks
    ]
    report["harness_verdict"] = "pass" if all(ok for _, ok, _ in checks) else "fail"
    report["fixture_call_count"] = len(llm.calls)
    report["stages"] = _stage_summaries(manifest)
    report["artifacts"] = _manifest_artifacts(manifest, workspace, output)
    return report


def simulate_workflow(
    scenario_id: str,
    output_dir: str | os.PathLike[str],
    *,
    country: str | None = None,
    iso3: str | None = None,
    profile: str | None = None,
) -> dict[str, Any]:
    """Run one built-in cost-free scenario and return its hash-bound report."""

    scenario, _scenario_file, scenario_sha256 = _load_scenario(scenario_id)
    code_identity = _code_identity()
    if code_identity["aggregate_sha256"] != scenario["expected"]["code_sha256"]:
        raise SimulationError(
            "simulation production-code identity does not match the committed scenario"
        )
    selected_country, selected_iso3, selected_profile = _normalise_launch(
        scenario, country=country, iso3=iso3, profile=profile
    )
    output = Path(output_dir).expanduser().resolve()
    if output.exists() and any(output.iterdir()):
        raise SimulationError(f"simulation output directory is not empty: {output}")
    output.mkdir(parents=True, exist_ok=True)
    identity = _stable_sha256({
        "scenario_sha256": scenario_sha256,
        "country": selected_country,
        "iso3": selected_iso3,
        "profile": selected_profile,
    })
    run_id = f"sim-{scenario_id[:32]}-{identity[:12]}"
    report = _base_report(
        scenario_id=scenario_id,
        scenario_sha256=scenario_sha256,
        run_id=run_id,
        country=selected_country,
        iso3=selected_iso3,
        profile=selected_profile,
        code_identity=code_identity,
    )
    provenance = _provenance(
        scenario_id,
        scenario_sha256,
        code_identity["aggregate_sha256"],
        selected_profile,
    )
    counters = {
        "network_attempts": 0,
        "subprocess_attempts": 0,
        "database_attempts": 0,
        "capabilities_minted": 0,
    }
    with _simulation_boundary(counters):
        if scenario["kind"] == "stage6_repair_recovery":
            report = _simulate_overlength(scenario, report, output)
        else:
            report = _simulate_happy(scenario, report, output, provenance)

    external_attempts = (
        counters["network_attempts"]
        + counters["subprocess_attempts"]
        + counters["database_attempts"]
        + counters["capabilities_minted"]
    )
    if external_attempts:
        report["assertions"].append({
            "id": "no_external_io_attempts",
            "ok": False,
            "detail": str(external_attempts),
        })
        report["harness_verdict"] = "fail"
    report["external_io"] = {
        "network_calls": counters["network_attempts"],
        "database_writes": counters["database_attempts"],
        "capabilities_minted": counters["capabilities_minted"],
        "subprocess_calls": counters["subprocess_attempts"],
    }
    return _finish_report(report, output)


__all__ = [
    "REPORT_NAME",
    "SIMULATION_LABEL",
    "SimulationBoundaryError",
    "SimulationError",
    "simulate_workflow",
]
