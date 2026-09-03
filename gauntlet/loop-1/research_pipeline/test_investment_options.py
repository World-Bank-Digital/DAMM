#!/usr/bin/env python3

import hashlib
import io
import json
import os
import re
import sys
import tempfile
import unittest
import zipfile
from unittest import mock

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)

import investment_options as I


class InvestmentOptionsTest(unittest.TestCase):
    def setUp(self):
        self.price_patcher = mock.patch.dict(I.V.PRICES["anthropic"], {
            "test-model": {"in_per_mtok": 5.0, "out_per_mtok": 25.0},
        })
        self.price_patcher.start()
        self.addCleanup(self.price_patcher.stop)

    @staticmethod
    def truncated(detail, max_tokens):
        return I.V.VendorOutputTruncated(
            vendor="anthropic",
            model="test-model",
            pass_name="investment",
            detail=detail,
            stop_reason="max_tokens",
            request_id="request-not-persisted",
            max_tokens=max_tokens,
            input_tokens=100,
            output_tokens=max_tokens,
            thinking_tokens=max_tokens - 10,
            partial_output_chars=42,
            partial_output_sha256="a" * 64,
        )

    @staticmethod
    def malformed(detail, max_tokens):
        return I.V.VendorMalformedOutput(
            vendor="openai",
            model="test-model",
            pass_name="investment",
            detail=detail,
            stop_reason="completed",
            request_id="request-not-persisted",
            max_tokens=max_tokens,
            input_tokens=100,
            output_tokens=max_tokens - 1,
            thinking_tokens=10,
            partial_output_chars=42,
            partial_output_sha256="b" * 64,
            parse_error="JSONDecodeError at character 41",
        )

    @staticmethod
    def rejected(detail, max_tokens):
        return I.V.VendorOutputRejected(
            vendor="gemini",
            model="test-model",
            pass_name="investment",
            detail=detail,
            stop_reason="SAFETY",
            request_id="request-not-persisted",
            max_tokens=max_tokens,
            input_tokens=100,
            output_tokens=12,
            thinking_tokens=2,
            partial_output_chars=0,
            partial_output_sha256=hashlib.sha256(b"").hexdigest(),
        )

    def test_internal_schemas_fit_their_bounded_output_allowances(self):
        mapped = I.CANDIDATE_MAP_SCHEMA["properties"]["candidates"]
        reduced = I.CANDIDATE_REDUCTION_SCHEMA["properties"]["candidates"]
        final = I.CANDIDATE_REGISTER_SCHEMA["properties"]["candidates"]
        compared = I.CANDIDATE_COMPARISON_SCHEMA["properties"]["candidates"]
        self.assertEqual((mapped["minItems"], mapped["maxItems"]), (0, 4))
        self.assertEqual((reduced["minItems"], reduced["maxItems"]), (1, 4))
        self.assertEqual((final["minItems"], final["maxItems"]), (3, 7))
        self.assertEqual((compared["minItems"], compared["maxItems"]), (1, 7))
        candidate = final["items"]
        self.assertLessEqual(
            candidate["properties"]["source_refs"]["maxItems"], 12)
        self.assertLessEqual(candidate["properties"]["problem"]["maxLength"], 500)
        self.assertLessEqual(
            candidate["properties"]["recommendation_rationale"]["maxLength"], 500)

        option = I.OPTION_BODY_SCHEMA["properties"]
        self.assertLessEqual(
            option["benefits"]["properties"]["quantified"]["maxItems"], 3)
        self.assertLessEqual(
            option["benefits"]["properties"]["qualitative"]["maxItems"], 4)
        self.assertLessEqual(option["sensitivity"]["maxItems"], 3)
        for name in (
                "distributional_effects", "climate_effects", "ai_and_data_risks",
                "implementation_risks", "data_gaps"):
            self.assertLessEqual(option[name]["maxItems"], 4)
            self.assertLessEqual(option[name]["items"]["maxLength"], 300)

    def test_investment_state_checkpoint_refuses_a_symlink(self):
        with tempfile.TemporaryDirectory() as directory:
            target = os.path.join(directory, "target.json")
            link = os.path.join(directory, "investment_state.json")
            with open(target, "w", encoding="utf-8") as handle:
                json.dump({"steps": {}}, handle)
            os.symlink(target, link)

            with self.assertRaisesRegex(ValueError, "regular file"):
                I.load_appraisal_state(link)

    def test_reusable_state_requires_its_authoritative_spend_ledger_prefix(self):
        sources = [{
            "ref": "SRC-001", "kind": "country_finding", "title": "Policy",
            "text": "Evidence", "source": "https://example.test/policy",
        }]
        limits = I.DEFAULT_APPRAISAL_LIMITS
        state = I.appraisal_state(
            "Exampleland", "EXP", sources, "anthropic", "test-model", limits)
        state["steps"] = {"candidate-register-0001": {
            "request_sha256": "a" * 64,
            "response_sha256": "b" * 64,
            "response": {"candidates": []},
        }}
        state["spend_prefix"] = {
            "call_count": 1,
            "calls_sha256": "c" * 64,
        }
        empty_ledger = I.V.Ledger(ceiling=500, label="test")

        with self.assertRaisesRegex(ValueError, "spend ledger"):
            I.appraisal_state(
                "Exampleland", "EXP", sources, "anthropic", "test-model", limits,
                loaded=state, ledger=empty_ledger,
            )

    def test_checkpointed_call_retries_one_truncation_with_more_room(self):
        test_case = self

        class TruncatingOnceLLM:
            def __init__(self):
                self.calls = []

            def json_call(
                    self, _system, _user, _schema, _pass_name,
                    max_tokens=8000, detail=""):
                self.calls.append((max_tokens, detail))
                if len(self.calls) == 1:
                    raise test_case.truncated(detail, max_tokens)
                return {"value": "complete"}

        llm = TruncatingOnceLLM()
        state = {"steps": {}}
        persisted = []
        response = I._checkpointed_json_call(
            llm,
            state,
            lambda value: persisted.append(json.loads(json.dumps(value))),
            "unit-step",
            "system",
            "user",
            {"type": "object", "properties": {"value": {"type": "string"}},
             "required": ["value"], "additionalProperties": False},
            100,
            "logical unit",
        )

        self.assertEqual(response, {"value": "complete"})
        self.assertEqual(llm.calls, [
            (100, "logical unit"),
            (200, "logical unit [truncation retry 2/2]"),
        ])
        self.assertEqual(len(persisted), 2)
        step = state["steps"]["unit-step"]
        self.assertEqual(step["status"], "complete")
        self.assertEqual(len(step["truncations"]), 1)
        self.assertNotIn("request_id", json.dumps(step))
        self.assertEqual(
            step["paid_request_sha256"],
            I.V.json_call_request_sha256(
                "system",
                "user",
                {"type": "object", "properties": {"value": {"type": "string"}},
                 "required": ["value"], "additionalProperties": False},
                I.PASS,
                200,
                "logical unit [truncation retry 2/2]",
            ),
        )

    def test_stage6_uses_exactly_two_one_shot_provider_attempts(self):
        ledger = I.V.Ledger(ceiling=500, label="test")
        llm = I.V.LLM("anthropic", ledger, model="test-model")
        attempts = []

        def transport(_system, _user, _schema, max_tokens):
            attempts.append(max_tokens)
            raise I.V._ProviderOutputTruncated(
                stop_reason="max_tokens",
                request_id=f"request-{len(attempts)}",
                max_tokens=max_tokens,
                input_tokens=10,
                output_tokens=max_tokens,
                thinking_tokens=max_tokens - 1,
                partial_output="",
            )

        llm._call_anthropic = transport
        with self.assertRaises(I.AppraisalOutputExhausted):
            I._checkpointed_json_call(
                llm, {"steps": {}}, None, "unit-step", "system", "user",
                {"type": "object"}, 100, "logical unit")

        self.assertEqual(attempts, [100, 200])

    def test_completed_checkpoint_requires_exact_paid_request_digest(self):
        class CompleteLLM:
            def __init__(self):
                self.calls = []

            def json_call(
                    self, _system, _user, _schema, _pass_name,
                    max_tokens=8000, detail=""):
                self.calls.append((max_tokens, detail))
                return {"value": "complete"}

        schema = {
            "type": "object",
            "properties": {"value": {"type": "string"}},
            "required": ["value"],
            "additionalProperties": False,
        }
        state = {"steps": {}}
        I._checkpointed_json_call(
            CompleteLLM(), state, None, "unit-step", "system", "user",
            schema, 100, "logical unit")

        for mutation in ("missing", "tampered"):
            with self.subTest(mutation=mutation):
                corrupted = json.loads(json.dumps(state))
                if mutation == "missing":
                    corrupted["steps"]["unit-step"].pop("paid_request_sha256")
                else:
                    corrupted["steps"]["unit-step"]["paid_request_sha256"] = "f" * 64
                resumed = CompleteLLM()
                with self.assertRaisesRegex(ValueError, "response mismatch"):
                    I._checkpointed_json_call(
                        resumed, corrupted, None, "unit-step", "system", "user",
                        schema, 100, "logical unit")
                self.assertEqual(resumed.calls, [])

    def test_completed_checkpoint_rejects_falsey_non_list_failure_history(self):
        class CompleteLLM:
            def __init__(self):
                self.calls = []

            def json_call(self, *_args, **_kwargs):
                self.calls.append(True)
                return {"value": "complete"}

        schema = {"type": "object"}
        state = {"steps": {}}
        I._checkpointed_json_call(
            CompleteLLM(), state, None, "unit-step", "system", "user",
            schema, 100, "logical unit")
        state["steps"]["unit-step"]["truncations"] = {}

        resumed = CompleteLLM()
        with self.assertRaisesRegex(ValueError, "truncations are invalid"):
            I._checkpointed_json_call(
                resumed, state, None, "unit-step", "system", "user",
                schema, 100, "logical unit")
        self.assertEqual(resumed.calls, [])

    def test_cached_candidate_response_must_match_its_paid_ledger_result(self):
        ledger = I.V.Ledger(ceiling=500, label="test")
        llm = I.V.LLM(
            "anthropic", ledger, model="test-model").enable_durable_outcomes()
        mapped = {"candidates": [{
            "title": f"Option {index}",
            "problem": f"Problem {index}",
            "recommendation_rationale": f"Rationale {index}",
            "source_refs": ["SRC-001"],
        } for index in range(1, 4)]}
        transport = mock.Mock(return_value=(mapped, 10, 20))
        llm._call_anthropic = transport
        state = {"steps": {}}

        I._checkpointed_candidate_call(
            llm,
            state,
            lambda current: I._bind_state_to_spend_prefix(current, ledger),
            "candidate-map-0001",
            "system",
            "user",
            I.CANDIDATE_MAP_SCHEMA,
            100,
            "investment candidate map batch 1/1",
            {"SRC-001"},
        )
        cached = state["steps"]["candidate-map-0001"]
        cached["response"]["candidates"][0]["problem"] = "p" * 501
        cached["response_sha256"] = I.V.stable_json_sha256(cached["response"])
        transport.reset_mock()
        transport.side_effect = AssertionError(
            "tampered state authorized a new paid repair")

        with self.assertRaisesRegex(
                I.AppraisalCheckpointUnsafe, "spend ledger result"):
            I._checkpointed_candidate_call(
                llm,
                state,
                lambda current: I._bind_state_to_spend_prefix(current, ledger),
                "candidate-map-0001",
                "system",
                "user",
                I.CANDIDATE_MAP_SCHEMA,
                100,
                "investment candidate map batch 1/1",
                {"SRC-001"},
            )

        transport.assert_not_called()
        self.assertEqual(len(ledger.calls), 1)

    def test_unclaimed_durable_success_is_recovered_without_provider_replay(self):
        ledger = I.V.Ledger(ceiling=500, label="test")
        llm = I.V.LLM(
            "anthropic", ledger, model="test-model").enable_durable_outcomes()
        schema = {
            "type": "object",
            "properties": {"value": {"type": "string"}},
            "required": ["value"],
            "additionalProperties": False,
        }
        transport = mock.Mock(return_value=({"value": "complete"}, 10, 20))
        llm._call_anthropic = transport

        # Simulate a crash after the atomically journaled provider result but before
        # the Stage 6 state write.
        llm.json_call_once(
            "system", "user", schema, I.PASS,
            max_tokens=100, detail="logical unit")
        self.assertIn("structured_result", ledger.calls[0])
        transport.reset_mock()
        transport.side_effect = AssertionError("paid provider request was replayed")

        state = {"steps": {}}
        response = I._checkpointed_json_call(
            llm,
            state,
            lambda current: I._bind_state_to_spend_prefix(current, ledger),
            "unit-step",
            "system",
            "user",
            schema,
            100,
            "logical unit",
        )

        self.assertEqual(response, {"value": "complete"})
        transport.assert_not_called()
        self.assertEqual(state["spend_prefix"]["call_count"], 1)
        self.assertEqual(state["steps"]["unit-step"]["status"], "complete")

    def test_unclaimed_truncation_resumes_with_only_the_larger_attempt(self):
        ledger = I.V.Ledger(ceiling=500, label="test")
        llm = I.V.LLM(
            "anthropic", ledger, model="test-model").enable_durable_outcomes()
        schema = {"type": "object"}
        attempts = []

        def transport(_system, _user, _schema, max_tokens):
            attempts.append(max_tokens)
            if max_tokens == 100:
                raise I.V._ProviderOutputTruncated(
                    stop_reason="max_tokens",
                    request_id="first",
                    max_tokens=max_tokens,
                    input_tokens=10,
                    output_tokens=max_tokens,
                    thinking_tokens=max_tokens - 1,
                    partial_output="",
                )
            return {"value": "complete"}, 11, 12

        llm._call_anthropic = transport
        with self.assertRaises(I.V.VendorOutputTruncated):
            llm.json_call_once(
                "system", "user", schema, I.PASS,
                max_tokens=100, detail="logical unit")
        self.assertEqual(attempts, [100])

        state = {"steps": {}}
        response = I._checkpointed_json_call(
            llm,
            state,
            lambda current: I._bind_state_to_spend_prefix(current, ledger),
            "unit-step",
            "system",
            "user",
            schema,
            100,
            "logical unit",
        )

        self.assertEqual(response, {"value": "complete"})
        self.assertEqual(attempts, [100, 200])
        self.assertEqual(len(ledger.calls), 2)
        self.assertEqual(state["spend_prefix"]["call_count"], 2)
        self.assertEqual(
            len(state["steps"]["unit-step"]["truncations"]), 1)

    def test_single_attempt_checkpoint_claims_pending_truncation_without_replay(self):
        ledger = I.V.Ledger(ceiling=500, label="test")
        llm = I.V.LLM(
            "anthropic", ledger, model="test-model"
        ).enable_durable_outcomes()
        schema = {"type": "object"}
        attempts = []

        def transport(_system, _user, _schema, max_tokens):
            attempts.append(max_tokens)
            raise I.V._ProviderOutputTruncated(
                stop_reason="max_tokens",
                request_id="single-attempt",
                max_tokens=max_tokens,
                input_tokens=10,
                output_tokens=max_tokens,
                thinking_tokens=max_tokens - 1,
                partial_output="",
            )

        llm._call_anthropic = transport
        with self.assertRaises(I.V.VendorOutputTruncated):
            llm.json_call_once(
                "system", "user", schema, I.PASS,
                max_tokens=100, detail="logical unit",
            )
        self.assertEqual(attempts, [100])

        state = {"steps": {}}
        with self.assertRaises(I.AppraisalOutputExhausted) as raised:
            I._checkpointed_json_call(
                llm,
                state,
                lambda current: I._bind_state_to_spend_prefix(current, ledger),
                "unit-step",
                "system",
                "user",
                schema,
                100,
                "logical unit",
                attempt_limit=1,
            )

        self.assertEqual(raised.exception.attempt_limit, 1)
        self.assertEqual(attempts, [100])
        self.assertEqual(len(ledger.calls), 1)
        self.assertEqual(state["spend_prefix"]["call_count"], 1)
        self.assertEqual(state["steps"]["unit-step"]["status"], "truncated")
        self.assertEqual(
            len(state["steps"]["unit-step"]["truncations"]), 1)

        with self.assertRaises(I.AppraisalOutputExhausted):
            I._checkpointed_json_call(
                llm, state, None, "unit-step", "system", "user", schema,
                100, "logical unit", attempt_limit=1,
            )
        self.assertEqual(attempts, [100])

    def test_checkpointed_call_adapts_once_after_malformed_output(self):
        test_case = self

        class MalformedOnceLLM:
            def __init__(self):
                self.calls = []

            def json_call(
                    self, _system, _user, _schema, _pass_name,
                    max_tokens=8000, detail=""):
                self.calls.append((max_tokens, detail))
                if len(self.calls) == 1:
                    raise test_case.malformed(detail, max_tokens)
                return {"value": "complete"}

        llm = MalformedOnceLLM()
        state = {"steps": {}}
        response = I._checkpointed_json_call(
            llm, state, None, "unit-step", "system", "user",
            {"type": "object"}, 100, "logical unit")

        self.assertEqual(response, {"value": "complete"})
        self.assertEqual([tokens for tokens, _detail in llm.calls], [100, 200])
        self.assertEqual(
            state["steps"]["unit-step"]["truncations"][0]["outcome"],
            "structured_output_malformed",
        )

    def test_exhausted_truncation_checkpoint_is_terminal_without_paid_replay(self):
        test_case = self

        class AlwaysTruncatedLLM:
            def __init__(self):
                self.calls = []

            def json_call(
                    self, _system, _user, _schema, _pass_name,
                    max_tokens=8000, detail=""):
                self.calls.append((max_tokens, detail))
                raise test_case.truncated(detail, max_tokens)

        schema = {
            "type": "object", "properties": {"value": {"type": "string"}},
            "required": ["value"], "additionalProperties": False,
        }
        state = {"steps": {}}
        first = AlwaysTruncatedLLM()
        with self.assertRaises(I.AppraisalOutputExhausted):
            I._checkpointed_json_call(
                first, state, None, "unit-step", "system", "user", schema,
                100, "logical unit")
        self.assertEqual(len(first.calls), 2)
        self.assertEqual(state["steps"]["unit-step"]["status"], "truncated")
        self.assertEqual(len(state["steps"]["unit-step"]["truncations"]), 2)

        resumed = AlwaysTruncatedLLM()
        with self.assertRaises(I.AppraisalOutputExhausted):
            I._checkpointed_json_call(
                resumed, state, None, "unit-step", "system", "user", schema,
                100, "logical unit")
        self.assertEqual(resumed.calls, [])

    def test_rejected_output_is_terminal_and_not_replayed_on_resume(self):
        test_case = self

        class RejectedLLM:
            def __init__(self):
                self.calls = []

            def json_call(
                    self, _system, _user, _schema, _pass_name,
                    max_tokens=8000, detail=""):
                self.calls.append((max_tokens, detail))
                raise test_case.rejected(detail, max_tokens)

        schema = {
            "type": "object", "properties": {"value": {"type": "string"}},
            "required": ["value"], "additionalProperties": False,
        }
        state = {"steps": {}}
        first = RejectedLLM()
        with self.assertRaises(I.AppraisalOutputRejected):
            I._checkpointed_json_call(
                first, state, None, "unit-step", "system", "user", schema,
                100, "logical unit")
        self.assertEqual(first.calls, [(100, "logical unit")])
        self.assertEqual(state["steps"]["unit-step"]["status"], "rejected")
        self.assertNotIn("request-not-persisted", json.dumps(state))

        corrupted = json.loads(json.dumps(state))
        corrupted["steps"]["unit-step"]["truncations"] = False
        rejected_resume = RejectedLLM()
        with self.assertRaisesRegex(ValueError, "truncations are invalid"):
            I._checkpointed_json_call(
                rejected_resume, corrupted, None, "unit-step", "system", "user",
                schema, 100, "logical unit")
        self.assertEqual(rejected_resume.calls, [])

        resumed = RejectedLLM()
        with self.assertRaises(I.AppraisalOutputRejected):
            I._checkpointed_json_call(
                resumed, state, None, "unit-step", "system", "user", schema,
                100, "logical unit")
        self.assertEqual(resumed.calls, [])

    def test_corrupt_failure_checkpoint_cannot_skip_an_attempt(self):
        test_case = self

        class AlwaysTruncatedLLM:
            def __init__(self):
                self.calls = []

            def json_call(
                    self, _system, _user, _schema, _pass_name,
                    max_tokens=8000, detail=""):
                self.calls.append((max_tokens, detail))
                raise test_case.truncated(detail, max_tokens)

        schema = {"type": "object"}
        state = {"steps": {}}
        with self.assertRaises(I.AppraisalOutputExhausted):
            I._checkpointed_json_call(
                AlwaysTruncatedLLM(), state, None, "unit-step",
                "system", "user", schema, 100, "logical unit")
        state["steps"]["unit-step"]["truncations"][0]["attempt"] = 2

        resumed = AlwaysTruncatedLLM()
        with self.assertRaisesRegex(ValueError, "truncations are invalid"):
            I._checkpointed_json_call(
                resumed, state, None, "unit-step",
                "system", "user", schema, 100, "logical unit")
        self.assertEqual(resumed.calls, [])

    def test_main_exposes_exhausted_truncation_as_nonretryable_exit(self):
        test_case = self

        class TruncatedLLM:
            def __init__(self):
                self.ledger = None
                self.calls = []

            def bind(self, ledger):
                self.ledger = ledger
                return self

            def enable_durable_outcomes(self):
                return self

            def json_call(
                    self, _system, _user, _schema, _pass_name,
                    max_tokens=8000, detail=""):
                self.calls.append((max_tokens, detail))
                self.ledger.record(
                    "anthropic", "investment", model="test-model",
                    in_tok=100, out_tok=max_tokens, detail=f"TRUNCATED {detail}",
                )
                raise test_case.truncated(detail, max_tokens)

        with tempfile.TemporaryDirectory() as directory:
            out = "EXP_terminal"
            payloads = {
                "scans": {
                    "country_findings": [{
                        "chapter_title": "Policy",
                        "statement": "A recorded interoperability gap.",
                        "source_url": "https://example.test/policy",
                    }],
                    "international_pointers": [],
                },
                "foresight": {"milestones": []},
                "ai_assessment": {},
            }
            for suffix, payload in payloads.items():
                with open(os.path.join(directory, f"{out}_{suffix}.json"), "w") as handle:
                    json.dump(payload, handle)
            llm = TruncatedLLM()
            argv = [
                "investment_options.py", "--country", "Exampleland", "--iso", "EXP",
                "--out", out, "--vendor", "anthropic/test-model",
            ]
            with (
                mock.patch.object(I, "LOOP1", directory),
                mock.patch.object(I.V, "load_env"),
                mock.patch.object(
                    I.V, "LLM", side_effect=lambda _vendor, ledger, model=None: (
                        llm.bind(ledger))),
                mock.patch.object(sys, "argv", argv),
            ):
                self.assertEqual(I.main(), I.NONRETRYABLE_STAGE_EXIT)

            self.assertEqual(llm.calls, [
                (I.DEFAULT_APPRAISAL_LIMITS.candidate_output_tokens,
                 "investment candidate map batch 1/1"),
                (I.DEFAULT_APPRAISAL_LIMITS.candidate_output_tokens * 2,
                 "investment candidate map batch 1/1 [truncation retry 2/2]"),
            ])
            state = I.V.strict_json_load(
                os.path.join(directory, f"{out}_investment_state.json"))
            self.assertEqual(state["steps"]["candidate-map-0001"]["status"], "truncated")
            self.assertFalse(os.path.exists(
                os.path.join(directory, f"{out}_investment_options.json")))

    def test_recorded_nigeria_multi_target_repair_continues_in_bounded_chunks(self):
        # Production incident e96a93fd-d4a9-4c83-96d9-3488483729a9: candidate-map
        # batch 2/3 had these four exact overlength fields. Its one bulk repair spent
        # both 4k and 8k output allowances entirely on thinking and emitted zero text.
        test_case = self
        compact_repairs = {
            "candidate-0.recommendation_rationale": "Compact rationale 1.",
            "candidate-1.recommendation_rationale": "Compact rationale 2.",
            "candidate-2.recommendation_rationale": "Compact rationale 3.",
            "candidate-3.problem": "Compact problem 4.",
        }

        class RecordedRepairLLM:
            def __init__(self):
                self.calls = []
                self.repair_target_counts = []
                self.repaired_keys = set()

            def json_call(
                    self, _system, user, schema, _pass_name,
                    max_tokens=8000, detail=""):
                self.calls.append((max_tokens, detail))
                mapped = re.fullmatch(
                    r"investment candidate map batch (\d)/3", detail)
                if mapped:
                    batch = int(mapped.group(1))
                    if batch != 2:
                        return {"candidates": [{
                            "title": f"Batch {batch} candidate",
                            "problem": f"Batch {batch} problem",
                            "recommendation_rationale": f"Batch {batch} rationale",
                            "source_refs": [f"SRC-{batch:03d}"],
                        }]}
                    lengths = (
                        (114, 478, 568),
                        (140, 497, 531),
                        (138, 456, 551),
                        (121, 555, 487),
                    )
                    return {"candidates": [{
                        "title": chr(65 + index) * title_length,
                        "problem": "p" * problem_length,
                        "recommendation_rationale": "r" * rationale_length,
                        "source_refs": ["SRC-002"],
                    } for index, (
                        title_length, problem_length, rationale_length,
                    ) in enumerate(lengths)]}

                repair_properties = (
                    schema.get("properties", {}).get("repairs", {})
                    .get("properties")
                )
                if repair_properties is not None:
                    keys = tuple(repair_properties)
                    self.repair_target_counts.append(len(keys))
                    if len(keys) > 2:
                        raise I.V.VendorOutputTruncated(
                            vendor="anthropic",
                            model="claude-opus-5",
                            pass_name=I.PASS,
                            detail=detail,
                            stop_reason="max_tokens",
                            request_id="recorded-shape-synthetic",
                            max_tokens=max_tokens,
                            input_tokens=2741,
                            output_tokens=max_tokens,
                            thinking_tokens=max_tokens,
                            partial_output_chars=0,
                            partial_output_sha256=hashlib.sha256(b"").hexdigest(),
                        )
                    self.repaired_keys.update(keys)
                    return {"repairs": {
                        key: compact_repairs[key]
                        for key in keys
                    }}

                if detail == "investment candidate final register":
                    supplied = user.split(
                        "SUPPORTED CANDIDATE BRIEFS:\n", 1
                    )[1].split("\n\nReturn", 1)[0]
                    return {"candidates": json.loads(supplied)}
                if detail.startswith("investment appraisal INV-"):
                    candidate = json.loads(
                        user.split("CANDIDATE:\n", 1)[1].split(
                            "\n\nCURRENT APPRAISAL", 1
                        )[0]
                    )
                    option = json.loads(json.dumps(
                        test_case.product()["options"][0]))
                    for field in (
                            "option_id", "title", "problem",
                            "recommendation_rationale", "financing_decision"):
                        option.pop(field)
                    option["costs"]["source_refs"] = candidate["source_refs"]
                    return {"option": option}
                if detail == "investment portfolio sequencing":
                    return {
                        "portfolio_sequencing": "Governance before procurement.",
                        "cross_cutting_data_gaps": ["Validate unit costs."],
                    }
                raise AssertionError(f"unexpected model call: {detail}")

        sources = [{
            "ref": f"SRC-{index:03d}",
            "kind": "country_finding",
            "title": f"Evidence {index}",
            "text": "e" * 180,
            "source": f"https://example.test/{index}",
        } for index in range(1, 4)]
        batch_characters = max(
            len(I.evidence_prompt([source])) for source in sources)
        limits = I.AppraisalLimits(
            evidence_batch_characters=batch_characters,
            candidate_output_tokens=4000,
            option_output_tokens=7000,
            portfolio_output_tokens=2500,
        )
        state = {"steps": {}}
        llm = RecordedRepairLLM()

        response = I.synthesize_appraisal(
            "Nigeria", sources, llm, limits=limits, state=state)
        product = I.build_product("Nigeria", "NGA", response, sources)

        self.assertEqual(I.validate_product(product), [])
        self.assertEqual(llm.repaired_keys, set(compact_repairs))
        self.assertTrue(llm.repair_target_counts)
        self.assertTrue(all(
            1 <= count <= 2 for count in llm.repair_target_counts))
        self.assertTrue(all(
            len(option["title"]) <= 160
            and len(option["problem"]) <= 500
            and len(option["recommendation_rationale"]) <= 500
            for option in response["options"]
        ))

        class SimulatedCrash(Exception):
            pass

        interrupted_state = {"steps": {}}
        durable_state = []
        interrupted_llm = RecordedRepairLLM()

        def crash_after_first_repair_chunk(current):
            if (
                    "candidate-map-0002-length-repair-chunk-0001"
                    in current["steps"]
                    and "candidate-map-0002-length-repair-chunk-0002"
                    not in current["steps"]):
                durable_state[:] = [json.loads(json.dumps(current))]
                raise SimulatedCrash()

        with self.assertRaises(SimulatedCrash):
            I.synthesize_appraisal(
                "Nigeria",
                sources,
                interrupted_llm,
                limits=limits,
                state=interrupted_state,
                save_checkpoint=crash_after_first_repair_chunk,
            )

        resumed_llm = RecordedRepairLLM()
        resumed = I.synthesize_appraisal(
            "Nigeria",
            sources,
            resumed_llm,
            limits=limits,
            state=durable_state[0],
        )
        self.assertEqual(resumed, response)
        self.assertNotIn(
            (4000,
             "investment candidate map batch 2/3 "
             "[local-length repair 1/1 chunk 1/2]"),
            resumed_llm.calls,
        )
        self.assertIn(
            (4000,
             "investment candidate map batch 2/3 "
             "[local-length repair 1/1 chunk 2/2]"),
            resumed_llm.calls,
        )

        class NoReplayLLM:
            def json_call(self, *_args, **_kwargs):
                raise AssertionError("completed paid work was replayed")

        self.assertEqual(
            I.synthesize_appraisal(
                "Nigeria", sources, NoReplayLLM(), limits=limits, state=state),
            response,
        )

    def test_multi_chunk_repair_semantic_failure_is_durable_and_not_replayed(self):
        class DuplicateTitleRepairLLM:
            def __init__(self):
                self.calls = []

            def json_call(
                    self, _system, _user, schema, _pass_name,
                    max_tokens=8000, detail=""):
                self.calls.append((max_tokens, detail))
                if detail == "investment candidate map batch 1/1":
                    return {"candidates": [
                        {
                            "title": "A" * 161,
                            "problem": "p" * 501,
                            "recommendation_rationale": "Rationale 1",
                            "source_refs": ["SRC-001"],
                        },
                        {
                            "title": "B" * 161,
                            "problem": "Problem 2",
                            "recommendation_rationale": "Rationale 2",
                            "source_refs": ["SRC-001"],
                        },
                        {
                            "title": "Option 3",
                            "problem": "Problem 3",
                            "recommendation_rationale": "Rationale 3",
                            "source_refs": ["SRC-001"],
                        },
                    ]}
                if "[local-length repair 1/1 chunk " in detail:
                    keys = schema["properties"]["repairs"]["required"]
                    replacements = {
                        "candidate-0.title": "Duplicate title",
                        "candidate-0.problem": "Compact problem 1",
                        "candidate-1.title": "Duplicate title",
                    }
                    return {"repairs": {
                        key: replacements[key] for key in keys
                    }}
                raise AssertionError(f"unexpected model call: {detail}")

        sources = [{
            "ref": "SRC-001",
            "kind": "country_finding",
            "title": "Policy",
            "text": "Evidence",
            "source": "https://example.test/policy",
        }]
        state = {"steps": {}}
        first = DuplicateTitleRepairLLM()

        with self.assertRaises(I.AppraisalOutputInvalid) as raised:
            I.synthesize_appraisal("Exampleland", sources, first, state=state)

        self.assertEqual(
            raised.exception.step_id,
            "candidate-map-0001-length-repair-chunk-0002",
        )
        self.assertEqual(
            state["steps"][
                "candidate-map-0001-length-repair-chunk-0001"
            ]["status"],
            "complete",
        )
        self.assertEqual(
            state["steps"][
                "candidate-map-0001-length-repair-chunk-0002"
            ]["status"],
            "invalid",
        )

        resumed = DuplicateTitleRepairLLM()
        with self.assertRaises(I.AppraisalOutputInvalid):
            I.synthesize_appraisal(
                "Exampleland", sources, resumed, state=state)
        self.assertEqual(resumed.calls, [])

    def test_sparse_overlong_candidate_text_uses_bounded_second_repair_without_replay(self):
        test_case = self
        compact_problem_4 = "Compact evidence-backed problem 4."
        original_rationale_1 = (
            "Original caveat: this option is not approved. " + "o" * 501
        )
        overlong_repair_rationale_1 = "r" * 501
        compact_rationale_1 = "Compact evidence-backed rationale 1."
        compact_rationale_2 = "Compact evidence-backed rationale 2."
        compact_rationale_4 = "Compact evidence-backed rationale 4."

        class RepairingLLM:
            def __init__(self):
                self.calls = []
                self.final_input = None
                self.mapped = [{
                    "title": f"Option {index + 1}",
                    "problem": "p" * 501 if index == 3 else f"Problem {index + 1}",
                    "recommendation_rationale": (
                        original_rationale_1
                        if index == 0
                        else "r" * 501
                        if index in {0, 1, 3}
                        else f"Rationale {index + 1}"
                    ),
                    "source_refs": ["SRC-001"],
                } for index in range(4)]

            def json_call(
                    self, _system, user, schema, _pass_name,
                    max_tokens=8000, detail=""):
                self.calls.append(detail)
                if detail == "investment candidate map batch 1/1":
                    test_case.assertIn(I.CANDIDATE_TEXT_LIMIT_GUIDANCE, user)
                    return {"candidates": self.mapped}
                if "[local-length repair 1/1 chunk " in detail:
                    required = json.loads(user.split(
                        "REQUIRED_REPAIRS:\n", 1
                    )[1].split("\n\nCANDIDATE_CONTEXT:\n", 1)[0])
                    keys = tuple(
                        schema["properties"]["repairs"]["required"])
                    test_case.assertEqual(
                        tuple(item["key"] for item in required), keys)
                    replacements = {
                        "candidate-0.recommendation_rationale": (
                            overlong_repair_rationale_1
                        ),
                        "candidate-1.recommendation_rationale": compact_rationale_2,
                        "candidate-3.problem": compact_problem_4,
                        "candidate-3.recommendation_rationale": compact_rationale_4,
                    }
                    return {"repairs": {
                        key: replacements[key] for key in keys
                    }}
                if detail == (
                        "investment candidate map batch 1/1 "
                        "[local-length repair 2/2]"):
                    required = json.loads(user.split(
                        "REQUIRED_REPAIRS:\n", 1
                    )[1].split("\n\nCANDIDATE_CONTEXT:\n", 1)[0])
                    test_case.assertEqual(required, [{
                        "key": "candidate-0.recommendation_rationale",
                        "max_characters": 450,
                    }])
                    supplied = json.loads(user.split(
                        "CANDIDATE_CONTEXT:\n", 1
                    )[1])
                    test_case.assertEqual(
                        supplied["candidates"][0]["recommendation_rationale"],
                        original_rationale_1,
                    )
                    test_case.assertNotIn(overlong_repair_rationale_1, user)
                    test_case.assertEqual(max_tokens, 706)
                    return {"repairs": {
                        "candidate-0.recommendation_rationale": compact_rationale_1,
                    }}
                if detail == "investment candidate final register":
                    supplied = user.split(
                        "SUPPORTED CANDIDATE BRIEFS:\n", 1
                    )[1].split("\n\nReturn", 1)[0]
                    self.final_input = json.loads(supplied)
                    return {"candidates": self.final_input}
                if detail.startswith("investment appraisal INV-"):
                    option = json.loads(json.dumps(
                        test_case.product()["options"][0]))
                    for field in (
                            "option_id", "title", "problem",
                            "recommendation_rationale", "financing_decision"):
                        option.pop(field)
                    return {"option": option}
                if detail == "investment portfolio sequencing":
                    return {
                        "portfolio_sequencing": "Governance before procurement.",
                        "cross_cutting_data_gaps": ["Validate unit costs."],
                    }
                raise AssertionError(f"unexpected model call: {detail}")

        sources = [{
            "ref": "SRC-001", "kind": "country_finding", "title": "Policy",
            "text": "Evidence", "source": "https://example.test/policy",
        }]
        state = {"steps": {}}
        llm = RepairingLLM()

        response = I.synthesize_appraisal(
            "Exampleland", sources, llm, state=state,
        )
        product = I.build_product("Exampleland", "EXP", response, sources)

        self.assertEqual(I.validate_product(product), [])
        by_title = {option["title"]: option for option in response["options"]}
        self.assertEqual(
            by_title["Option 1"]["recommendation_rationale"], compact_rationale_1)
        self.assertEqual(
            by_title["Option 2"]["recommendation_rationale"], compact_rationale_2)
        self.assertEqual(by_title["Option 4"]["problem"], compact_problem_4)
        self.assertEqual(
            by_title["Option 4"]["recommendation_rationale"], compact_rationale_4)
        self.assertTrue(all(
            len(option["title"]) <= 160
            and len(option["problem"]) <= 500
            and len(option["recommendation_rationale"]) <= 500
            for option in response["options"]
        ))
        self.assertEqual(
            set(by_title), {"Option 1", "Option 2", "Option 3", "Option 4"})
        self.assertEqual(by_title["Option 1"]["problem"], "Problem 1")
        self.assertEqual(by_title["Option 2"]["problem"], "Problem 2")
        self.assertEqual(by_title["Option 3"]["problem"], "Problem 3")
        self.assertEqual(
            by_title["Option 3"]["recommendation_rationale"], "Rationale 3")
        expected_repaired = json.loads(json.dumps(llm.mapped))
        expected_repaired[0]["recommendation_rationale"] = compact_rationale_1
        expected_repaired[1]["recommendation_rationale"] = compact_rationale_2
        expected_repaired[3].update({
            "problem": compact_problem_4,
            "recommendation_rationale": compact_rationale_4,
        })
        self.assertCountEqual(llm.final_input, expected_repaired)
        self.assertEqual(llm.calls.count(
            "investment candidate map batch 1/1"), 1)
        self.assertEqual(llm.calls.count(
            "investment candidate map batch 1/1 "
            "[local-length repair 1/1 chunk 1/2]"), 1)
        self.assertEqual(llm.calls.count(
            "investment candidate map batch 1/1 "
            "[local-length repair 1/1 chunk 2/2]"), 1)
        self.assertEqual(llm.calls.count(
            "investment candidate map batch 1/1 "
            "[local-length repair 2/2]"), 1)
        self.assertEqual(
            state["steps"][
                "candidate-map-0001-length-repair-chunk-0001"
            ]["status"],
            "complete",
        )
        self.assertEqual(
            state["steps"][
                "candidate-map-0001-length-repair-chunk-0002"
            ]["status"],
            "complete",
        )
        self.assertEqual(
            state["steps"]["candidate-map-0001-length-repair-0002"]["status"],
            "complete",
        )

        class NoReplayLLM:
            def json_call(self, *_args, **_kwargs):
                raise AssertionError("completed paid work was replayed")

        resumed = I.synthesize_appraisal(
            "Exampleland", sources, NoReplayLLM(), state=state,
        )
        self.assertEqual(resumed, response)

    def test_overlong_candidate_title_remains_a_repairable_target(self):
        original = {"candidates": [{
            "title": "t" * 161 if index == 0 else f"Option {index + 1}",
            "problem": f"Problem {index + 1}",
            "recommendation_rationale": f"Rationale {index + 1}",
            "source_refs": ["SRC-001"],
        } for index in range(3)]}

        repair = I._candidate_response_or_length_repair(
            original, {"SRC-001"}, I.CANDIDATE_MAP_SCHEMA)

        self.assertIsInstance(repair, I._CandidateLengthRepair)
        self.assertEqual(repair.targets, ((0, "title", 160),))
        repaired = I._apply_candidate_length_repairs(
            {"repairs": {"candidate-0.title": "Compact title"}},
            repair.response,
            repair.targets,
            {"SRC-001"},
            I.CANDIDATE_MAP_SCHEMA,
            I._candidate_length_repair_schema(repair.targets),
        )
        self.assertEqual(repaired["candidates"][0]["title"], "Compact title")
        self.assertEqual(repaired["candidates"][0]["problem"], "Problem 1")
        self.assertEqual(repaired["candidates"][1:], original["candidates"][1:])

    def test_structurally_invalid_candidate_repair_is_terminal_and_not_replayed(self):
        class InvalidRepairLLM:
            def __init__(self):
                self.calls = []

            def json_call(self, *_args, detail="", **_kwargs):
                self.calls.append(detail)
                if detail == "investment candidate map batch 1/1":
                    return {"candidates": [{
                        "title": f"Option {index}",
                        "problem": "p" * 501 if index == 1 else f"Problem {index}",
                        "recommendation_rationale": f"Rationale {index}",
                        "source_refs": ["SRC-001"],
                    } for index in range(1, 4)]}
                if detail == (
                        "investment candidate map batch 1/1 "
                        "[local-length repair 1/1]"):
                    return {"repairs": {"candidate-0.problem": 42}}
                raise AssertionError(f"unexpected model call: {detail}")

        sources = [{
            "ref": "SRC-001", "kind": "country_finding", "title": "Policy",
            "text": "Evidence", "source": "https://example.test/policy",
        }]
        state = {"steps": {}}
        first = InvalidRepairLLM()

        with self.assertRaises(I.AppraisalOutputInvalid) as raised:
            I.synthesize_appraisal(
                "Exampleland", sources, first, state=state,
            )

        self.assertEqual(
            raised.exception.step_id, "candidate-map-0001-length-repair")
        self.assertEqual(first.calls, [
            "investment candidate map batch 1/1",
            "investment candidate map batch 1/1 [local-length repair 1/1]",
        ])

        resumed = InvalidRepairLLM()
        with self.assertRaises(I.AppraisalOutputInvalid):
            I.synthesize_appraisal(
                "Exampleland", sources, resumed, state=state,
            )
        self.assertEqual(resumed.calls, [])

    def test_second_overlong_candidate_repair_is_terminal_and_not_replayed(self):
        class TwiceOverlongLLM:
            def __init__(self):
                self.calls = []

            def json_call(self, *_args, detail="", **_kwargs):
                self.calls.append(detail)
                if detail == "investment candidate map batch 1/1":
                    return {"candidates": [{
                        "title": f"Option {index}",
                        "problem": "p" * 501 if index == 1 else f"Problem {index}",
                        "recommendation_rationale": f"Rationale {index}",
                        "source_refs": ["SRC-001"],
                    } for index in range(1, 4)]}
                if detail == (
                        "investment candidate map batch 1/1 "
                        "[local-length repair 1/1]"):
                    return {"repairs": {"candidate-0.problem": "p" * 501}}
                if detail == (
                        "investment candidate map batch 1/1 "
                        "[local-length repair 2/2]"):
                    return {"repairs": {"candidate-0.problem": "p" * 451}}
                raise AssertionError(f"unexpected model call: {detail}")

        sources = [{
            "ref": "SRC-001", "kind": "country_finding", "title": "Policy",
            "text": "Evidence", "source": "https://example.test/policy",
        }]
        state = {"steps": {}}
        first = TwiceOverlongLLM()

        with self.assertRaises(I.AppraisalOutputInvalid) as raised:
            I.synthesize_appraisal(
                "Exampleland", sources, first, state=state,
            )

        self.assertEqual(
            raised.exception.step_id,
            "candidate-map-0001-length-repair-0002",
        )
        self.assertEqual(first.calls, [
            "investment candidate map batch 1/1",
            "investment candidate map batch 1/1 [local-length repair 1/1]",
            "investment candidate map batch 1/1 [local-length repair 2/2]",
        ])
        self.assertEqual(
            state["steps"]["candidate-map-0001-length-repair"]["status"],
            "complete",
        )
        self.assertEqual(
            state["steps"]["candidate-map-0001-length-repair-0002"]["status"],
            "invalid",
        )

        resumed = TwiceOverlongLLM()
        with self.assertRaises(I.AppraisalOutputInvalid):
            I.synthesize_appraisal(
                "Exampleland", sources, resumed, state=state,
            )
        self.assertEqual(resumed.calls, [])

    def test_second_length_repair_has_one_bounded_attempt_and_is_not_replayed(self):
        test_case = self

        class TruncatedSecondRepairLLM:
            def __init__(self):
                self.calls = []

            def json_call(
                    self, _system, _user, _schema, _pass_name,
                    max_tokens=8000, detail=""):
                self.calls.append((max_tokens, detail))
                if detail == "investment candidate map batch 1/1":
                    return {"candidates": [{
                        "title": f"Option {index}",
                        "problem": "p" * 501 if index == 1 else f"Problem {index}",
                        "recommendation_rationale": f"Rationale {index}",
                        "source_refs": ["SRC-001"],
                    } for index in range(1, 4)]}
                if detail == (
                        "investment candidate map batch 1/1 "
                        "[local-length repair 1/1]"):
                    return {"repairs": {"candidate-0.problem": "p" * 501}}
                if detail == (
                        "investment candidate map batch 1/1 "
                        "[local-length repair 2/2]"):
                    raise test_case.truncated(detail, max_tokens)
                raise AssertionError(f"unexpected model call: {detail}")

        sources = [{
            "ref": "SRC-001", "kind": "country_finding", "title": "Policy",
            "text": "Evidence", "source": "https://example.test/policy",
        }]
        state = {"steps": {}}
        first = TruncatedSecondRepairLLM()

        with self.assertRaises(I.AppraisalOutputExhausted) as raised:
            I.synthesize_appraisal(
                "Exampleland", sources, first, state=state,
            )

        self.assertEqual(raised.exception.attempt_limit, 1)
        self.assertEqual(first.calls[-1], (
            706,
            "investment candidate map batch 1/1 [local-length repair 2/2]",
        ))
        self.assertEqual(len(first.calls), 3)
        retry_step = state["steps"][
            "candidate-map-0001-length-repair-0002"
        ]
        self.assertEqual(retry_step["status"], "truncated")
        self.assertEqual(len(retry_step["truncations"]), 1)

        resumed = TruncatedSecondRepairLLM()
        with self.assertRaises(I.AppraisalOutputExhausted):
            I.synthesize_appraisal(
                "Exampleland", sources, resumed, state=state,
            )
        self.assertEqual(resumed.calls, [])

    def test_durable_overlong_repair_is_claimed_before_second_repair(self):
        test_case = self

        class SimulatedCrash(Exception):
            pass

        sources = [{
            "ref": "SRC-001", "kind": "country_finding", "title": "Policy",
            "text": "Evidence", "source": "https://example.test/policy",
        }]
        mapped = [{
            "title": f"Option {index}",
            "problem": "p" * 501 if index == 1 else f"Problem {index}",
            "recommendation_rationale": f"Rationale {index}",
            "source_refs": ["SRC-001"],
        } for index in range(1, 4)]
        ledger = I.V.Ledger(ceiling=500, label="test")
        llm = I.V.LLM(
            "anthropic", ledger, model="test-model"
        ).enable_durable_outcomes()
        crashed = False
        first_repair_transports = 0
        second_repair_transports = 0
        second_repair_token_limits = []

        def transport(_system, user, _schema, transport_max_tokens):
            nonlocal first_repair_transports, second_repair_transports
            if "The first targeted candidate-register repair" in user:
                second_repair_transports += 1
                second_repair_token_limits.append(transport_max_tokens)
                return ({
                    "repairs": {
                        "candidate-0.problem": "Compact problem.",
                    },
                }, 10, 20)
            if "REQUIRED_REPAIRS:\n" in user:
                if crashed:
                    raise AssertionError("durable repair outcome was replayed")
                first_repair_transports += 1
                return ({
                    "repairs": {
                        "candidate-0.problem": "p" * 501,
                    },
                }, 10, 20)
            if "\n\nMap this batch to 0-4" in user:
                return ({"candidates": mapped}, 10, 20)
            if "SUPPORTED CANDIDATE BRIEFS:\n" in user:
                supplied = user.split(
                    "SUPPORTED CANDIDATE BRIEFS:\n", 1
                )[1].split("\n\nReturn", 1)[0]
                return ({"candidates": json.loads(supplied)}, 10, 20)
            if "CANDIDATE:\n" in user:
                option = json.loads(json.dumps(
                    test_case.product()["options"][0]))
                for field in (
                        "option_id", "title", "problem",
                        "recommendation_rationale", "financing_decision"):
                    option.pop(field)
                return ({"option": option}, 10, 20)
            if "BOUNDED OPTION PROJECTIONS:\n" in user:
                return ({
                    "portfolio_sequencing": "Governance before procurement.",
                    "cross_cutting_data_gaps": ["Validate unit costs."],
                }, 10, 20)
            raise AssertionError("unexpected model request")

        llm._call_anthropic = transport
        state = {"steps": {}}
        durable_state = []

        def crashing_save(current):
            nonlocal crashed
            if (
                    not crashed
                    and "candidate-map-0001-length-repair" in current["steps"]):
                crashed = True
                raise SimulatedCrash()
            I._bind_state_to_spend_prefix(current, ledger)
            durable_state[:] = [json.loads(json.dumps(current))]

        with self.assertRaises(SimulatedCrash):
            I.synthesize_appraisal(
                "Exampleland", sources, llm,
                state=state, save_checkpoint=crashing_save,
            )

        self.assertTrue(crashed)
        self.assertEqual(first_repair_transports, 1)
        self.assertEqual(second_repair_transports, 0)
        self.assertEqual(
            set(durable_state[0]["steps"]), {"candidate-map-0001"})
        self.assertEqual(durable_state[0]["spend_prefix"]["call_count"], 1)
        self.assertEqual(len(ledger.calls), 2)

        resumed_state = durable_state[0]

        def resumed_save(current):
            I._bind_state_to_spend_prefix(current, ledger)

        response = I.synthesize_appraisal(
            "Exampleland", sources, llm,
            state=resumed_state, save_checkpoint=resumed_save,
        )
        product = I.build_product("Exampleland", "EXP", response, sources)

        self.assertEqual(I.validate_product(product), [])
        self.assertEqual(first_repair_transports, 1)
        self.assertEqual(second_repair_transports, 1)
        self.assertEqual(second_repair_token_limits, [706])

    def test_mixed_candidate_defects_are_terminal_without_length_repair(self):
        class MixedInvalidLLM:
            def __init__(self):
                self.calls = []

            def json_call(self, *_args, detail="", **_kwargs):
                self.calls.append(detail)
                return {"candidates": [{
                    "title": f"Option {index}",
                    "problem": "p" * 501 if index == 1 else f"Problem {index}",
                    "recommendation_rationale": f"Rationale {index}",
                    "source_refs": [
                        "SRC-NOT-KNOWN" if index == 1 else "SRC-001"
                    ],
                } for index in range(1, 4)]}

        sources = [{
            "ref": "SRC-001", "kind": "country_finding", "title": "Policy",
            "text": "Evidence", "source": "https://example.test/policy",
        }]
        state = {"steps": {}}
        first = MixedInvalidLLM()

        with self.assertRaises(I.AppraisalOutputInvalid) as raised:
            I.synthesize_appraisal(
                "Exampleland", sources, first, state=state,
            )

        self.assertEqual(raised.exception.step_id, "candidate-map-0001")
        self.assertEqual(first.calls, ["investment candidate map batch 1/1"])

        resumed = MixedInvalidLLM()
        with self.assertRaises(I.AppraisalOutputInvalid):
            I.synthesize_appraisal(
                "Exampleland", sources, resumed, state=state,
            )
        self.assertEqual(resumed.calls, [])

    def test_invalid_candidate_response_is_terminal_and_not_replayed(self):
        class InvalidLLM:
            def __init__(self):
                self.calls = 0

            def json_call(self, *_args, **_kwargs):
                self.calls += 1
                return {"candidates": [{
                    "title": f"Option {index}",
                    "problem": "Problem",
                    "recommendation_rationale": "Rationale",
                    "source_refs": ["SRC-NOT-KNOWN"],
                } for index in range(1, 4)]}

        sources = [{
            "ref": "SRC-001", "kind": "country_finding", "title": "Policy",
            "text": "Evidence", "source": "https://example.test/policy",
        }]
        state = {"steps": {}}
        persisted = []
        first = InvalidLLM()

        with self.assertRaises(I.AppraisalOutputInvalid):
            I.synthesize_appraisal(
                "Exampleland", sources, first, state=state,
                save_checkpoint=lambda checkpoint: persisted.append(
                    json.loads(json.dumps(checkpoint))),
            )

        self.assertEqual(first.calls, 1)
        step = state["steps"]["candidate-map-0001"]
        self.assertEqual(step["status"], "invalid")
        self.assertNotIn("SRC-NOT-KNOWN", json.dumps(step))
        self.assertEqual(len(persisted), 1)

        resumed = InvalidLLM()
        with self.assertRaises(I.AppraisalOutputInvalid):
            I.synthesize_appraisal(
                "Exampleland", sources, resumed, state=state)
        self.assertEqual(resumed.calls, 0)

    def test_invalid_option_response_is_terminal(self):
        test_case = self

        class InvalidOptionLLM:
            def json_call(self, _system, _user, _schema, _pass_name,
                          max_tokens=8000, detail=""):
                del max_tokens
                if (detail.startswith("investment candidate map batch ")
                        or detail == "investment candidate final register"):
                    return {"candidates": [{
                        "title": f"Option {index}",
                        "problem": "Problem",
                        "recommendation_rationale": "Rationale",
                        "source_refs": ["SRC-001"],
                    } for index in range(1, 4)]}
                option = json.loads(json.dumps(test_case.product()["options"][0]))
                for field in (
                        "option_id", "title", "problem", "recommendation_rationale",
                        "financing_decision"):
                    option.pop(field)
                option["costs"]["low"] = 100.0
                option["costs"]["high"] = None
                return {"option": option}

        sources = [{
            "ref": "SRC-001", "kind": "country_finding", "title": "Policy",
            "text": "Evidence", "source": "https://example.test/policy",
        }]
        state = {"steps": {}}

        with self.assertRaises(I.AppraisalOutputInvalid):
            I.synthesize_appraisal(
                "Exampleland", sources, InvalidOptionLLM(), state=state,
            )

        self.assertEqual(set(state["steps"]), {
            "candidate-map-0001", "candidate-final-register",
            "option-001-batch-0001",
        })
        self.assertEqual(
            state["steps"]["option-001-batch-0001"]["status"], "invalid")

    def test_main_resumes_after_portfolio_failure_without_replaying_prior_calls(self):
        test_case = self

        class CheckpointLLM:
            def __init__(self):
                self.calls = []
                self.ledger = None
                self.fail_portfolio_once = True

            def bind(self, ledger):
                self.ledger = ledger
                return self

            def json_call(
                    self, system, user, schema, pass_name, max_tokens=8000, detail=""):
                self.calls.append(detail)
                if (detail == "investment portfolio sequencing"
                        and self.fail_portfolio_once):
                    # A pre-transport failure has no paid result to recover and may be
                    # retried. A paid failure without a durable outcome must fail closed.
                    self.fail_portfolio_once = False
                    raise I.V.VendorError("temporary portfolio failure")
                refs = sorted(set(re.findall(r"SRC-\d{3}", user)))
                if detail.startswith("investment candidate map batch "):
                    self.register = [{
                        "title": f"Option {index}",
                        "problem": f"Problem {index}",
                        "recommendation_rationale": f"Rationale {index}",
                        "source_refs": refs,
                    } for index in range(1, 4)]
                    response = {"candidates": self.register}
                elif detail == "investment candidate final register":
                    response = {"candidates": self.register}
                elif detail.startswith("investment appraisal INV-"):
                    option = json.loads(json.dumps(test_case.product()["options"][0]))
                    for field in (
                            "option_id", "title", "problem",
                            "recommendation_rationale", "financing_decision"):
                        option.pop(field)
                    response = {"option": option}
                elif detail == "investment portfolio sequencing":
                    response = {
                        "portfolio_sequencing": "Governance before procurement.",
                        "cross_cutting_data_gaps": ["Validate unit costs."],
                    }
                else:
                    raise AssertionError(f"legacy monolithic call was used: {detail}")
                request_sha256 = I.V.json_call_request_sha256(
                    system, user, schema, pass_name, max_tokens, detail)
                self.ledger.record(
                    "anthropic", "investment", model="test-model",
                    in_tok=10, out_tok=20, detail=detail,
                    structured_result={
                        "schema_version": "damm.structured-result/v1",
                        "request_sha256": request_sha256,
                        "outcome": "complete",
                        "response_sha256": I.V.stable_json_sha256(response),
                        "response": response,
                    },
                )
                return response

        with tempfile.TemporaryDirectory() as directory:
            out = "EXP_checkpoint"
            payloads = {
                "scans": {
                    "country_findings": [{
                        "chapter_title": "Policy",
                        "statement": "A recorded interoperability gap.",
                        "source_url": "https://example.test/policy",
                    }],
                    "international_pointers": [],
                },
                "foresight": {"milestones": []},
                "ai_assessment": {},
            }
            for suffix, payload in payloads.items():
                with open(os.path.join(directory, f"{out}_{suffix}.json"), "w") as handle:
                    json.dump(payload, handle)
            llm = CheckpointLLM()
            argv = [
                "investment_options.py", "--country", "Exampleland", "--iso", "EXP",
                "--out", out, "--vendor", "anthropic/test-model",
            ]
            with (
                mock.patch.object(I, "LOOP1", directory),
                mock.patch.object(I.V, "load_env"),
                mock.patch.object(
                    I.V, "LLM", side_effect=lambda _vendor, ledger, model=None: llm.bind(ledger)),
                mock.patch.object(sys, "argv", argv),
            ):
                self.assertEqual(I.main(), 1)
                calls_after_failure = list(llm.calls)
                self.assertTrue(calls_after_failure)
                self.assertTrue(os.path.exists(
                    os.path.join(directory, f"{out}_investment_state.json")))
                self.assertFalse(os.path.exists(
                    os.path.join(directory, f"{out}_investment_options.json")))
                with mock.patch.object(sys, "argv", argv + ["--resume"]):
                    self.assertEqual(I.main(), 0)
                    calls_after_resume = list(llm.calls)
                    self.assertEqual(
                        calls_after_resume,
                        calls_after_failure + ["investment portfolio sequencing"],
                    )
                    self.assertEqual(I.main(), 0)

            self.assertEqual(llm.calls, calls_after_resume)

    def test_adaptive_synthesis_batches_evidence_and_appraises_one_option_at_a_time(self):
        sources = [
            {
                "ref": f"SRC-{index:03d}",
                "kind": "country_finding",
                "title": f"Evidence {index}",
                "text": f"MARKER-{index:03d} " + ("evidence " * 35),
                "source": f"https://example.test/{index}",
            }
            for index in range(1, 13)
        ]

        class TracingLLM:
            def __init__(self):
                self.calls = []
                self.mapped_titles = []

            def json_call(
                    self, system, user, schema, pass_name, max_tokens=8000, detail=""):
                self.calls.append({
                    "user": user,
                    "schema": schema,
                    "pass_name": pass_name,
                    "max_tokens": max_tokens,
                    "detail": detail,
                })
                if detail.startswith("investment candidate map batch "):
                    candidates = schema["properties"]["candidates"]
                    self.assert_candidate_bounds(candidates, 0, 4)
                    refs = sorted(set(re.findall(r"SRC-\d{3}", user)))
                    title = f"Mapped {refs[0]}"
                    self.mapped_titles.append(title)
                    return {"candidates": [{
                        "title": title,
                        "problem": f"Problem evidenced by {refs[0]}",
                        "recommendation_rationale": f"Rationale from {refs[0]}",
                        "source_refs": refs,
                    }]}
                if detail.startswith("investment candidate reduction round "):
                    candidates = schema["properties"]["candidates"]
                    self.assert_candidate_bounds(candidates, 1, 4)
                    refs = sorted(set(re.findall(r"SRC-\d{3}", user)))
                    return {"candidates": [{
                        "title": f"Reduced option {index}",
                        "problem": f"Reduced problem {index}",
                        "recommendation_rationale": f"Reduced rationale {index}",
                        "source_refs": refs,
                    } for index in range(1, 5)]}
                if detail == "investment candidate final register":
                    candidates = schema["properties"]["candidates"]
                    self.assert_candidate_bounds(candidates, 1, 7)
                    refs = sorted(set(re.findall(r"SRC-\d{3}", user)))
                    return {"candidates": [{
                        "title": f"Final option {index}",
                        "problem": f"Final problem {index}",
                        "recommendation_rationale": f"Final rationale {index}",
                        "source_refs": refs,
                    } for index in range(1, 4)]}
                if detail.startswith("investment appraisal INV-"):
                    if set(schema["properties"]) != {"option"}:
                        raise AssertionError("each appraisal call must return one option")
                    assembler_owned = {
                        "option_id", "title", "problem", "recommendation_rationale",
                        "financing_decision",
                    }
                    if assembler_owned & set(schema["properties"]["option"]["properties"]):
                        raise AssertionError("assembler-owned fields leaked into model schema")
                    option_id = re.search(r"INV-\d+", detail).group(0)
                    source_refs = sorted(set(re.findall(r"SRC-\d{3}", user)))
                    return {"option": self.option(option_id, source_refs)}
                if detail == "investment portfolio sequencing":
                    return {
                        "portfolio_sequencing": "Sequence governance before procurement.",
                        "cross_cutting_data_gaps": ["Validate unit costs."],
                    }
                raise AssertionError(f"unexpected model call: {detail}")

            @staticmethod
            def assert_candidate_bounds(candidates, minimum, maximum):
                if (candidates.get("minItems") != minimum
                        or candidates.get("maxItems") != maximum):
                    raise AssertionError(
                        f"candidate schema must enforce {minimum}-{maximum} options")

            @staticmethod
            def option(option_id, source_refs):
                del option_id
                return {
                    "baseline": "The evidence records a fragmented baseline.",
                    "counterfactual": "Fragmentation continues without the investment.",
                    "costs": {
                        "currency": "USD", "base_year": 2026,
                        "low": None, "high": None,
                        "basis": "Not quantified; validate during appraisal.",
                        "source_refs": source_refs[:1],
                    },
                    "benefits": {
                        "quantified": [],
                        "qualitative": ["Potentially reduces fragmentation."],
                    },
                    "horizon_years": 5,
                    "discount_rate": None,
                    "npv_low": None, "npv_high": None,
                    "bcr_low": None, "bcr_high": None,
                    "sensitivity": [{
                        "scenario": "Cost uncertainty",
                        "changes": "Validate scope and unit costs.",
                        "result": "Defer quantified appraisal until evidence exists.",
                    }],
                    "distributional_effects": ["Validate effects on women and smallholders."],
                    "climate_effects": ["Validate climate effects."],
                    "ai_and_data_risks": ["Validate data governance safeguards."],
                    "implementation_risks": ["Institutional fragmentation."],
                    "data_gaps": ["Validated unit costs."],
                    "evidence_status": "Preliminary; validation required.",
                }

        limits = I.AppraisalLimits(
            evidence_batch_characters=500,
            candidate_output_tokens=3000,
            option_output_tokens=7000,
            portfolio_output_tokens=2500,
        )
        batches = I.batch_evidence(sources, limits.evidence_batch_characters)
        llm = TracingLLM()

        response = I.synthesize_appraisal(
            "Exampleland", sources, llm, limits=limits,
        )

        self.assertGreater(len(batches), 1)
        self.assertEqual(
            [source["ref"] for batch in batches for source in batch],
            [source["ref"] for source in sources],
        )
        self.assertTrue(all(
            len(I.evidence_prompt(batch)) <= limits.evidence_batch_characters
            for batch in batches
        ))
        map_calls = [
            call for call in llm.calls
            if call["detail"].startswith("investment candidate map batch ")
        ]
        self.assertEqual(len(map_calls), len(batches))
        for source in sources:
            self.assertTrue(any(source["ref"] in call["user"] for call in map_calls))
        reduction_calls = [
            call for call in llm.calls
            if call["detail"].startswith("investment candidate reduction round ")
        ]
        self.assertTrue(reduction_calls)
        self.assertTrue(all(
            any(title in call["user"] for call in reduction_calls)
            for title in llm.mapped_titles
        ))
        final_calls = [
            call for call in llm.calls
            if call["detail"] == "investment candidate final register"
        ]
        self.assertEqual(len(final_calls), 1)
        option_calls = [
            call for call in llm.calls
            if call["detail"].startswith("investment appraisal INV-")
        ]
        self.assertEqual(
            {re.search(r"INV-\d+", call["detail"]).group(0) for call in option_calls},
            {"INV-1", "INV-2", "INV-3"},
        )
        self.assertGreaterEqual(len(option_calls), 3)
        self.assertTrue(all(
            call["max_tokens"] <= limits.option_output_tokens for call in option_calls
        ))
        seen_by_option = {}
        for call in option_calls:
            option_id = re.search(r"INV-\d+", call["detail"]).group(0)
            seen = seen_by_option.setdefault(option_id, set())
            batch_text = call["user"].split(
                "RELEVANT EVIDENCE BATCH:\n", 1)[1].split("\n\nReturn one", 1)[0]
            seen.update(re.findall(r"SRC-\d{3}", batch_text))
            candidate_text = call["user"].split(
                "CANDIDATE:\n", 1)[1].split("\n\nCURRENT APPRAISAL", 1)[0]
            self.assertEqual(set(json.loads(candidate_text)["source_refs"]), seen)
        portfolio_calls = [
            call for call in llm.calls
            if call["detail"] == "investment portfolio sequencing"
        ]
        self.assertEqual(len(portfolio_calls), 1)
        projection_text = portfolio_calls[0]["user"].split(
            "BOUNDED OPTION PROJECTIONS:\n", 1
        )[1].split("\n\nDescribe preliminary sequencing", 1)[0]
        projections = json.loads(projection_text)
        self.assertEqual(len(projections), 3)
        self.assertEqual(set(projections[0]), set(I.PORTFOLIO_PROJECTION_FIELDS))
        self.assertIn("benefits", projections[0])
        self.assertIn("sensitivity", projections[0])
        self.assertIn("distributional_effects", projections[0])
        self.assertIn("implementation_risks", projections[0])
        self.assertEqual(len(response["options"]), 3)
        product = I.build_product("Exampleland", "EXP", response, sources)
        self.assertEqual(I.validate_product(product), [])
        self.assertEqual(product["schema_version"], "damm.investment-options/v1")
        self.assertEqual(
            [row["ref"] for row in product["source_inventory"]],
            [row["ref"] for row in sources],
        )
        self.assertEqual(
            set(response),
            {"options", "portfolio_sequencing", "cross_cutting_data_gaps"},
        )

    def test_duplicate_reduction_may_collapse_below_three_and_fails_closed(self):
        sources = [{
            "ref": f"SRC-{index:03d}",
            "kind": "country_finding",
            "title": f"Evidence {index}",
            "text": "Evidence for the same investment concept.",
            "source": f"https://example.test/{index}",
        } for index in range(1, 9)]

        class DuplicateLLM:
            def __init__(self):
                self.details = []

            def json_call(
                    self, _system, user, schema, _pass_name,
                    max_tokens=8000, detail=""):
                del max_tokens
                self.details.append(detail)
                refs = sorted(set(re.findall(r"SRC-\d{3}", user)))
                if detail.startswith("investment candidate map batch "):
                    return {"candidates": [{
                        "title": "Shared registry",
                        "problem": "The same fragmented-registry problem.",
                        "recommendation_rationale": "The same supported rationale.",
                        "source_refs": refs,
                    }]}
                if detail.startswith("investment candidate reduction round "):
                    self.assert_reduction_allows_one(schema)
                    return {"candidates": [{
                        "title": "Shared registry",
                        "problem": "The duplicate concepts are one investment.",
                        "recommendation_rationale": "All supplied briefs support it.",
                        "source_refs": refs,
                    }]}
                raise AssertionError(f"unexpected call after duplicate reduction: {detail}")

            @staticmethod
            def assert_reduction_allows_one(schema):
                candidates = schema["properties"]["candidates"]
                if candidates.get("minItems") != 1:
                    raise AssertionError("duplicate reduction must allow one candidate")

        llm = DuplicateLLM()
        limits = I.AppraisalLimits(evidence_batch_characters=100)
        with self.assertRaisesRegex(ValueError, "fewer than three distinct candidates"):
            I.synthesize_appraisal("Exampleland", sources, llm, limits=limits)

        self.assertTrue(any(
            detail.startswith("investment candidate reduction round ")
            for detail in llm.details
        ))
        self.assertNotIn("investment candidate final register", llm.details)

    def test_final_comparison_may_expose_duplicates_below_three_without_padding(self):
        sources = [{
            "ref": f"SRC-{index:03d}",
            "kind": "country_finding",
            "title": f"Evidence {index}",
            "text": "Evidence for the same investment concept.",
            "source": f"https://example.test/{index}",
        } for index in range(1, 5)]

        class DuplicateLLM:
            def __init__(self):
                self.details = []

            def json_call(
                    self, _system, user, schema, _pass_name,
                    max_tokens=8000, detail=""):
                del max_tokens
                self.details.append(detail)
                refs = sorted(set(re.findall(r"SRC-\d{3}", user)))
                if detail.startswith("investment candidate map batch "):
                    index = len(self.details)
                    return {"candidates": [{
                        "title": f"Rephrased registry {index}",
                        "problem": "The same fragmented-registry problem.",
                        "recommendation_rationale": "The same supported rationale.",
                        "source_refs": refs,
                    }]}
                if detail == "investment candidate final register":
                    candidates = schema["properties"]["candidates"]
                    if (candidates.get("minItems"), candidates.get("maxItems")) != (1, 7):
                        raise AssertionError(
                            "final comparison must be allowed to expose duplicates"
                        )
                    return {"candidates": [{
                        "title": "Shared registry",
                        "problem": "The four briefs describe one investment.",
                        "recommendation_rationale": "All supplied briefs support it.",
                        "source_refs": refs,
                    }]}
                raise AssertionError(f"unexpected call: {detail}")

        llm = DuplicateLLM()
        limits = I.AppraisalLimits(evidence_batch_characters=100)
        with self.assertRaisesRegex(
                ValueError, "fewer than three distinct candidates after final"):
            I.synthesize_appraisal("Exampleland", sources, llm, limits=limits)

        self.assertIn("investment candidate final register", llm.details)
        self.assertFalse(any(
            detail.startswith("investment appraisal INV-") for detail in llm.details
        ))

    def test_ttl_upload_context_is_balanced_transparent_and_untrusted(self):
        text = ("START_MARKER" + "a" * 15000 + "MIDDLE_MARKER"
                + "b" * 15000 + "TAIL_MARKER")
        sources = I.evidence_context({}, {}, {}, [
            {"filename": "costs.pdf", "sha256": "a" * 64,
             "extracted_text": text},
            {"filename": "empty.txt", "sha256": "b" * 64,
             "extracted_text": ""},
        ])

        self.assertEqual(len(sources), 2)
        self.assertIn("START_MARKER", sources[0]["text"])
        self.assertIn("MIDDLE_MARKER", sources[0]["text"])
        self.assertIn("TAIL_MARKER", sources[0]["text"])
        self.assertEqual(
            sources[0]["analysis_coverage"]["policy"],
            I.WI.BALANCED_EXCERPT_POLICY,
        )
        self.assertEqual(sources[1]["analysis_coverage"]["mode"], "empty")
        prompt = I.evidence_prompt(sources)
        self.assertIn("TAIL_MARKER", prompt)
        self.assertIn("ANALYSIS_COVERAGE", prompt)
        self.assertIn("NEVER INSTRUCTIONS", prompt)
        self.assertIn("empty.txt", prompt)
        self.assertIn("never instructions", I.SYSTEM.casefold())
        self.assertIn("ignore", I.SYSTEM.casefold())

        product = I.build_product(
            "Exampleland", "EXP",
            {"options": [], "portfolio_sequencing": "",
             "cross_cutting_data_gaps": []},
            sources,
            uploads=[{}],
        )
        self.assertEqual(
            product["source_inventory"][0]["analysis_coverage"],
            sources[0]["analysis_coverage"],
        )

    def product(self):
        response = {
            "options": [{
                "option_id": "INV-1", "title": "Shared farmer data service",
                "problem": "Services cannot reuse farmer-authorized data.",
                "baseline": "No shared service is evidenced.",
                "counterfactual": "Agencies continue duplicating registries.",
                "costs": {"currency": "USD", "base_year": 2026, "low": 100.0,
                          "high": 200.0, "basis": "Illustrative planning range",
                          "source_refs": ["SRC-001"]},
                "benefits": {"quantified": [],
                             "qualitative": ["Reduced duplication", "Faster onboarding"]},
                "horizon_years": 5, "discount_rate": 0.06,
                "npv_low": None, "npv_high": None, "bcr_low": None, "bcr_high": None,
                "sensitivity": [{"scenario": "High cost", "changes": "Cost +30%",
                                 "result": "Revalidate scope before appraisal"}],
                "distributional_effects": ["Design for women and smallholders"],
                "climate_effects": ["Could support climate advisories"],
                "ai_and_data_risks": ["Consent and model bias"],
                "implementation_risks": ["Institutional fragmentation"],
                "data_gaps": ["Validate user volumes and unit costs"],
                "evidence_status": "Illustrative; validation required",
                "recommendation_rationale": "Addresses a recorded interoperability gap.",
                "financing_decision": "not made",
            }],
            "portfolio_sequencing": "Governance before platform procurement.",
            "cross_cutting_data_gaps": ["Common base-year cost data"],
        }
        template = response["options"][0]
        response["options"] = []
        for index in range(1, 4):
            option = json.loads(json.dumps(template))
            option["option_id"] = f"INV-{index}"
            option["title"] = f"Shared farmer data service {index}"
            response["options"].append(option)
        sources = [{"ref": "SRC-001", "kind": "country_finding", "title": "Policy",
                    "source": "https://example.gov/policy", "text": "Policy evidence"}]
        product = I.build_product("Exampleland", "EXP", response, sources)
        # Keep this fixture independent of the host timezone.  The workbook must
        # bind its metadata to the product's assessment date, not to the calendar
        # date on whichever machine happens to run the test.
        product["assessment_date"] = "2026-09-02"
        return product

    def test_valid_product_makes_no_financing_decision(self):
        product = self.product()
        self.assertEqual(I.validate_product(product), [])
        self.assertEqual(product["decision_status"], "no_financing_decision_made")
        self.assertIn("No financing decision", I.render_markdown(product))

    def test_invalid_range_and_unknown_source_are_rejected(self):
        product = self.product()
        option = product["options"][0]
        option["costs"]["low"] = 300
        option["costs"]["source_refs"] = ["SRC-NOT-REAL"]
        errors = I.validate_product(product)
        self.assertTrue(any("cost range is invalid" in error for error in errors))
        self.assertTrue(any("unknown sources" in error for error in errors))

    def test_product_rejects_fewer_than_three_or_more_than_seven_options(self):
        for count in (2, 8):
            with self.subTest(count=count):
                product = self.product()
                template = product["options"][0]
                product["options"] = []
                for index in range(1, count + 1):
                    option = json.loads(json.dumps(template))
                    option["option_id"] = f"INV-{index}"
                    option["title"] = f"Option {index}"
                    product["options"].append(option)

                self.assertIn(
                    "options must contain 3-7 investments",
                    I.validate_product(product),
                )

    def test_product_independently_enforces_bounded_schema(self):
        product = self.product()
        product["options"][0]["horizon_years"] = 0
        product["options"][0]["benefits"]["qualitative"] = [
            f"Benefit {index}" for index in range(5)
        ]
        product["portfolio_sequencing"] = "x" * 1601

        errors = I.validate_product(product)

        self.assertTrue(any("horizon_years" in error and "minimum" in error
                            for error in errors))
        self.assertTrue(any("qualitative" in error and "maxItems" in error
                            for error in errors))
        self.assertTrue(any("portfolio_sequencing" in error and "maxLength" in error
                            for error in errors))

    def test_quantified_benefit_ranges_are_paired_and_ordered_without_crashing(self):
        cases = (
            (10.0, None, ["options[0] benefit range must provide both bounds or neither"]),
            (None, 20.0, ["options[0] benefit range must provide both bounds or neither"]),
            (None, None, []),
            (10.0, 20.0, []),
            (20.0, 10.0, ["options[0] benefit range is invalid"]),
        )
        for low, high, expected in cases:
            with self.subTest(low=low, high=high):
                product = self.product()
                product["options"][0]["benefits"]["quantified"] = [{
                    "name": "Yield improvement",
                    "low": low,
                    "high": high,
                    "unit": "percent",
                    "basis": "Illustrative planning assumption",
                    "source_refs": ["SRC-001"],
                }]

                errors = [
                    error for error in I.validate_product(product)
                    if "benefit range" in error
                ]

                self.assertEqual(errors, expected)

    def test_partial_cost_range_is_rejected_without_crashing(self):
        for low, high in ((100.0, None), (None, 200.0)):
            with self.subTest(low=low, high=high):
                product = self.product()
                product["options"][0]["costs"]["low"] = low
                product["options"][0]["costs"]["high"] = high

                errors = I.validate_product(product)

                self.assertIn(
                    "options[0].cost range must provide both bounds or neither",
                    errors,
                )

    def test_workbook_contains_required_sheets(self):
        with tempfile.TemporaryDirectory() as directory:
            path = os.path.join(directory, "cba.xlsx")
            I.write_workbook(self.product(), path)
            from openpyxl import load_workbook
            workbook = load_workbook(path, read_only=False, data_only=False)
            self.assertEqual(workbook.sheetnames, [
                "Executive Summary", "Cost Ranges", "Options", "Benefits",
                "Sensitivity", "Assumptions & Gaps", "Sources",
            ])
            self.assertEqual(workbook["Options"].freeze_panes, "A2")
            self.assertTrue(workbook["Options"].auto_filter.ref)
            self.assertGreater(workbook["Options"].column_dimensions["B"].width, 20)
            self.assertEqual(workbook["Options"]["L2"].number_format, "0.0%")
            self.assertLessEqual(workbook["Executive Summary"]["A1"].font.sz, 20)
            self.assertTrue(workbook["Executive Summary"]["A1"].alignment.wrap_text)
            self.assertIn(
                "No financing decision",
                " ".join(str(cell.value or "") for row in workbook["Executive Summary"]
                         for cell in row),
            )
            assumptions = " ".join(
                str(cell.value or "") for row in workbook["Assumptions & Gaps"]
                for cell in row
            )
            self.assertIn("Illustrative planning range", assumptions)
            self.assertIn("Validate user volumes and unit costs", assumptions)
            ranges_text = " ".join(
                str(cell.value or "") for row in workbook["Cost Ranges"] for cell in row
            )
            self.assertIn("All options have paired low–high ranges", ranges_text)
            self.assertNotIn("No defensible range", ranges_text)
            workbook.close()

    def test_unbounded_benefit_is_presented_as_quantification_pending(self):
        product = self.product()
        product["options"][0]["benefits"]["quantified"] = [{
            "name": "Avoided duplicate enrollment",
            "low": None,
            "high": None,
            "unit": "households",
            "basis": "Quantify during appraisal",
            "source_refs": ["SRC-001"],
        }]

        rendered = I.render_html(product, "Investment options and CBA")

        self.assertIn(
            '<div class="value">0</div><div class="label">Quantified benefits</div>',
            rendered,
        )
        self.assertIn("Quantification pending", rendered)
        with tempfile.TemporaryDirectory() as directory:
            path = os.path.join(directory, "pending-benefit.xlsx")
            I.write_workbook(product, path)
            from openpyxl import load_workbook
            workbook = load_workbook(path, read_only=False, data_only=False)
            self.assertEqual(workbook["Benefits"]["B2"].value, "Quantification pending")
            workbook.close()

    def test_html_is_offline_escaped_and_visualizes_ranges_without_ranking(self):
        product = self.product()
        product["options"][0]["title"] = "Registry <script>alert(1)</script>"
        product["options"][1]["costs"].update({
            "currency": "NGN", "low": 1000.0, "high": 2400.0,
        })
        product["options"][2]["costs"].update({"low": None, "high": None})

        rendered = I.render_html(product, "Investment options and CBA")

        self.assertEqual(
            rendered,
            I.render_html(product, "Investment options and CBA"),
        )
        self.assertIn('<html lang="en">', rendered)
        self.assertIn('role="img"', rendered)
        self.assertIn("USD — independently scaled", rendered)
        self.assertIn("NGN — independently scaled", rendered)
        self.assertIn("not a ranking", rendered)
        self.assertIn("1 option has no defensible cost range", rendered)
        self.assertIn("Registry &lt;script&gt;alert(1)&lt;/script&gt;", rendered)
        self.assertNotIn("<script>alert(1)</script>", rendered)
        self.assertNotIn("<link", rendered)
        self.assertNotIn("@import", rendered)

    def test_workbook_has_separate_currency_visuals_and_stable_bytes(self):
        product = self.product()
        product["options"][1]["costs"].update({
            "currency": "NGN", "low": 1000.0, "high": 2400.0,
        })
        product["options"][2]["costs"].update({"low": None, "high": None})
        with tempfile.TemporaryDirectory() as directory:
            first = os.path.join(directory, "first.xlsx")
            second = os.path.join(directory, "second.xlsx")

            I.write_workbook(product, first)
            I.write_workbook(product, second)

            with open(first, "rb") as handle:
                first_bytes = handle.read()
            with open(second, "rb") as handle:
                second_bytes = handle.read()
            self.assertEqual(first_bytes, second_bytes)
            with zipfile.ZipFile(first) as archive:
                core_properties = archive.read("docProps/core.xml").decode("utf-8")
            self.assertIn(
                "<dcterms:modified", core_properties,
            )
            self.assertIn(
                ">2026-09-02T00:00:00Z</dcterms:modified>", core_properties,
            )

            from openpyxl import load_workbook
            workbook = load_workbook(first, read_only=False, data_only=False)
            ranges = workbook["Cost Ranges"]
            text = " ".join(str(cell.value or "") for row in ranges for cell in row)
            self.assertIn("USD", text)
            self.assertIn("NGN", text)
            self.assertIn("separate scale", text)
            self.assertIn("not a ranking", text)
            self.assertIn("No defensible range", text)
            self.assertEqual(len(ranges._charts), 2)
            self.assertTrue(all(chart.legend is None for chart in ranges._charts))
            self.assertEqual(ranges["D7"].data_type, "f")
            self.assertEqual(workbook["Executive Summary"]["B9"].data_type, "f")
            self.assertEqual(workbook["Executive Summary"]["E9"].data_type, "f")
            self.assertEqual(workbook["Executive Summary"]["H9"].data_type, "f")
            for sheet in workbook.worksheets:
                expected_width = 2 if sheet.title == "Options" else 1
                self.assertEqual(sheet.page_setup.fitToWidth, expected_width)
                self.assertEqual(sheet.page_setup.fitToHeight, 0)
                self.assertEqual(sheet.page_setup.orientation, "landscape")
                self.assertTrue(sheet.sheet_properties.pageSetUpPr.fitToPage)
            self.assertEqual(workbook["Options"].print_title_cols, "$A:$B")
            workbook.close()

    def test_workbook_normalizer_rejects_missing_or_ambiguous_core_metadata(self):
        def archive(core_properties):
            payload = io.BytesIO()
            with zipfile.ZipFile(payload, "w") as workbook:
                workbook.writestr("xl/workbook.xml", b"<workbook/>")
                if core_properties is not None:
                    workbook.writestr("docProps/core.xml", core_properties)
            return payload.getvalue()

        duplicate_created = (
            b'<core xmlns:dcterms="urn:test">'
            b"<dcterms:created>first</dcterms:created>"
            b"<dcterms:created>second</dcterms:created>"
            b"<dcterms:modified>only</dcterms:modified>"
            b"</core>"
        )
        timestamp = I._workbook_timestamp("2026-09-02")
        for label, raw in (
            ("missing core", archive(None)),
            ("duplicate timestamp", archive(duplicate_created)),
        ):
            with self.subTest(label=label):
                with self.assertRaisesRegex(ValueError, "unique"):
                    I._stable_workbook_bytes(raw, timestamp)

    def test_workbook_never_turns_model_or_source_text_into_excel_formulas(self):
        product = self.product()
        product["options"][0]["title"] = '=HYPERLINK("https://example.test","click")'
        product["options"][0]["problem"] = "+2+3"
        product["options"][0]["costs"]["currency"] = "USD\x01"
        product["options"][0]["recommendation_rationale"] = "bad\uffffvalue"
        product["options"][0]["data_gaps"] = ["bad\ud800value"]
        product["source_inventory"][0]["title"] = "@SUM(A1:A2)"
        with tempfile.TemporaryDirectory() as directory:
            path = os.path.join(directory, "safe.xlsx")
            I.write_workbook(product, path)

            from openpyxl import load_workbook
            workbook = load_workbook(path, read_only=False, data_only=False)
            self.assertNotEqual(workbook["Options"]["B2"].data_type, "f")
            self.assertNotEqual(workbook["Options"]["C2"].data_type, "f")
            self.assertNotEqual(workbook["Sources"]["C2"].data_type, "f")
            self.assertTrue(str(workbook["Options"]["B2"].value).startswith("'="))
            self.assertTrue(str(workbook["Options"]["C2"].value).startswith("'+"))
            self.assertTrue(str(workbook["Sources"]["C2"].value).startswith("'@"))
            self.assertEqual(workbook["Options"]["F2"].value, "USD\ufffd")
            self.assertEqual(workbook["Options"]["R2"].value, "bad\ufffdvalue")
            self.assertEqual(workbook["Options"]["T2"].value, "bad\ufffdvalue")
            self.assertIsNone(workbook["Cost Ranges"]._charts[0].legend)
            workbook.close()


if __name__ == "__main__":
    unittest.main()
