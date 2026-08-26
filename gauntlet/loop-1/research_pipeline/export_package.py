#!/usr/bin/env python3
"""Fail-closed Stage 8 exporter for the canonical DAR workflow.

The workflow-run manifest is the only artifact index accepted by this module.  It
must bind the canonical workflow contract, identify completed stages 1--7, and
carry a hash-bound artifact record for every product required by those stages.
Stage 8 does not discover similarly named files in a working directory.

Narrative source artifacts may be Markdown or HTML.  They are normalized to
Markdown, then exported to standalone HTML and DOCX with pandoc; PDF is produced
from that DOCX with LibreOffice/soffice.  Missing converters, invalid output, or
any artifact/hash mismatch is a terminal packaging error.
"""

from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
import datetime as _datetime
import hashlib
import json
import os
from pathlib import Path
import re
import shutil
import subprocess
import sys
import tempfile
from typing import Any, Callable, Iterable, Mapping, Sequence
import zipfile


HERE = Path(__file__).resolve().parent
WORKFLOW_CONTRACT = HERE.parents[2] / "workflow" / "dar-workflow-v1.json"

WORKFLOW_SCHEMA = "damm.workflow-run/v1"
PACKAGE_SCHEMA = "damm.dar-package/v1"
UPLOADS_SCHEMA = "damm.uploads-manifest/v1"
EXPECTED_WORKFLOW_ID = "dar-canonical-v1"
EXPECTED_STAGE_IDS = (
    "damm_diagnostic",
    "country_research",
    "ai_digital_agriculture",
    "international_lessons",
    "strategic_foresight",
    "investment_options",
    "draft_dar",
    "export_package",
)
NARRATIVE_ARTIFACTS = {
    "damm_diagnostic": "diagnostic_report",
    "country_research": "country_research_report",
    "ai_digital_agriculture": "ai_assessment_report",
    "international_lessons": "international_lessons_report",
    "strategic_foresight": "foresight_report",
    "investment_options": "investment_options_report",
    "draft_dar": "draft_dar_report",
}
SOURCE_INVENTORY_STAGES = EXPECTED_STAGE_IDS[:6]
HASH_RE = re.compile(r"^[0-9a-f]{64}$")
ISO3_RE = re.compile(r"^[A-Z]{3}$")
UPLOAD_KINDS = frozenset(
    {
        "country_context_documents",
        "ai_documents",
        "international_strategy_documents",
        "foresight_documents",
        "investment_documents",
    }
)

PandocConverter = Callable[[Path, Path], None]
PdfConverter = Callable[[Path, Path], None]


class PackagingError(RuntimeError):
    """A terminal Stage 8 failure."""


class ManifestValidationError(PackagingError):
    """The workflow-run manifest does not authorize packaging."""


@dataclass(frozen=True)
class ArtifactBinding:
    stage_id: str
    key: str
    path: str
    sha256: str
    media_type: str


@dataclass(frozen=True)
class ValidatedWorkflow:
    country: str
    iso3: str
    workflow_id: str
    workflow_version: str
    contract_sha256: str
    input_snapshot_path: str
    input_snapshot_sha256: str
    uploads_manifest_path: str | None
    uploads_manifest_sha256: str | None
    uploads_document_count: int
    artifacts: Mapping[str, Mapping[str, ArtifactBinding]]
    artifact_records: Mapping[str, Mapping[str, tuple[ArtifactBinding, ...]]]


@dataclass(frozen=True)
class VerifiedUploadDocument:
    id: str
    kind: str
    original_filename: str
    content_path: str
    content_sha256: str
    content_media_type: str
    content: bytes
    original_path: str
    original_sha256: str
    original_size_bytes: int
    original: bytes
    metadata: Mapping[str, Any]


@dataclass(frozen=True)
class VerifiedUploads:
    manifest_path: str
    manifest_sha256: str
    manifest_content: bytes
    documents: tuple[VerifiedUploadDocument, ...]


@dataclass(frozen=True)
class PackageResult:
    package_dir: Path
    package_manifest: Path
    zip_path: Path


def _reject_duplicate_keys(pairs: Sequence[tuple[str, Any]]) -> dict[str, Any]:
    result: dict[str, Any] = {}
    for key, value in pairs:
        if key in result:
            raise ManifestValidationError(f"JSON contains duplicate key {key!r}")
        result[key] = value
    return result


def load_json_bytes(content: bytes, label: str) -> Any:
    try:
        return json.loads(
            content.decode("utf-8"), object_pairs_hook=_reject_duplicate_keys
        )
    except (UnicodeDecodeError, json.JSONDecodeError) as error:
        raise ManifestValidationError(f"{label} is not valid UTF-8 JSON: {error}") from error


def load_json_file(path: Path, label: str | None = None) -> Any:
    try:
        content = path.read_bytes()
    except OSError as error:
        raise PackagingError(f"cannot read {label or path}: {error}") from error
    return load_json_bytes(content, label or str(path))


def sha256_bytes(content: bytes) -> str:
    return hashlib.sha256(content).hexdigest()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    try:
        with path.open("rb") as handle:
            for block in iter(lambda: handle.read(1024 * 1024), b""):
                digest.update(block)
    except OSError as error:
        raise PackagingError(f"cannot hash {path}: {error}") from error
    return digest.hexdigest()


def _canonical_json_bytes(value: Any) -> bytes:
    return (json.dumps(value, sort_keys=True, ensure_ascii=False, indent=2) + "\n").encode(
        "utf-8"
    )


def _normalized_country(value: Any) -> str:
    return " ".join(str(value or "").split())


def _require_safe_relative_path(value: Any, label: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ManifestValidationError(f"{label}.path is empty")
    if "\\" in value:
        raise ManifestValidationError(f"{label}.path must use portable forward slashes")
    path = Path(value)
    if path.is_absolute() or ".." in path.parts or value.startswith("/"):
        raise ManifestValidationError(f"{label}.path is not a safe relative path")
    if any(part in {"", "."} for part in path.parts):
        raise ManifestValidationError(f"{label}.path is not normalized")
    return path.as_posix()


def _contract_stage_map(contract: Mapping[str, Any]) -> dict[str, Mapping[str, Any]]:
    if contract.get("schema_version") != "damm.dar-workflow/v1":
        raise ManifestValidationError("workflow contract has the wrong schema_version")
    if contract.get("workflow_id") != EXPECTED_WORKFLOW_ID:
        raise ManifestValidationError("workflow contract has the wrong workflow_id")
    stages = contract.get("stages")
    if not isinstance(stages, list):
        raise ManifestValidationError("workflow contract stages is not an array")
    ids = [stage.get("id") for stage in stages if isinstance(stage, dict)]
    if tuple(ids) != EXPECTED_STAGE_IDS:
        raise ManifestValidationError("workflow contract does not contain the canonical eight stages")
    profiles = contract.get("export_profiles")
    expected_profiles = {
        "narrative": ["md", "docx", "pdf", "html"],
        "structured": ["xlsx", "csv", "json"],
        "package": ["zip", "json"],
    }
    if profiles != expected_profiles:
        raise ManifestValidationError("workflow contract export_profiles is not canonical")
    return {stage["id"]: stage for stage in stages}


def _artifact_binding(stage_id: str, record: Any) -> ArtifactBinding:
    label = f"workflow stage {stage_id} artifact"
    if not isinstance(record, dict):
        raise ManifestValidationError(f"{label} is not an object")
    key = record.get("key")
    if not isinstance(key, str) or not key:
        raise ManifestValidationError(f"{label} has no key")
    path = _require_safe_relative_path(record.get("path"), f"{label} {key}")
    digest = record.get("sha256")
    if not isinstance(digest, str) or not HASH_RE.fullmatch(digest):
        raise ManifestValidationError(f"{label} {key} has no valid SHA-256")
    media_type = record.get("media_type")
    if not isinstance(media_type, str) or not media_type.strip():
        raise ManifestValidationError(f"{label} {key} has no media_type")
    return ArtifactBinding(stage_id, key, path, digest, media_type.strip())


def _validate_artifact_extension(binding: ArtifactBinding) -> None:
    suffix = Path(binding.path).suffix.lower()
    narrative_key = NARRATIVE_ARTIFACTS.get(binding.stage_id)
    if binding.key == narrative_key:
        if suffix not in {".md", ".markdown", ".html", ".htm"}:
            raise ManifestValidationError(
                f"{binding.stage_id}.{binding.key} must be a Markdown or HTML source"
            )
        return
    if binding.key == "cost_benefit_workbook":
        if suffix != ".xlsx":
            raise ManifestValidationError("cost_benefit_workbook must be an .xlsx file")
        return
    if suffix != ".json":
        raise ManifestValidationError(
            f"{binding.stage_id}.{binding.key} must be a structured JSON artifact"
        )


def validate_workflow_manifest(
    manifest: Any,
    contract: Mapping[str, Any],
    *,
    country: str,
    iso3: str,
    contract_sha256: str,
) -> ValidatedWorkflow:
    """Purely validate a workflow-run manifest and return its trusted bindings.

    This function performs no filesystem access.  File resolution, byte hashing, and
    stage-manifest replay happen separately in :func:`build_export_package`.
    """

    contract_stages = _contract_stage_map(contract)
    if not isinstance(manifest, dict):
        raise ManifestValidationError("workflow manifest is not an object")
    if manifest.get("schema_version") != WORKFLOW_SCHEMA:
        raise ManifestValidationError(f"workflow manifest schema_version is not {WORKFLOW_SCHEMA}")
    if manifest.get("workflow_id") != contract.get("workflow_id"):
        raise ManifestValidationError("workflow manifest workflow_id does not match the contract")
    if manifest.get("workflow_version") != contract.get("workflow_version"):
        raise ManifestValidationError("workflow manifest workflow_version does not match the contract")
    if not HASH_RE.fullmatch(str(contract_sha256 or "")):
        raise ManifestValidationError("canonical contract SHA-256 is invalid")
    if manifest.get("contract_sha256") != contract_sha256:
        raise ManifestValidationError("workflow manifest is not bound to the canonical contract bytes")

    requested_country = _normalized_country(country)
    manifest_country = _normalized_country(manifest.get("country"))
    if not requested_country:
        raise ManifestValidationError("country is empty")
    if requested_country.casefold() != manifest_country.casefold():
        raise ManifestValidationError("workflow manifest country does not match --country")
    requested_iso = str(iso3 or "").strip().upper()
    if not ISO3_RE.fullmatch(requested_iso):
        raise ManifestValidationError("--iso must be a three-letter ISO code")
    if str(manifest.get("iso3") or "").strip().upper() != requested_iso:
        raise ManifestValidationError("workflow manifest ISO does not match --iso")

    snapshot = manifest.get("input_snapshot")
    if not isinstance(snapshot, dict):
        raise ManifestValidationError("workflow manifest input_snapshot is not an object")
    snapshot_path = _require_safe_relative_path(snapshot.get("path"), "input_snapshot")
    snapshot_hash = snapshot.get("sha256")
    if not isinstance(snapshot_hash, str) or not HASH_RE.fullmatch(snapshot_hash):
        raise ManifestValidationError("workflow manifest input_snapshot has no valid SHA-256")

    uploads = manifest.get("uploads_manifest")
    uploads_path: str | None = None
    uploads_hash: str | None = None
    uploads_count = 0
    if uploads is not None:
        if not isinstance(uploads, dict):
            raise ManifestValidationError(
                "workflow manifest uploads_manifest is not an object"
            )
        uploads_path = _require_safe_relative_path(
            uploads.get("path"), "uploads_manifest"
        )
        if uploads_path != "inputs/uploads-manifest.json":
            raise ManifestValidationError(
                "workflow manifest uploads_manifest path is not canonical"
            )
        uploads_hash = uploads.get("sha256")
        if not isinstance(uploads_hash, str) or not HASH_RE.fullmatch(uploads_hash):
            raise ManifestValidationError(
                "workflow manifest uploads_manifest has no valid SHA-256"
            )
        uploads_count = uploads.get("document_count")
        if (
            isinstance(uploads_count, bool)
            or not isinstance(uploads_count, int)
            or uploads_count < 0
        ):
            raise ManifestValidationError(
                "workflow manifest uploads_manifest document_count is invalid"
            )

    stages = manifest.get("stages")
    if not isinstance(stages, list):
        raise ManifestValidationError("workflow manifest stages is not an array")
    if len(stages) != len(EXPECTED_STAGE_IDS):
        raise ManifestValidationError("workflow manifest must contain exactly eight stages")
    stage_ids = [stage.get("id") for stage in stages if isinstance(stage, dict)]
    if tuple(stage_ids) != EXPECTED_STAGE_IDS:
        raise ManifestValidationError("workflow manifest stages are missing, duplicated, or out of order")

    status = manifest.get("status")
    current_stage = manifest.get("current_stage")
    stage8_status = stages[-1].get("status") if isinstance(stages[-1], dict) else None
    active_stage8 = (
        status in {"running", "retrying"}
        and current_stage == "export_package"
        and stage8_status in {"running", "retrying"}
    )
    completed_run = status == "complete" and current_stage is None and stage8_status == "complete"
    if not (active_stage8 or completed_run):
        raise ManifestValidationError(
            "workflow manifest does not represent active Stage 8 or a completed workflow"
        )

    by_stage: dict[str, dict[str, ArtifactBinding]] = {}
    records_by_stage: dict[str, dict[str, tuple[ArtifactBinding, ...]]] = {}
    all_paths: dict[str, tuple[str, str]] = {}
    for stage, stage_id in zip(stages[:7], EXPECTED_STAGE_IDS[:7]):
        if not isinstance(stage, dict):
            raise ManifestValidationError(f"workflow stage {stage_id} is not an object")
        if stage.get("status") != "complete":
            raise ManifestValidationError(f"workflow stage {stage_id} is not complete")
        records = stage.get("artifacts")
        if not isinstance(records, list):
            raise ManifestValidationError(f"workflow stage {stage_id} artifacts is not an array")
        grouped: dict[str, list[ArtifactBinding]] = {}
        for record in records:
            binding = _artifact_binding(stage_id, record)
            prior = all_paths.get(binding.path)
            if prior is not None and prior[1] != binding.sha256:
                raise ManifestValidationError(
                    f"artifact path {binding.path} is assigned conflicting hashes by "
                    f"{prior[0]} and {stage_id}.{binding.key}"
                )
            all_paths[binding.path] = (f"{stage_id}.{binding.key}", binding.sha256)
            grouped.setdefault(binding.key, []).append(binding)
        required = contract_stages[stage_id].get("required_artifacts")
        if not isinstance(required, list):
            raise ManifestValidationError(f"contract stage {stage_id} has no required_artifacts")
        missing = [key for key in required if key not in grouped]
        if missing:
            raise ManifestValidationError(
                f"workflow stage {stage_id} is missing required artifacts: {', '.join(missing)}"
            )
        selected: dict[str, ArtifactBinding] = {}
        required_records: dict[str, tuple[ArtifactBinding, ...]] = {}
        narrative_key = NARRATIVE_ARTIFACTS[stage_id]
        for key in required:
            candidates = grouped[key]
            for binding in candidates:
                _validate_artifact_extension(binding)
            if key != narrative_key and len(candidates) != 1:
                raise ManifestValidationError(
                    f"workflow stage {stage_id} duplicates artifact key {key}"
                )
            if key == narrative_key:
                normalized_suffixes = [
                    "md"
                    if Path(binding.path).suffix.lower() in {".md", ".markdown"}
                    else "html"
                    for binding in candidates
                ]
                if len(normalized_suffixes) != len(set(normalized_suffixes)):
                    raise ManifestValidationError(
                        f"workflow stage {stage_id} has duplicate narrative source formats"
                    )
                candidates = sorted(
                    candidates,
                    key=lambda binding: (
                        0
                        if Path(binding.path).suffix.lower() in {".md", ".markdown"}
                        else 1,
                        binding.path,
                    ),
                )
            selected[key] = candidates[0]
            required_records[key] = tuple(candidates)
        by_stage[stage_id] = selected
        records_by_stage[stage_id] = required_records

    return ValidatedWorkflow(
        country=manifest_country,
        iso3=requested_iso,
        workflow_id=str(contract["workflow_id"]),
        workflow_version=str(contract["workflow_version"]),
        contract_sha256=contract_sha256,
        input_snapshot_path=snapshot_path,
        input_snapshot_sha256=snapshot_hash,
        uploads_manifest_path=uploads_path,
        uploads_manifest_sha256=uploads_hash,
        uploads_document_count=uploads_count,
        artifacts=by_stage,
        artifact_records=records_by_stage,
    )


def _resolve_source(root: Path, relative: str, label: str) -> Path:
    try:
        root_resolved = root.resolve(strict=True)
        resolved = (root_resolved / relative).resolve(strict=True)
    except OSError as error:
        raise PackagingError(f"{label} does not exist: {error}") from error
    try:
        resolved.relative_to(root_resolved)
    except ValueError as error:
        raise PackagingError(f"{label} resolves outside the workflow directory") from error
    if not resolved.is_file():
        raise PackagingError(f"{label} is not a regular file")
    return resolved


def _validate_input_snapshot(
    snapshot: Any,
    *,
    workflow: ValidatedWorkflow,
    workflow_manifest: Mapping[str, Any],
) -> None:
    """Require the immutable launch snapshot to bind every Stage 8 input."""

    expected_uploads = None
    if workflow.uploads_manifest_path is not None:
        expected_uploads = {
            "path": workflow.uploads_manifest_path,
            "sha256": workflow.uploads_manifest_sha256,
            "document_count": workflow.uploads_document_count,
        }
    expected = {
        "schema_version": "damm.workflow-input-snapshot/v1",
        "country": workflow.country,
        "iso3": workflow.iso3,
        "contract_sha256": workflow.contract_sha256,
        "uploads_manifest": expected_uploads,
        "ceiling_usd": workflow_manifest.get("ceiling_usd"),
        "vendor": workflow_manifest.get("vendor"),
    }
    if snapshot != expected:
        raise ManifestValidationError(
            "input_snapshot does not exactly bind the canonical workflow launch inputs"
        )


def _resolve_upload_source(
    root: Path,
    relative: Any,
    *,
    label: str,
    required_parent: str,
) -> tuple[str, Path]:
    """Resolve a frozen upload path without permitting traversal or symlinks."""

    portable = _require_safe_relative_path(relative, label)
    path = Path(portable)
    if tuple(path.parts[:2]) != ("inputs", required_parent):
        raise ManifestValidationError(
            f"{label}.path must be below inputs/{required_parent}/"
        )
    try:
        root_resolved = root.resolve(strict=True)
        unresolved = root_resolved / path
        resolved = unresolved.resolve(strict=True)
        resolved.relative_to(root_resolved)
    except (OSError, ValueError) as error:
        raise PackagingError(f"{label} is missing or escapes the workflow directory") from error
    if unresolved.is_symlink() or resolved != unresolved.absolute():
        raise PackagingError(f"{label} may not traverse a symbolic link")
    if not resolved.is_file():
        raise PackagingError(f"{label} is not a regular file")
    return portable, resolved


def _required_nonempty_string(value: Any, label: str) -> str:
    if not isinstance(value, str) or not value.strip():
        raise ManifestValidationError(f"{label} must be a nonempty string")
    return value


def _read_upload_bytes(path: Path, expected_sha256: Any, label: str) -> bytes:
    if not isinstance(expected_sha256, str) or not HASH_RE.fullmatch(expected_sha256):
        raise ManifestValidationError(f"{label} has no valid lowercase SHA-256")
    content = _read_verified_source(path, expected_sha256, label)
    return content


def _verify_frozen_uploads(
    root: Path, workflow: ValidatedWorkflow
) -> VerifiedUploads | None:
    """Verify the upload envelope, extracted text, and immutable original bytes."""

    if workflow.uploads_manifest_path is None:
        return None
    if workflow.uploads_manifest_sha256 is None:
        raise ManifestValidationError("uploads manifest binding is incomplete")
    manifest_relative = workflow.uploads_manifest_path
    # The manifest itself is the one canonical file directly below inputs/; document
    # payloads are constrained more narrowly by _resolve_upload_source below.
    try:
        root_resolved = root.resolve(strict=True)
        unresolved_manifest = root_resolved / manifest_relative
        resolved_manifest = unresolved_manifest.resolve(strict=True)
        resolved_manifest.relative_to(root_resolved)
    except (OSError, ValueError) as error:
        raise PackagingError("uploads_manifest is missing or escapes the workflow directory") from error
    if (
        unresolved_manifest.is_symlink()
        or resolved_manifest != unresolved_manifest.absolute()
        or not resolved_manifest.is_file()
    ):
        raise PackagingError("uploads_manifest may not traverse a symbolic link")
    manifest_source = resolved_manifest
    manifest_content = _read_upload_bytes(
        manifest_source, workflow.uploads_manifest_sha256, "uploads_manifest"
    )
    envelope = load_json_bytes(manifest_content, "uploads_manifest")
    if not isinstance(envelope, dict) or envelope.get("schema_version") != UPLOADS_SCHEMA:
        raise ManifestValidationError(
            f"uploads_manifest must be a {UPLOADS_SCHEMA} object"
        )
    raw_documents = envelope.get("documents")
    if not isinstance(raw_documents, list):
        raise ManifestValidationError("uploads_manifest documents must be an array")
    if len(raw_documents) != workflow.uploads_document_count:
        raise ManifestValidationError(
            "uploads_manifest document_count does not match the workflow binding"
        )

    documents: list[VerifiedUploadDocument] = []
    seen_ids: set[str] = set()
    seen_paths: set[str] = set()
    for index, raw_document in enumerate(raw_documents):
        label = f"uploads_manifest documents[{index}]"
        if not isinstance(raw_document, dict):
            raise ManifestValidationError(f"{label} is not an object")
        document_id = _required_nonempty_string(raw_document.get("id"), f"{label}.id")
        if document_id in seen_ids:
            raise ManifestValidationError(f"uploads_manifest duplicates id {document_id!r}")
        seen_ids.add(document_id)
        kind = _required_nonempty_string(raw_document.get("kind"), f"{label}.kind")
        if kind not in UPLOAD_KINDS:
            raise ManifestValidationError(f"{label}.kind is not canonical")
        original_filename = _required_nonempty_string(
            raw_document.get("original_filename"), f"{label}.original_filename"
        )
        content_path, content_source = _resolve_upload_source(
            root,
            raw_document.get("content_path"),
            label=f"{label}.content_path",
            required_parent="upload-content",
        )
        original_path, original_source = _resolve_upload_source(
            root,
            raw_document.get("original_path"),
            label=f"{label}.original_path",
            required_parent="upload-originals",
        )
        for portable in (content_path, original_path):
            if portable in seen_paths:
                raise ManifestValidationError(
                    f"uploads_manifest duplicates input path {portable!r}"
                )
            seen_paths.add(portable)
        media_type = _required_nonempty_string(
            raw_document.get("content_media_type"), f"{label}.content_media_type"
        )
        if media_type != "text/plain":
            raise ManifestValidationError(f"{label}.content_media_type must be text/plain")
        content_sha = raw_document.get("content_sha256")
        content = _read_upload_bytes(content_source, content_sha, f"{label} extracted text")
        try:
            decoded = content.decode("utf-8")
        except UnicodeDecodeError as error:
            raise ManifestValidationError(
                f"{label} extracted text is not valid UTF-8"
            ) from error
        original_sha = raw_document.get("original_sha256")
        original = _read_upload_bytes(original_source, original_sha, f"{label} original")
        original_size = raw_document.get("original_size_bytes")
        if (
            isinstance(original_size, bool)
            or not isinstance(original_size, int)
            or original_size < 0
            or original_size != len(original)
        ):
            raise ManifestValidationError(
                f"{label}.original_size_bytes does not match the original"
            )
        metadata = raw_document.get("metadata")
        if not isinstance(metadata, dict):
            raise ManifestValidationError(f"{label}.metadata is not an object")
        for field in (
            "app_upload_kind",
            "source_mime_type",
            "uploaded_at",
            "uploaded_by",
            "extraction_status",
        ):
            _required_nonempty_string(metadata.get(field), f"{label}.metadata.{field}")
        extracted_characters = metadata.get("extracted_characters")
        if (
            isinstance(extracted_characters, bool)
            or not isinstance(extracted_characters, int)
            or extracted_characters != len(decoded)
        ):
            raise ManifestValidationError(
                f"{label}.metadata.extracted_characters does not match the text"
            )
        if metadata.get("app_upload_kind") != kind:
            raise ManifestValidationError(
                f"{label}.metadata.app_upload_kind does not match kind"
            )
        if metadata.get("extraction_status") != "extracted":
            raise ManifestValidationError(
                f"{label}.metadata.extraction_status must be extracted"
            )
        documents.append(
            VerifiedUploadDocument(
                id=document_id,
                kind=kind,
                original_filename=original_filename,
                content_path=content_path,
                content_sha256=str(content_sha),
                content_media_type=media_type,
                content=content,
                original_path=original_path,
                original_sha256=str(original_sha),
                original_size_bytes=original_size,
                original=original,
                metadata=dict(metadata),
            )
        )
    return VerifiedUploads(
        manifest_path=manifest_relative,
        manifest_sha256=workflow.uploads_manifest_sha256,
        manifest_content=manifest_content,
        documents=tuple(documents),
    )


def _read_verified_source(path: Path, expected_sha256: str, label: str) -> bytes:
    try:
        content = path.read_bytes()
    except OSError as error:
        raise PackagingError(f"cannot read {label}: {error}") from error
    actual = sha256_bytes(content)
    if actual != expected_sha256:
        raise PackagingError(
            f"{label} SHA-256 mismatch: expected {expected_sha256}, found {actual}"
        )
    return content


def _stage_output_hashes(value: Any, label: str) -> dict[str, list[str]]:
    records: dict[str, list[str]] = {}

    def add(key: Any, digest: Any) -> None:
        if not isinstance(key, str) or not isinstance(digest, str):
            return
        if not HASH_RE.fullmatch(digest):
            return
        records.setdefault(key, []).append(digest)

    if isinstance(value, dict):
        for key, item in value.items():
            if isinstance(item, list):
                for child in item:
                    add(key, child.get("sha256") if isinstance(child, dict) else child)
            else:
                add(key, item.get("sha256") if isinstance(item, dict) else item)
    elif isinstance(value, list):
        for item in value:
            if not isinstance(item, dict):
                continue
            key = item.get("key") or item.get("artifact_id") or item.get("id")
            digest = item.get("sha256")
            add(key, digest)
    else:
        raise ManifestValidationError(f"{label}.output_hashes is not an object or array")
    return records


def _validate_stage_manifest(
    content: bytes,
    *,
    stage_id: str,
    workflow: ValidatedWorkflow,
    contract: Mapping[str, Any],
) -> None:
    label = f"{stage_id}.stage_manifest"
    value = load_json_bytes(content, label)
    if not isinstance(value, dict):
        raise ManifestValidationError(f"{label} is not an object")
    required_fields = contract.get("stage_manifest_required_fields")
    if not isinstance(required_fields, list):
        raise ManifestValidationError("workflow contract has no stage_manifest_required_fields")
    missing = [field for field in required_fields if field not in value]
    if missing:
        raise ManifestValidationError(f"{label} is missing fields: {', '.join(missing)}")
    if value.get("workflow_id") != workflow.workflow_id:
        raise ManifestValidationError(f"{label}.workflow_id does not match")
    if value.get("workflow_version") != workflow.workflow_version:
        raise ManifestValidationError(f"{label}.workflow_version does not match")
    if value.get("stage_id") != stage_id:
        raise ManifestValidationError(f"{label}.stage_id does not match")
    if value.get("status") != "complete":
        raise ManifestValidationError(f"{label}.status is not complete")
    if not isinstance(value.get("execution_mode"), str) or not value["execution_mode"].strip():
        raise ManifestValidationError(f"{label}.execution_mode is empty")
    spend = value.get("spend_usd")
    if isinstance(spend, bool) or not isinstance(spend, (int, float)) or spend < 0:
        raise ManifestValidationError(f"{label}.spend_usd is invalid")
    if not isinstance(value.get("input_hashes"), (dict, list)):
        raise ManifestValidationError(f"{label}.input_hashes is invalid")
    if not isinstance(value.get("quality_checks"), (dict, list)):
        raise ManifestValidationError(f"{label}.quality_checks is invalid")
    checks: Iterable[Any]
    quality = value["quality_checks"]
    checks = quality.values() if isinstance(quality, dict) else quality
    for check in checks:
        if check is False or (isinstance(check, dict) and check.get("ok") is False):
            raise ManifestValidationError(f"{label} contains a failed quality check")
    if value.get("source_inventory") is None:
        raise ManifestValidationError(f"{label}.source_inventory is missing")

    output_hashes = _stage_output_hashes(value.get("output_hashes"), label)
    for key, bindings in workflow.artifact_records[stage_id].items():
        if key == "stage_manifest":
            continue
        expected = sorted(binding.sha256 for binding in bindings)
        actual = sorted(output_hashes.get(key, []))
        if actual != expected:
            raise ManifestValidationError(
                f"{label}.output_hashes does not bind {key} to the workflow artifact hash"
            )


def _normalize_text(content: bytes, label: str) -> str:
    try:
        text = content.decode("utf-8")
    except UnicodeDecodeError as error:
        raise PackagingError(f"{label} is not UTF-8 text") from error
    if "\x00" in text:
        raise PackagingError(f"{label} contains a NUL byte")
    text = text.replace("\r\n", "\n").replace("\r", "\n")
    text = text.rstrip() + "\n"
    if not text.strip():
        raise PackagingError(f"{label} is empty")
    return text


def default_pandoc_converter(source: Path, target: Path) -> None:
    executable = shutil.which("pandoc")
    if not executable:
        raise PackagingError("pandoc is required for narrative exports but was not found")
    command = [executable, str(source), "--standalone", "--output", str(target)]
    if target.suffix.lower() == ".md":
        command[2:2] = ["--to", "gfm"]
    try:
        result = subprocess.run(
            command, text=True, capture_output=True, check=False, timeout=180
        )
    except (OSError, subprocess.TimeoutExpired) as error:
        raise PackagingError(f"pandoc conversion failed for {source.name}: {error}") from error
    if result.returncode != 0:
        detail = (result.stderr or result.stdout).strip()[-1000:]
        raise PackagingError(
            f"pandoc conversion failed for {source.name} with exit {result.returncode}: {detail}"
        )


def default_pdf_converter(source_docx: Path, target_pdf: Path) -> None:
    executable = shutil.which("soffice") or shutil.which("libreoffice")
    if not executable:
        raise PackagingError("soffice is required for PDF exports but was not found")
    target_pdf.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.TemporaryDirectory(prefix="damm-soffice-") as temp_name:
        temp = Path(temp_name)
        profile = temp / "profile"
        converted = temp / source_docx.with_suffix(".pdf").name
        command = [
            executable,
            f"-env:UserInstallation={profile.resolve().as_uri()}",
            "--headless",
            "--convert-to",
            "pdf",
            "--outdir",
            str(temp),
            str(source_docx),
        ]
        try:
            result = subprocess.run(
                command, text=True, capture_output=True, check=False, timeout=180
            )
        except (OSError, subprocess.TimeoutExpired) as error:
            raise PackagingError(
                f"soffice conversion failed for {source_docx.name}: {error}"
            ) from error
        if result.returncode != 0 or not converted.is_file():
            detail = (result.stderr or result.stdout).strip()[-1000:]
            raise PackagingError(
                f"soffice conversion failed for {source_docx.name} with exit "
                f"{result.returncode}: {detail}"
            )
        shutil.copyfile(converted, target_pdf)


def _require_conversion_output(path: Path, kind: str) -> None:
    if not path.is_file() or path.stat().st_size == 0:
        raise PackagingError(f"{kind} converter did not create {path.name}")
    if kind == "DOCX":
        if not zipfile.is_zipfile(path):
            raise PackagingError(f"pandoc output {path.name} is not a valid DOCX container")
        try:
            with zipfile.ZipFile(path) as archive:
                names = set(archive.namelist())
        except (OSError, zipfile.BadZipFile) as error:
            raise PackagingError(f"cannot inspect DOCX output {path.name}: {error}") from error
        if "[Content_Types].xml" not in names or "word/document.xml" not in names:
            raise PackagingError(f"pandoc output {path.name} is missing DOCX document parts")
    elif kind == "PDF":
        try:
            header = path.read_bytes()[:5]
        except OSError as error:
            raise PackagingError(f"cannot inspect PDF output {path.name}: {error}") from error
        if header != b"%PDF-":
            raise PackagingError(f"soffice output {path.name} is not a PDF")


def _write_verified_bytes(destination: Path, content: bytes) -> None:
    destination.parent.mkdir(parents=True, exist_ok=True)
    try:
        destination.write_bytes(content)
    except OSError as error:
        raise PackagingError(f"cannot write {destination}: {error}") from error
    if sha256_file(destination) != sha256_bytes(content):
        raise PackagingError(f"written artifact {destination} failed read-back verification")


def _convert_narrative(
    *,
    source_content: bytes,
    source_suffix: str,
    destination_stem: Path,
    work_dir: Path,
    pandoc_converter: PandocConverter,
    pdf_converter: PdfConverter,
) -> list[Path]:
    destination_stem.parent.mkdir(parents=True, exist_ok=True)
    markdown_path = destination_stem.with_suffix(".md")
    html_path = destination_stem.with_suffix(".html")
    docx_path = destination_stem.with_suffix(".docx")
    pdf_path = destination_stem.with_suffix(".pdf")

    suffix = source_suffix.lower()
    if suffix in {".md", ".markdown"}:
        markdown_path.write_text(
            _normalize_text(source_content, destination_stem.name), encoding="utf-8", newline="\n"
        )
    else:
        source_html = work_dir / f"{destination_stem.name}-source.html"
        source_html.write_text(
            _normalize_text(source_content, destination_stem.name), encoding="utf-8", newline="\n"
        )
        pandoc_converter(source_html, markdown_path)
        if not markdown_path.is_file() or markdown_path.stat().st_size == 0:
            raise PackagingError(f"pandoc did not create {markdown_path.name}")
        normalized = _normalize_text(markdown_path.read_bytes(), markdown_path.name)
        markdown_path.write_text(normalized, encoding="utf-8", newline="\n")

    pandoc_converter(markdown_path, html_path)
    if not html_path.is_file() or html_path.stat().st_size == 0:
        raise PackagingError(f"pandoc did not create {html_path.name}")
    normalized_html = _normalize_text(html_path.read_bytes(), html_path.name)
    if "<html" not in normalized_html.casefold():
        raise PackagingError(f"pandoc output {html_path.name} is not standalone HTML")
    html_path.write_text(normalized_html, encoding="utf-8", newline="\n")

    pandoc_converter(markdown_path, docx_path)
    _require_conversion_output(docx_path, "DOCX")
    pdf_converter(docx_path, pdf_path)
    _require_conversion_output(pdf_path, "PDF")
    return [markdown_path, docx_path, pdf_path, html_path]


def _source_inventory_rows(value: Any, stage_id: str) -> list[dict[str, Any]]:
    if isinstance(value, dict):
        if isinstance(value.get("source_inventory"), list):
            value = value["source_inventory"]
        elif isinstance(value.get("sources"), list):
            value = value["sources"]
    if not isinstance(value, list):
        raise ManifestValidationError(f"{stage_id}.source_inventory is not an array")
    if not value:
        raise ManifestValidationError(f"{stage_id}.source_inventory is empty")
    rows: list[dict[str, Any]] = []
    reserved = {"stage_id", "source_inventory_artifact", "source_index"}
    for index, item in enumerate(value, 1):
        if not isinstance(item, dict):
            raise ManifestValidationError(
                f"{stage_id}.source_inventory[{index - 1}] is not an object"
            )
        row: dict[str, Any] = {
            "stage_id": stage_id,
            "source_inventory_artifact": "source_inventory",
            "source_index": index,
        }
        for key, cell in item.items():
            if not isinstance(key, str) or not key:
                raise ManifestValidationError(
                    f"{stage_id}.source_inventory[{index - 1}] has an invalid field name"
                )
            row[f"record_{key}" if key in reserved else key] = cell
        rows.append(row)
    return rows


def _tabular_value(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (str, int, float, bool)):
        return value
    return json.dumps(value, sort_keys=True, ensure_ascii=False, separators=(",", ":"))


def _inventory_columns(rows: Sequence[Mapping[str, Any]]) -> list[str]:
    preferred = [
        "stage_id",
        "source_inventory_artifact",
        "source_index",
        "ref",
        "source_id",
        "title",
        "publisher",
        "kind",
        "url",
        "source",
        "tier",
        "accessed_on",
        "sha256",
    ]
    available = {key for row in rows for key in row}
    return [key for key in preferred if key in available] + sorted(available - set(preferred))


def _write_inventory_csv(path: Path, rows: Sequence[Mapping[str, Any]]) -> None:
    columns = _inventory_columns(rows)
    path.parent.mkdir(parents=True, exist_ok=True)
    try:
        with path.open("w", encoding="utf-8", newline="") as handle:
            writer = csv.DictWriter(handle, fieldnames=columns, lineterminator="\n")
            writer.writeheader()
            for row in rows:
                writer.writerow({key: _tabular_value(row.get(key)) for key in columns})
    except (OSError, csv.Error) as error:
        raise PackagingError(f"cannot write consolidated source inventory CSV: {error}") from error


def _write_inventory_xlsx(path: Path, rows: Sequence[Mapping[str, Any]]) -> None:
    try:
        from openpyxl import Workbook, load_workbook
        from openpyxl.styles import Font
    except ImportError as error:
        raise PackagingError("openpyxl is required for structured XLSX exports") from error
    columns = _inventory_columns(rows)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Sources"
    sheet.append(columns)
    for row in rows:
        sheet.append([_tabular_value(row.get(key)) for key in columns])
    for cell in sheet[1]:
        cell.font = Font(bold=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions
    for index, column in enumerate(columns, 1):
        observed = [len(str(column))]
        observed.extend(len(str(_tabular_value(row.get(column)))) for row in rows[:200])
        sheet.column_dimensions[sheet.cell(1, index).column_letter].width = min(
            max(observed) + 2, 60
        )
    try:
        workbook.save(path)
        workbook.close()
        verification = load_workbook(path, read_only=True, data_only=False)
        if verification.sheetnames != ["Sources"]:
            raise PackagingError("consolidated source inventory XLSX has unexpected sheets")
        verification.close()
    except (OSError, ValueError, zipfile.BadZipFile) as error:
        raise PackagingError(f"cannot create consolidated source inventory XLSX: {error}") from error


def _validate_xlsx(content: bytes, label: str, work_dir: Path) -> None:
    try:
        from openpyxl import load_workbook
    except ImportError as error:
        raise PackagingError("openpyxl is required to validate cost_benefit.xlsx") from error
    candidate = work_dir / "cost_benefit-validation.xlsx"
    candidate.write_bytes(content)
    try:
        workbook = load_workbook(candidate, read_only=True, data_only=False)
        if not workbook.sheetnames:
            raise PackagingError(f"{label} has no worksheets")
        workbook.close()
    except Exception as error:
        if isinstance(error, PackagingError):
            raise
        raise PackagingError(f"{label} is not a readable XLSX workbook: {error}") from error


def _file_record(
    package_root: Path,
    path: Path,
    *,
    category: str,
    stage_id: str | None = None,
    artifact_id: str | None = None,
    source_sha256: str | None = None,
    input_id: str | None = None,
    input_kind: str | None = None,
) -> dict[str, Any]:
    relative = path.relative_to(package_root).as_posix()
    record: dict[str, Any] = {
        "path": relative,
        "sha256": sha256_file(path),
        "bytes": path.stat().st_size,
        "category": category,
    }
    if stage_id:
        record["stage_id"] = stage_id
    if artifact_id:
        record["artifact_id"] = artifact_id
    if source_sha256:
        record["source_sha256"] = source_sha256
    if input_id:
        record["input_id"] = input_id
    if input_kind:
        record["input_kind"] = input_kind
    return record


def validate_package_files(package_root: Path, manifest: Mapping[str, Any]) -> None:
    """Verify that the package contains exactly the hash-bound payload files."""

    if manifest.get("schema_version") != PACKAGE_SCHEMA:
        raise PackagingError("package manifest has the wrong schema_version")
    nodes = list(package_root.rglob("*"))
    if any(path.is_symlink() for path in nodes):
        raise PackagingError("package payload must not contain symbolic links")
    records = manifest.get("files")
    if not isinstance(records, list) or manifest.get("file_count") != len(records):
        raise PackagingError("package manifest file_count is invalid")
    expected: set[str] = set()
    for index, record in enumerate(records):
        if not isinstance(record, dict):
            raise PackagingError(f"package manifest files[{index}] is not an object")
        relative = _require_safe_relative_path(record.get("path"), f"package files[{index}]")
        if relative in expected:
            raise PackagingError(f"package manifest duplicates {relative}")
        expected.add(relative)
        path = package_root / relative
        if not path.is_file():
            raise PackagingError(f"package payload {relative} is missing")
        if path.stat().st_size != record.get("bytes"):
            raise PackagingError(f"package payload {relative} byte count does not match")
        if sha256_file(path) != record.get("sha256"):
            raise PackagingError(f"package payload {relative} SHA-256 does not match")
    actual = {
        path.relative_to(package_root).as_posix()
        for path in nodes
        if path.is_file()
        and path.relative_to(package_root).as_posix() != "package-manifest.json"
    }
    if actual != expected:
        raise PackagingError("package payload set does not match the SHA-256 manifest")


def _required_package_record(
    manifest: Mapping[str, Any], relative_path: str
) -> Mapping[str, Any]:
    records = manifest.get("files")
    if not isinstance(records, list):
        raise PackagingError("package manifest files is not an array")
    matches = [
        record
        for record in records
        if isinstance(record, dict) and record.get("path") == relative_path
    ]
    if len(matches) != 1:
        raise PackagingError(
            f"package manifest must bind exactly one {relative_path} payload"
        )
    return matches[0]


def _workflow_artifact_signature(workflow: ValidatedWorkflow) -> dict[str, Any]:
    """Return the immutable Stage 1--7 portion of a workflow-run identity."""

    return {
        stage_id: {
            key: sorted(
                (
                    binding.path,
                    binding.sha256,
                    binding.media_type,
                )
                for binding in bindings
            )
            for key, bindings in sorted(workflow.artifact_records[stage_id].items())
        }
        for stage_id in EXPECTED_STAGE_IDS[:7]
    }


def _upload_input_signature(uploads: VerifiedUploads | None) -> dict[str, Any] | None:
    if uploads is None:
        return None
    return {
        "schema_version": UPLOADS_SCHEMA,
        "manifest_path": uploads.manifest_path,
        "manifest_sha256": uploads.manifest_sha256,
        "document_count": len(uploads.documents),
        "documents": [
            {
                "id": document.id,
                "kind": document.kind,
                "content_path": document.content_path,
                "content_sha256": document.content_sha256,
                "original_path": document.original_path,
                "original_sha256": document.original_sha256,
                "original_size_bytes": document.original_size_bytes,
            }
            for document in uploads.documents
        ],
    }


def _validate_packaged_upload_records(
    package_root: Path,
    package_manifest: Mapping[str, Any],
    uploads: VerifiedUploads | None,
) -> None:
    expected_signature = _upload_input_signature(uploads)
    if package_manifest.get("upload_inputs") != expected_signature:
        raise PackagingError("existing package upload input identity does not match")
    if uploads is None:
        return
    expected = [
        (
            uploads.manifest_path,
            uploads.manifest_sha256,
            "uploads_manifest",
            None,
            None,
            uploads.manifest_content,
        )
    ]
    for document in uploads.documents:
        expected.extend(
            [
                (
                    document.content_path,
                    document.content_sha256,
                    "upload_extracted_text",
                    document.id,
                    document.kind,
                    document.content,
                ),
                (
                    document.original_path,
                    document.original_sha256,
                    "upload_original",
                    document.id,
                    document.kind,
                    document.original,
                ),
            ]
        )
    for relative, digest, artifact_id, input_id, input_kind, content in expected:
        record = _required_package_record(package_manifest, relative)
        if (
            record.get("category") != "input"
            or record.get("artifact_id") != artifact_id
            or record.get("source_sha256") != digest
            or record.get("sha256") != digest
            or record.get("input_id") != input_id
            or record.get("input_kind") != input_kind
        ):
            raise PackagingError(
                f"existing package input payload {relative} is not provenance-bound"
            )
        try:
            packaged = (package_root / relative).read_bytes()
        except OSError as error:
            raise PackagingError(
                f"cannot read existing package input payload {relative}: {error}"
            ) from error
        if packaged != content:
            raise PackagingError(
                f"existing package input payload {relative} is not byte-identical"
            )


def _validate_existing_package(
    *,
    package_root: Path,
    workflow: ValidatedWorkflow,
    contract: Mapping[str, Any],
    contract_content: bytes,
    snapshot_content: bytes,
    uploads: VerifiedUploads | None,
) -> Path:
    """Validate a published package without trusting filenames or prior process state."""

    if not package_root.is_dir() or package_root.is_symlink():
        raise PackagingError(
            f"versioned package path is not a regular directory: {package_root}"
        )
    manifest_path = package_root / "package-manifest.json"
    try:
        manifest_content = manifest_path.read_bytes()
    except OSError as error:
        raise PackagingError(f"cannot read existing package manifest: {error}") from error
    manifest = load_json_bytes(manifest_content, "existing package manifest")
    if not isinstance(manifest, dict):
        raise PackagingError("existing package manifest is not an object")
    if manifest_content != _canonical_json_bytes(manifest):
        raise PackagingError("existing package manifest is not canonical SHA-bound JSON")

    expected_identity = {
        "schema_version": PACKAGE_SCHEMA,
        "package_version": workflow.workflow_version,
        "workflow_id": workflow.workflow_id,
        "workflow_version": workflow.workflow_version,
        "workflow_contract_sha256": workflow.contract_sha256,
        "input_snapshot_sha256": workflow.input_snapshot_sha256,
        "country": workflow.country,
        "iso3": workflow.iso3,
        "lifecycle_state": contract["execution_policy"]["output_lifecycle_state"],
        "export_profiles": contract["export_profiles"],
    }
    for key, expected in expected_identity.items():
        if manifest.get(key) != expected:
            raise PackagingError(
                f"existing package identity mismatch for {key}: "
                f"expected {expected!r}, found {manifest.get(key)!r}"
            )
    packaged_manifest_sha = manifest.get("workflow_manifest_sha256")
    if not isinstance(packaged_manifest_sha, str) or not HASH_RE.fullmatch(
        packaged_manifest_sha
    ):
        raise PackagingError("existing package has no valid workflow_manifest_sha256")

    validate_package_files(package_root, manifest)
    _validate_packaged_upload_records(package_root, manifest, uploads)
    identity_files = {
        "workflow/workflow_manifest.json": packaged_manifest_sha,
        "workflow/dar-workflow-v1.json": sha256_bytes(contract_content),
        "workflow/input_snapshot.json": sha256_bytes(snapshot_content),
    }
    for relative, expected_sha in identity_files.items():
        record = _required_package_record(manifest, relative)
        if record.get("sha256") != expected_sha or record.get("source_sha256") != expected_sha:
            raise PackagingError(
                f"existing package identity payload {relative} is not SHA-bound"
            )

    try:
        packaged_contract_content = (
            package_root / "workflow/dar-workflow-v1.json"
        ).read_bytes()
        packaged_snapshot_content = (
            package_root / "workflow/input_snapshot.json"
        ).read_bytes()
        packaged_workflow_content = (
            package_root / "workflow/workflow_manifest.json"
        ).read_bytes()
    except OSError as error:
        raise PackagingError(
            f"cannot read existing package identity payload: {error}"
        ) from error
    if packaged_contract_content != contract_content:
        raise PackagingError("existing package workflow contract bytes do not match")
    if packaged_snapshot_content != snapshot_content:
        raise PackagingError("existing package input snapshot bytes do not match")

    if sha256_bytes(packaged_workflow_content) != packaged_manifest_sha:
        raise PackagingError("existing packaged workflow manifest SHA-256 does not match")
    packaged_workflow_value = load_json_bytes(
        packaged_workflow_content, "existing packaged workflow manifest"
    )
    packaged_workflow = validate_workflow_manifest(
        packaged_workflow_value,
        contract,
        country=workflow.country,
        iso3=workflow.iso3,
        contract_sha256=workflow.contract_sha256,
    )
    if _workflow_artifact_signature(packaged_workflow) != _workflow_artifact_signature(
        workflow
    ):
        raise PackagingError(
            "existing package Stage 1--7 artifact identity does not match the workflow"
        )
    return manifest_path


def validate_package_zip(package_root: Path, target: Path) -> None:
    """Verify that a ZIP contains exactly the byte-identical package directory."""

    if not target.is_file() or target.is_symlink():
        raise PackagingError(f"complete bundle is not a regular ZIP file: {target}")
    expected_paths = sorted(path for path in package_root.rglob("*") if path.is_file())
    expected = {
        f"{package_root.name}/{path.relative_to(package_root).as_posix()}": path
        for path in expected_paths
    }
    try:
        with zipfile.ZipFile(target) as archive:
            infos = archive.infolist()
            names = [info.filename for info in infos]
            if len(names) != len(set(names)):
                raise PackagingError("complete bundle ZIP contains duplicate members")
            if set(names) != set(expected) or any(info.is_dir() for info in infos):
                raise PackagingError("complete bundle ZIP content set does not match package")
            if archive.testzip() is not None:
                raise PackagingError("complete bundle ZIP failed CRC verification")
            for info in infos:
                source = expected[info.filename]
                if info.file_size != source.stat().st_size:
                    raise PackagingError(
                        f"complete bundle ZIP byte count mismatch for {info.filename}"
                    )
                digest = hashlib.sha256()
                with archive.open(info, "r") as stream:
                    for block in iter(lambda: stream.read(1024 * 1024), b""):
                        digest.update(block)
                if digest.hexdigest() != sha256_file(source):
                    raise PackagingError(
                        f"complete bundle ZIP SHA-256 mismatch for {info.filename}"
                    )
    except PackagingError:
        raise
    except (OSError, zipfile.BadZipFile, RuntimeError) as error:
        raise PackagingError(f"cannot validate complete bundle ZIP: {error}") from error


def _zip_package(package_root: Path, target: Path) -> None:
    try:
        with zipfile.ZipFile(target, "w", compression=zipfile.ZIP_DEFLATED) as archive:
            for path in sorted(p for p in package_root.rglob("*") if p.is_file()):
                archive.write(
                    path,
                    arcname=f"{package_root.name}/{path.relative_to(package_root).as_posix()}",
                )
    except (OSError, zipfile.BadZipFile) as error:
        raise PackagingError(f"cannot create complete bundle ZIP: {error}") from error
    validate_package_zip(package_root, target)


def _recreate_zip_atomically(package_root: Path, target: Path) -> None:
    """Publish a recovered ZIP with no observable partial public file."""

    target.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temp_name = tempfile.mkstemp(
        prefix=f".{target.name}-recover-", suffix=".tmp", dir=target.parent
    )
    os.close(descriptor)
    temp_path = Path(temp_name)
    try:
        _zip_package(package_root, temp_path)
        os.replace(temp_path, target)
        validate_package_zip(package_root, target)
    finally:
        try:
            temp_path.unlink()
        except FileNotFoundError:
            pass


def _created_at(value: str | None) -> str:
    if value is not None:
        if not isinstance(value, str) or not value.strip():
            raise PackagingError("created_at must be a nonempty timestamp")
        return value
    return (
        _datetime.datetime.now(_datetime.timezone.utc)
        .replace(microsecond=0)
        .isoformat()
        .replace("+00:00", "Z")
    )


def build_export_package(
    *,
    country: str,
    iso3: str,
    out: Path | str,
    workflow_manifest: Path | str,
    pandoc_converter: PandocConverter = default_pandoc_converter,
    pdf_converter: PdfConverter = default_pdf_converter,
    contract_path: Path | str = WORKFLOW_CONTRACT,
    created_at: str | None = None,
    resume: bool = False,
) -> PackageResult:
    """Build or idempotently resume one versioned DAR package and ZIP."""

    manifest_path = Path(workflow_manifest)
    contract_file = Path(contract_path)
    manifest_content = manifest_path.read_bytes()
    contract_content = contract_file.read_bytes()
    manifest_value = load_json_bytes(manifest_content, "workflow manifest")
    contract_value = load_json_bytes(contract_content, "workflow contract")
    if not isinstance(contract_value, dict):
        raise ManifestValidationError("workflow contract is not an object")
    workflow = validate_workflow_manifest(
        manifest_value,
        contract_value,
        country=country,
        iso3=iso3,
        contract_sha256=sha256_bytes(contract_content),
    )

    manifest_root = manifest_path.parent
    snapshot_source = _resolve_source(
        manifest_root, workflow.input_snapshot_path, "input_snapshot"
    )
    snapshot_content = _read_verified_source(
        snapshot_source, workflow.input_snapshot_sha256, "input_snapshot"
    )
    snapshot_value = load_json_bytes(snapshot_content, "input_snapshot")
    _validate_input_snapshot(
        snapshot_value,
        workflow=workflow,
        workflow_manifest=manifest_value,
    )
    uploads = _verify_frozen_uploads(manifest_root, workflow)

    out_path = Path(out)
    if not out_path.name or out_path.name in {".", ".."}:
        raise PackagingError("--out must include a package filename prefix")
    output_parent = out_path.parent
    output_parent.mkdir(parents=True, exist_ok=True)
    package_name = f"{out_path.name}_dar_package_v{workflow.workflow_version}"
    final_package = output_parent / package_name
    final_zip = Path(f"{out_path}_dar_package.zip")

    if resume:
        if final_zip.exists() and not final_package.exists():
            raise PackagingError(
                "cannot resume from a ZIP without its versioned package directory"
            )
        if final_package.exists():
            final_manifest = _validate_existing_package(
                package_root=final_package,
                workflow=workflow,
                contract=contract_value,
                contract_content=contract_content,
                snapshot_content=snapshot_content,
                uploads=uploads,
            )
            if final_zip.exists():
                validate_package_zip(final_package, final_zip)
            else:
                _recreate_zip_atomically(final_package, final_zip)
            return PackageResult(final_package, final_manifest, final_zip)
    if final_package.exists():
        raise PackagingError(f"versioned package directory already exists: {final_package}")
    if final_zip.exists():
        raise PackagingError(f"complete bundle ZIP already exists: {final_zip}")

    resolved: dict[tuple[str, str], tuple[ArtifactBinding, Path, bytes]] = {}
    for stage_id, grouped in workflow.artifact_records.items():
        for key, bindings in grouped.items():
            candidates: list[tuple[ArtifactBinding, Path, bytes]] = []
            for binding in bindings:
                label = f"{stage_id}.{key} ({binding.path})"
                source = _resolve_source(manifest_root, binding.path, label)
                content = _read_verified_source(source, binding.sha256, label)
                if Path(binding.path).suffix.lower() == ".json":
                    load_json_bytes(content, label)
                candidates.append((binding, source, content))
            selected_path = workflow.artifacts[stage_id][key].path
            resolved[(stage_id, key)] = next(
                candidate for candidate in candidates if candidate[0].path == selected_path
            )

    for stage_id in EXPECTED_STAGE_IDS[:7]:
        binding, _source, content = resolved[(stage_id, "stage_manifest")]
        _validate_stage_manifest(
            content,
            stage_id=stage_id,
            workflow=workflow,
            contract=contract_value,
        )

    temp_parent = Path(tempfile.mkdtemp(prefix=f".{out_path.name}-stage8-", dir=output_parent))
    package_root = temp_parent / package_name
    work_dir = temp_parent / "conversion-work"
    temp_zip = temp_parent / final_zip.name
    published_package = False
    published_zip = False
    try:
        package_root.mkdir()
        work_dir.mkdir()
        records: list[dict[str, Any]] = []

        for ordinal, stage_id in enumerate(EXPECTED_STAGE_IDS[:7], 1):
            artifact_id = NARRATIVE_ARTIFACTS[stage_id]
            binding, _source, content = resolved[(stage_id, artifact_id)]
            stem = package_root / "narratives" / f"{ordinal:02d}_{stage_id}" / artifact_id
            outputs = _convert_narrative(
                source_content=content,
                source_suffix=Path(binding.path).suffix,
                destination_stem=stem,
                work_dir=work_dir,
                pandoc_converter=pandoc_converter,
                pdf_converter=pdf_converter,
            )
            for output in outputs:
                records.append(
                    _file_record(
                        package_root,
                        output,
                        category="narrative",
                        stage_id=stage_id,
                        artifact_id=artifact_id,
                        source_sha256=binding.sha256,
                    )
                )

        inventory_rows: list[dict[str, Any]] = []
        for ordinal, stage_id in enumerate(EXPECTED_STAGE_IDS[:7], 1):
            bindings = workflow.artifacts[stage_id]
            for artifact_id in sorted(bindings):
                if artifact_id == NARRATIVE_ARTIFACTS[stage_id]:
                    continue
                binding, _source, content = resolved[(stage_id, artifact_id)]
                if artifact_id == "source_inventory":
                    destination = (
                        package_root
                        / "source-inventory"
                        / "raw"
                        / f"{ordinal:02d}_{stage_id}_source_inventory.json"
                    )
                    _write_verified_bytes(destination, content)
                    inventory_rows.extend(
                        _source_inventory_rows(
                            load_json_bytes(content, f"{stage_id}.source_inventory"), stage_id
                        )
                    )
                    category = "source_inventory"
                elif artifact_id == "cost_benefit_workbook":
                    _validate_xlsx(content, f"{stage_id}.{artifact_id}", work_dir)
                    destination = package_root / "structured" / f"{ordinal:02d}_{stage_id}" / "cost_benefit.xlsx"
                    _write_verified_bytes(destination, content)
                    category = "structured"
                else:
                    destination = (
                        package_root
                        / "structured"
                        / f"{ordinal:02d}_{stage_id}"
                        / f"{artifact_id}.json"
                    )
                    _write_verified_bytes(destination, content)
                    category = "structured"
                records.append(
                    _file_record(
                        package_root,
                        destination,
                        category=category,
                        stage_id=stage_id,
                        artifact_id=artifact_id,
                        source_sha256=binding.sha256,
                    )
                )

        if not inventory_rows:
            raise PackagingError("no source inventory records were available to consolidate")
        inventory_csv = package_root / "source-inventory" / "source_inventory.csv"
        inventory_xlsx = package_root / "source-inventory" / "source_inventory.xlsx"
        _write_inventory_csv(inventory_csv, inventory_rows)
        _write_inventory_xlsx(inventory_xlsx, inventory_rows)
        records.extend(
            [
                _file_record(
                    package_root, inventory_csv, category="source_inventory_consolidated"
                ),
                _file_record(
                    package_root, inventory_xlsx, category="source_inventory_consolidated"
                ),
            ]
        )

        workflow_dir = package_root / "workflow"
        packaged_workflow_manifest = workflow_dir / "workflow_manifest.json"
        packaged_contract = workflow_dir / "dar-workflow-v1.json"
        packaged_snapshot = workflow_dir / "input_snapshot.json"
        _write_verified_bytes(packaged_workflow_manifest, manifest_content)
        _write_verified_bytes(packaged_contract, contract_content)
        _write_verified_bytes(packaged_snapshot, snapshot_content)
        records.extend(
            [
                _file_record(
                    package_root,
                    packaged_workflow_manifest,
                    category="workflow",
                    artifact_id="workflow_manifest",
                    source_sha256=sha256_bytes(manifest_content),
                ),
                _file_record(
                    package_root,
                    packaged_contract,
                    category="workflow",
                    artifact_id="workflow_contract",
                    source_sha256=workflow.contract_sha256,
                ),
                _file_record(
                    package_root,
                    packaged_snapshot,
                    category="workflow",
                    artifact_id="input_snapshot",
                    source_sha256=workflow.input_snapshot_sha256,
                ),
            ]
        )

        if uploads is not None:
            packaged_upload_manifest = package_root / uploads.manifest_path
            _write_verified_bytes(packaged_upload_manifest, uploads.manifest_content)
            records.append(
                _file_record(
                    package_root,
                    packaged_upload_manifest,
                    category="input",
                    artifact_id="uploads_manifest",
                    source_sha256=uploads.manifest_sha256,
                )
            )
            for document in uploads.documents:
                packaged_content = package_root / document.content_path
                packaged_original = package_root / document.original_path
                _write_verified_bytes(packaged_content, document.content)
                _write_verified_bytes(packaged_original, document.original)
                records.extend(
                    [
                        _file_record(
                            package_root,
                            packaged_content,
                            category="input",
                            artifact_id="upload_extracted_text",
                            source_sha256=document.content_sha256,
                            input_id=document.id,
                            input_kind=document.kind,
                        ),
                        _file_record(
                            package_root,
                            packaged_original,
                            category="input",
                            artifact_id="upload_original",
                            source_sha256=document.original_sha256,
                            input_id=document.id,
                            input_kind=document.kind,
                        ),
                    ]
                )

        records.sort(key=lambda record: record["path"])
        package_manifest_value = {
            "schema_version": PACKAGE_SCHEMA,
            "package_version": workflow.workflow_version,
            "workflow_id": workflow.workflow_id,
            "workflow_version": workflow.workflow_version,
            "workflow_contract_sha256": workflow.contract_sha256,
            "workflow_manifest_sha256": sha256_bytes(manifest_content),
            "input_snapshot_sha256": workflow.input_snapshot_sha256,
            "country": workflow.country,
            "iso3": workflow.iso3,
            "lifecycle_state": contract_value["execution_policy"]["output_lifecycle_state"],
            "created_at": _created_at(created_at),
            "export_profiles": contract_value["export_profiles"],
            "upload_inputs": _upload_input_signature(uploads),
            "generating_code": {
                "file": Path(__file__).name,
                "sha256": sha256_file(Path(__file__)),
            },
            "files": records,
            "file_count": len(records),
        }
        package_manifest_path = package_root / "package-manifest.json"
        _write_verified_bytes(package_manifest_path, _canonical_json_bytes(package_manifest_value))
        validate_package_files(package_root, package_manifest_value)

        if sha256_bytes(manifest_path.read_bytes()) != sha256_bytes(manifest_content):
            raise PackagingError("workflow manifest changed while Stage 8 was packaging it")
        _zip_package(package_root, temp_zip)
        os.replace(package_root, final_package)
        published_package = True
        os.replace(temp_zip, final_zip)
        published_zip = True
        final_manifest = final_package / "package-manifest.json"
        validate_package_files(final_package, load_json_file(final_manifest))
        validate_package_zip(final_package, final_zip)
        return PackageResult(final_package, final_manifest, final_zip)
    except Exception:
        if published_zip and final_zip.exists():
            final_zip.unlink()
        if published_package and final_package.exists():
            shutil.rmtree(final_package)
        raise
    finally:
        shutil.rmtree(temp_parent, ignore_errors=True)


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="Build the canonical Stage 8 DAR package")
    parser.add_argument("--country", required=True)
    parser.add_argument("--iso", required=True, dest="iso3")
    parser.add_argument("--out", required=True)
    parser.add_argument("--workflow-manifest", required=True)
    parser.add_argument(
        "--resume",
        action="store_true",
        help="reuse a verified package or recover its missing ZIP atomically",
    )
    return parser


def main(argv: Sequence[str] | None = None) -> int:
    args = _parser().parse_args(argv)
    try:
        result = build_export_package(
            country=args.country,
            iso3=args.iso3,
            out=args.out,
            workflow_manifest=args.workflow_manifest,
            resume=args.resume,
        )
    except (PackagingError, OSError, ValueError) as error:
        print(f"!! export package failed: {error}", file=sys.stderr)
        return 1
    print(
        json.dumps(
            {
                "schema_version": "damm.workflow-event/v1",
                "event": "product_written",
                "stage_id": "export_package",
                "package_dir": str(result.package_dir),
                "package_manifest": {
                    "path": str(result.package_manifest),
                    "sha256": sha256_file(result.package_manifest),
                },
                "complete_bundle": {
                    "path": str(result.zip_path),
                    "sha256": sha256_file(result.zip_path),
                },
            },
            sort_keys=True,
            separators=(",", ":"),
        )
    )
    return 0


if __name__ == "__main__":
    sys.exit(main())
