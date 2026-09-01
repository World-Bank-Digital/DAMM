#!/usr/bin/env python3
"""Focused tests for the canonical Stage 8 export packager."""

from __future__ import annotations

import csv
import datetime
import hashlib
import io
import json
from pathlib import Path
import shutil
import tempfile
import types
import unittest
from unittest import mock
import zipfile
import xml.etree.ElementTree as ET

from openpyxl import Workbook, load_workbook

import export_package as E


HERE = Path(__file__).resolve().parent
CONTRACT_PATH = HERE.parents[2] / "workflow" / "dar-workflow-v1.json"


def digest(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def write_json(path: Path, value) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(
        json.dumps(value, sort_keys=True, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )


def write_minimal_docx(path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with zipfile.ZipFile(path, "w") as archive:
        archive.writestr(
            "[Content_Types].xml",
            "<?xml version='1.0'?><Types xmlns='http://schemas.openxmlformats.org/"
            "package/2006/content-types'><Default Extension='rels' ContentType='"
            "application/vnd.openxmlformats-package.relationships+xml'/><Default "
            "Extension='xml' ContentType='application/xml'/><Override PartName='"
            "/word/document.xml' ContentType='application/vnd.openxmlformats-officedocument."
            "wordprocessingml.document.main+xml'/></Types>",
        )
        archive.writestr(
            "_rels/.rels",
            "<?xml version='1.0'?><Relationships xmlns='http://schemas.openxmlformats.org/"
            "package/2006/relationships'><Relationship Id='rId1' Type='http://schemas."
            "openxmlformats.org/officeDocument/2006/relationships/officeDocument' "
            "Target='word/document.xml'/></Relationships>",
        )
        archive.writestr(
            "word/document.xml",
            "<?xml version='1.0'?><w:document xmlns:w='http://schemas.openxmlformats.org/"
            "wordprocessingml/2006/main'><w:body/></w:document>",
        )


def write_minimal_pdf(path: Path, text: str = "Fixture") -> None:
    escaped = text.replace("\\", "\\\\").replace("(", "\\(").replace(")", "\\)")
    stream = f"BT /F1 12 Tf 72 720 Td ({escaped}) Tj ET\n".encode("ascii", "replace")
    objects = [
        b"<< /Type /Catalog /Pages 2 0 R >>",
        b"<< /Type /Pages /Kids [3 0 R] /Count 1 >>",
        b"<< /Type /Page /Parent 2 0 R /MediaBox [0 0 612 792] "
        b"/Resources << /Font << /F1 4 0 R >> >> /Contents 5 0 R >>",
        b"<< /Type /Font /Subtype /Type1 /BaseFont /Helvetica >>",
        b"<< /Length " + str(len(stream)).encode("ascii") + b" >>\nstream\n"
        + stream + b"endstream",
    ]
    content = bytearray(b"%PDF-1.4\n% deterministic fixture\n")
    offsets = [0]
    for index, obj in enumerate(objects, 1):
        offsets.append(len(content))
        content.extend(f"{index} 0 obj\n".encode("ascii") + obj + b"\nendobj\n")
    xref = len(content)
    content.extend(f"xref\n0 {len(objects) + 1}\n".encode("ascii"))
    content.extend(b"0000000000 65535 f \n")
    for offset in offsets[1:]:
        content.extend(f"{offset:010d} 00000 n \n".encode("ascii"))
    content.extend(
        f"trailer\n<< /Size {len(objects) + 1} /Root 1 0 R >>\nstartxref\n"
        f"{xref}\n%%EOF\n".encode("ascii")
    )
    path.write_bytes(bytes(content))


class WorkflowFixture:
    def __init__(self, root: Path):
        self.root = root
        self.contract = json.loads(CONTRACT_PATH.read_text(encoding="utf-8"))
        self.contract_sha256 = digest(CONTRACT_PATH)
        self.manifest_path = root / "workflow-manifest.json"
        self.out = root / "exports" / "EGY_20260826"
        self.artifact_paths: dict[tuple[str, str], Path] = {}
        self.manifest = self._build()
        self.write_manifest()

    def _record(self, stage_id: str, key: str, path: Path, media_type: str):
        self.artifact_paths[(stage_id, key)] = path
        return {
            "key": key,
            "path": path.relative_to(self.root).as_posix(),
            "sha256": digest(path),
            "media_type": media_type,
        }

    def _workbook(self, path: Path) -> None:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Options"
        sheet.append(["Option", "Cost low", "Cost high"])
        sheet.append(["I-1", 10, 20])
        workbook.save(path)
        workbook.close()

    def _build_stage(self, ordinal: int, stage: dict):
        stage_id = stage["id"]
        stage_dir = self.root / "stages" / f"{ordinal:02d}_{stage_id}"
        stage_dir.mkdir(parents=True)
        records = []
        output_hashes = {}
        narrative = E.NARRATIVE_ARTIFACTS[stage_id]
        required_keys = list(stage["required_artifacts"])
        if stage_id == "damm_diagnostic":
            required_keys.extend(E.REQUIRED_SUPPLEMENTAL_ARTIFACTS[stage_id])
        for key in required_keys:
            if key == "stage_manifest":
                continue
            companion = None
            if key == narrative:
                if stage_id in {"damm_diagnostic", "ai_digital_agriculture"}:
                    path = stage_dir / f"{key}.html"
                    path.write_text(
                        "<!doctype html>\r\n<html><body><h1>Stage report</h1>"
                        "<p>Draft evidence.</p></body></html>\r\n",
                        encoding="utf-8",
                    )
                    media_type = "text/html"
                else:
                    path = stage_dir / f"{key}.md"
                    path.write_bytes(
                        f"# {stage['title']}\r\n\r\nDraft evidence for Egypt.\r\n".encode()
                    )
                    media_type = "text/markdown"
                    if stage_id == "country_research":
                        companion = stage_dir / f"{key}.html"
                        companion.write_text(
                            "<!doctype html>\n<html><body><h1>Country research</h1>"
                            "</body></html>\n",
                            encoding="utf-8",
                        )
            elif key == "cost_benefit_workbook":
                path = stage_dir / "cost_benefit.xlsx"
                self._workbook(path)
                media_type = (
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                )
            else:
                path = stage_dir / f"{key}.json"
                if key == "engine_input":
                    value = {
                        "1.1": {
                            "value": "Exact machine-filled fixture row",
                            "cls": "Documented",
                            "level": 3,
                            "year": 2026,
                            "src": "Fixture source",
                            "tier": "T1",
                            "url": "https://example.test/engine-input",
                        }
                    }
                elif key == "source_inventory":
                    value = [
                        {
                            "ref": f"SRC-{ordinal}",
                            "title": f"Source for {stage_id}",
                            "url": f"https://example.test/{stage_id}",
                            "metadata": {"tier": "A", "ordinal": ordinal},
                        }
                    ]
                else:
                    value = {
                        "schema_version": f"damm.{key}/v1",
                        "country": "Egypt",
                        "iso3": "EGY",
                        "stage_id": stage_id,
                    }
                write_json(path, value)
                media_type = "application/json"
            record = self._record(stage_id, key, path, media_type)
            records.append(record)
            output_hashes[key] = record["sha256"]
            if companion is not None:
                companion_record = self._record(
                    stage_id, key, companion, "text/html"
                )
                records.append(companion_record)
                output_hashes[key] = [record["sha256"], companion_record["sha256"]]

        stage_manifest_path = stage_dir / "stage_manifest.json"
        write_json(
            stage_manifest_path,
            {
                "schema_version": "damm.stage-manifest/v1",
                "workflow_id": "dar-canonical-v1",
                "workflow_version": self.contract["workflow_version"],
                "stage_id": stage_id,
                "input_hashes": {},
                "output_hashes": output_hashes,
                "source_inventory": "source_inventory"
                if "source_inventory" in stage["required_artifacts"]
                else [],
                "quality_checks": [{"id": "fixture", "ok": True}],
                "execution_mode": "test_fixture",
                "spend_usd": 0,
                "status": "complete",
            },
        )
        records.append(
            self._record(
                stage_id, "stage_manifest", stage_manifest_path, "application/json"
            )
        )
        return {"id": stage_id, "status": "complete", "artifacts": records}

    def _build(self):
        snapshot = self.root / "input-snapshot.json"
        self.snapshot_path = snapshot
        write_json(
            snapshot,
            {
                "schema_version": "damm.workflow-input-snapshot/v1",
                "country": "Egypt",
                "iso3": "EGY",
                "contract_sha256": self.contract_sha256,
                "uploads_manifest": None,
                "ceiling_usd": None,
                "vendor": None,
            },
        )
        stages = [
            self._build_stage(ordinal, stage)
            for ordinal, stage in enumerate(self.contract["stages"][:7], 1)
        ]
        stages.append({"id": "export_package", "status": "running", "artifacts": []})
        return {
            "schema_version": "damm.workflow-run/v1",
            "workflow_id": "dar-canonical-v1",
            "workflow_version": self.contract["workflow_version"],
            "contract_sha256": self.contract_sha256,
            "country": "Egypt",
            "iso3": "EGY",
            "status": "running",
            "current_stage": "export_package",
            "input_snapshot": {
                "path": snapshot.relative_to(self.root).as_posix(),
                "sha256": digest(snapshot),
            },
            "uploads_manifest": None,
            "ceiling_usd": None,
            "vendor": None,
            "stages": stages,
        }

    def add_upload(
        self,
        *,
        document_id: str = "ttl-country-1",
        kind: str = "country_context_documents",
        original: bytes = b"%PDF-1.4\noriginal fixture\n",
        extracted_text: str = "Egypt digital agriculture evidence.\n",
    ) -> dict:
        content_path = self.root / f"inputs/upload-content/{document_id}.txt"
        original_path = self.root / f"inputs/upload-originals/{document_id}.pdf"
        content_path.parent.mkdir(parents=True, exist_ok=True)
        original_path.parent.mkdir(parents=True, exist_ok=True)
        content_path.write_text(extracted_text, encoding="utf-8")
        original_path.write_bytes(original)
        document = {
            "id": document_id,
            "kind": kind,
            "original_filename": "country-evidence.pdf",
            "content_path": content_path.relative_to(self.root).as_posix(),
            "content_sha256": digest(content_path),
            "content_media_type": "text/plain",
            "original_path": original_path.relative_to(self.root).as_posix(),
            "original_sha256": digest(original_path),
            "original_size_bytes": len(original),
            "metadata": {
                "extracted_characters": len(extracted_text),
                "app_upload_kind": kind,
                "source_mime_type": "application/pdf",
                "uploaded_at": "2026-08-26T10:00:00Z",
                "uploaded_by": "fixture-ttl",
                "extraction_status": "extracted",
            },
        }
        uploads_path = self.root / "inputs/uploads-manifest.json"
        write_json(
            uploads_path,
            {"schema_version": "damm.uploads-manifest/v1", "documents": [document]},
        )
        uploads_record = {
            "path": uploads_path.relative_to(self.root).as_posix(),
            "sha256": digest(uploads_path),
            "document_count": 1,
        }
        self.manifest["uploads_manifest"] = uploads_record
        write_json(
            self.snapshot_path,
            {
                "schema_version": "damm.workflow-input-snapshot/v1",
                "country": "Egypt",
                "iso3": "EGY",
                "contract_sha256": self.contract_sha256,
                "uploads_manifest": uploads_record,
                "ceiling_usd": None,
                "vendor": None,
            },
        )
        self.manifest["input_snapshot"]["sha256"] = digest(self.snapshot_path)
        self.write_manifest()
        return document

    def rebind_upload_manifest(self) -> None:
        uploads_path = self.root / "inputs/uploads-manifest.json"
        envelope = json.loads(uploads_path.read_text(encoding="utf-8"))
        uploads_record = {
            "path": uploads_path.relative_to(self.root).as_posix(),
            "sha256": digest(uploads_path),
            "document_count": len(envelope["documents"]),
        }
        self.manifest["uploads_manifest"] = uploads_record
        snapshot = json.loads(self.snapshot_path.read_text(encoding="utf-8"))
        snapshot["uploads_manifest"] = uploads_record
        write_json(self.snapshot_path, snapshot)
        self.manifest["input_snapshot"]["sha256"] = digest(self.snapshot_path)
        self.write_manifest()

    def write_manifest(self):
        write_json(self.manifest_path, self.manifest)

    def artifact_record(self, stage_id: str, key: str):
        stage = next(stage for stage in self.manifest["stages"] if stage["id"] == stage_id)
        return next(record for record in stage["artifacts"] if record["key"] == key)


class FakeConverters:
    def __init__(self):
        self.pandoc_calls: list[tuple[str, str]] = []
        self.pdf_calls: list[tuple[str, str]] = []
        self.docx_sources: list[str] = []

    def pandoc(self, source: Path, target: Path):
        self.pandoc_calls.append((source.suffix, target.suffix))
        target.parent.mkdir(parents=True, exist_ok=True)
        if target.suffix == ".md":
            target.write_text("# Converted HTML\n\nDraft evidence.\n", encoding="utf-8")
        elif target.suffix == ".html":
            target.write_text(
                "<!doctype html>\n<html><body><p>Draft evidence.</p></body></html>\n",
                encoding="utf-8",
            )
        elif target.suffix == ".docx":
            self.docx_sources.append(source.read_text(encoding="utf-8"))
            write_minimal_docx(target)
        else:
            raise AssertionError(f"unexpected pandoc target: {target}")

    def pdf(self, source: Path, target: Path):
        self.pdf_calls.append((source.suffix, target.suffix))
        write_minimal_pdf(target)


class ExportPackageTest(unittest.TestCase):
    def setUp(self):
        self.temp = tempfile.TemporaryDirectory(prefix="damm-export-test-")
        self.addCleanup(self.temp.cleanup)
        self.root = Path(self.temp.name)
        self.fixture = WorkflowFixture(self.root)
        self.converters = FakeConverters()

    def build(self, *, resume: bool = False):
        return E.build_export_package(
            country="Egypt",
            iso3="EGY",
            out=self.fixture.out,
            workflow_manifest=self.fixture.manifest_path,
            contract_path=CONTRACT_PATH,
            pandoc_converter=self.converters.pandoc,
            pdf_converter=self.converters.pdf,
            created_at="2026-08-26T12:00:00Z",
            resume=resume,
        )

    def test_consolidated_inventory_xlsx_is_stable_for_package_timestamp(self):
        rows = [{
            "stage_id": "country_research",
            "source_inventory_artifact": "source_inventory",
            "source_index": 1,
            "ref": "SRC-001",
            "title": "Country strategy",
        }]
        first = self.root / "first-inventory.xlsx"
        second = self.root / "second-inventory.xlsx"
        created_at = "2026-08-26T12:00:00Z"

        def openpyxl_clock(second_value):
            return types.SimpleNamespace(
                datetime=types.SimpleNamespace(
                    now=lambda tz=None: datetime.datetime(
                        2026, 9, 2, 0, 0, second_value, tzinfo=tz
                    )
                ),
                timezone=datetime.timezone,
            )

        first_clock = openpyxl_clock(2)
        second_clock = openpyxl_clock(8)
        with (
            mock.patch("openpyxl.packaging.core.datetime", first_clock),
            mock.patch("openpyxl.writer.excel.datetime", first_clock),
        ):
            E._write_inventory_xlsx(first, rows, created_at)
        with (
            mock.patch("openpyxl.packaging.core.datetime", second_clock),
            mock.patch("openpyxl.writer.excel.datetime", second_clock),
        ):
            E._write_inventory_xlsx(second, rows, created_at)

        self.assertEqual(first.read_bytes(), second.read_bytes())
        with zipfile.ZipFile(first) as archive:
            self.assertTrue(all(
                item.date_time == (1980, 1, 1, 0, 0, 0)
                for item in archive.infolist()
            ))
            core = archive.read("docProps/core.xml").decode("utf-8")
        self.assertIn(
            ">2026-08-26T12:00:00Z</dcterms:created>", core
        )
        self.assertIn(
            ">2026-08-26T12:00:00Z</dcterms:modified>", core
        )

    def test_inventory_xlsx_normalizer_rejects_ambiguous_core_metadata(self):
        def archive(core_properties):
            payload = io.BytesIO()
            with zipfile.ZipFile(payload, "w") as workbook:
                workbook.writestr("xl/workbook.xml", b"<workbook/>")
                if core_properties is not None:
                    workbook.writestr("docProps/core.xml", core_properties)
            return payload.getvalue()

        duplicate_modified = (
            b'<core xmlns:dcterms="urn:test">'
            b"<dcterms:created>only</dcterms:created>"
            b"<dcterms:modified>first</dcterms:modified>"
            b"<dcterms:modified>second</dcterms:modified>"
            b"</core>"
        )
        for label, raw in (
            ("missing core", archive(None)),
            ("duplicate timestamp", archive(duplicate_modified)),
        ):
            with self.subTest(label=label):
                with self.assertRaisesRegex(E.PackagingError, "unique"):
                    E._stable_xlsx_bytes(raw, b"2026-08-26T12:00:00Z")

    def test_legacy_html_fragment_is_wrapped_offline_instead_of_rejected(self):
        fragment = (
            b'<meta charset="utf-8"><title>Legacy diagnostic</title>'
            b'<link rel="stylesheet" href="https://fonts.example.test/report.css">'
            b'<style>.legacy{color:#17322a}</style><div class="legacy">Evidence</div>'
        )
        root = self.root / "legacy-conversion"
        root.mkdir()
        outputs = E._convert_narrative(
            source_content=fragment,
            source_suffix=".html",
            destination_stem=root / "diagnostic_report",
            work_dir=root,
            pandoc_converter=self.converters.pandoc,
            pdf_converter=self.converters.pdf,
        )

        html_path = next(path for path in outputs if path.suffix == ".html")
        rendered = html_path.read_text(encoding="utf-8")
        self.assertTrue(rendered.startswith("<!doctype html>"))
        self.assertIn('<html lang="en">', rendered)
        self.assertIn('<div class="legacy">Evidence</div>', rendered)
        self.assertNotIn("fonts.example.test", rendered)
        self.assertNotIn('<link rel="stylesheet"', rendered)

    def test_office_conversion_retains_consulting_header_and_lifecycle(self):
        source = (
            b'<!doctype html><html lang="en"><head><meta charset="utf-8"></head>'
            b'<body><article class="page"><header class="masthead">'
            b'<h1>Investment options</h1><p>Nigeria</p></header>'
            b'<div class="lifecycle"><strong>Lifecycle</strong>'
            b'<span>Draft - no financing decision</span></div>'
            b'<main><section><h2>Executive frame</h2>'
            b'<svg viewBox="0 0 10 10" role="img" aria-label="Evidence mix">'
            b'<rect width="10" height="10" fill="#245844"/></svg></section></main>'
            b'<footer>Evidence-backed working paper</footer></article></body></html>'
        )
        root = self.root / "office-conversion"
        root.mkdir()

        E._convert_narrative(
            source_content=source,
            source_suffix=".html",
            destination_stem=root / "investment_options_report",
            work_dir=root,
            pandoc_converter=self.converters.pandoc,
            pdf_converter=self.converters.pdf,
        )

        office_source = self.converters.docx_sources[-1]
        self.assertIn("Investment options", office_source)
        self.assertIn("Nigeria", office_source)
        self.assertIn("Draft - no financing decision", office_source)
        self.assertNotIn("<header", office_source)
        self.assertNotIn("<main", office_source)
        self.assertNotIn("<footer", office_source)
        self.assertNotIn("<svg", office_source)
        self.assertIn("data:image/svg+xml;base64,", office_source)
        self.assertIn('alt="Evidence mix"', office_source)

    def test_office_conversion_prepares_consulting_blocks_for_word(self):
        source = (
            b'<!doctype html><html lang="en"><head><title>Stage 7 roadmap - Nigeria</title>'
            b'</head><body><article class="page"><header class="masthead">'
            b'<h1>Stage 7 roadmap</h1><p class="country">Nigeria</p></header><main>'
            b'<div class="notice proposal"><strong>Publication hold</strong>'
            b'SIMULATED - NOT ACCEPTANCE EVIDENCE</div>'
            b'<div class="cards"><div class="card"><div class="value">6</div>'
            b'<div class="label">Investment options</div>'
            b'<div class="note">No financing decision</div></div>'
            b'<div class="card"><div class="value">100%</div>'
            b'<div class="label">Figure traceability</div></div></div>'
            b'<div class="table-wrap short-table"><table><caption class="sr-only">'
            b'Table columns: Option; Status</caption><thead><tr><th>Option</th>'
            b'<th>Status</th></tr></thead><tbody><tr><td>INV-1</td>'
            b'<td>Review</td></tr></tbody></table></div>'
            b'<div class="table-wrap short-table"><table><caption class="sr-only">'
            b'Table columns: Empty</caption><thead><tr><th>Empty</th></tr></thead>'
            b'<tbody></tbody></table></div></main></article></body></html>'
        )
        root = self.root / "office-consulting-blocks"
        root.mkdir()

        E._convert_narrative(
            source_content=source,
            source_suffix=".html",
            destination_stem=root / "draft_dar_report",
            work_dir=root,
            pandoc_converter=self.converters.pandoc,
            pdf_converter=self.converters.pdf,
        )

        office_source = self.converters.docx_sources[-1]
        self.assertNotIn("<title>", office_source)
        self.assertEqual(office_source.count("Stage 7 roadmap"), 1)
        self.assertIn("<strong>Publication hold</strong><br>", office_source)
        self.assertIn('<table class="office-metrics">', office_source)
        self.assertIn("<th>Metric</th><th>Value</th><th>Context</th>", office_source)
        self.assertIn(
            "<td>Investment options</td><td>6</td>"
            "<td>No financing decision</td>",
            office_source,
        )
        self.assertNotIn('class="cards"', office_source)
        self.assertNotIn('class="card"', office_source)
        self.assertNotIn("Table columns:", office_source)
        self.assertNotIn("<th>Empty</th>", office_source)

    def test_conversion_validation_rejects_structural_stubs(self):
        incomplete_docx = self.root / "incomplete.docx"
        with zipfile.ZipFile(incomplete_docx, "w") as archive:
            archive.writestr("[Content_Types].xml", "<Types/>")
            archive.writestr("word/document.xml", "<w:document/>")
        with self.assertRaisesRegex(E.PackagingError, "relationship"):
            E._require_conversion_output(incomplete_docx, "DOCX")

        incomplete_pdf = self.root / "incomplete.pdf"
        incomplete_pdf.write_bytes(b"%PDF-1.4\n% header only\n")
        with self.assertRaisesRegex(E.PackagingError, "complete PDF"):
            E._require_conversion_output(incomplete_pdf, "PDF")

    @unittest.skipUnless(shutil.which("pandoc"), "pandoc is not installed")
    def test_real_pandoc_docx_retains_visible_report_identity(self):
        raw_html = (
            '<!doctype html><html><head><title>Stage 6 appraisal - Nigeria</title></head>'
            '<body><article><header><h1>Stage 6 appraisal</h1>'
            '<p>Nigeria</p></header><div><strong>Lifecycle</strong>'
            '<span>Draft - no financing decision</span></div><main>'
            '<h2>Executive frame</h2><div class="notice"><strong>Decision boundary</strong>'
            'Human review is required.</div><div class="cards"><div class="card">'
            '<div class="value">3</div><div class="label">Options</div>'
            '<div class="note">Review required</div></div></div>'
            '<svg viewBox="0 0 10 10" role="img" '
            'aria-label="Evidence mix"><rect width="10" height="10" '
            'fill="#245844"/></svg><table><caption class="sr-only">'
            'Table columns: Option; Status</caption><thead><tr><th>Option</th>'
            '<th>Status</th></tr></thead>'
            '<tbody><tr><td>INV-1</td><td>Review</td></tr></tbody></table>'
            '<table><thead><tr><th>Empty register</th></tr></thead><tbody></tbody></table>'
            '</main><footer>Evidence-backed working paper'
            '</footer></article></body></html>'
        )
        destination = self.root / "office-source"

        outputs = E._convert_narrative(
            source_content=raw_html.encode("utf-8"),
            source_suffix=".html",
            destination_stem=destination,
            work_dir=self.root,
            pandoc_converter=E.default_pandoc_converter,
            pdf_converter=lambda _source, target: write_minimal_pdf(target),
        )
        target = next(path for path in outputs if path.suffix == ".docx")
        E._require_conversion_output(target, "DOCX")

        with zipfile.ZipFile(target) as archive:
            document = ET.fromstring(archive.read("word/document.xml"))
            styles = archive.read("word/styles.xml")
            media = [name for name in archive.namelist() if name.startswith("word/media/")]
        visible_text = " ".join(text.strip() for text in document.itertext() if text.strip())
        self.assertEqual(visible_text.count("Stage 6 appraisal"), 1)
        self.assertIn("Stage 6 appraisal", visible_text)
        self.assertIn("Nigeria", visible_text)
        self.assertIn("Draft - no financing decision", visible_text)
        self.assertNotIn("Table columns:", visible_text)
        self.assertNotIn("Empty register", visible_text)
        self.assertIn(b"245844", styles)
        self.assertTrue(media, "the inline report visualization must be embedded in DOCX")
        notice = next(
            paragraph
            for paragraph in document.findall(f".//{{{E._WORD_NS}}}p")
            if "Decision boundary" in "".join(paragraph.itertext())
        )
        self.assertIsNotNone(notice.find(f".//{{{E._WORD_NS}}}br"))
        tables = document.findall(f".//{{{E._WORD_NS}}}tbl")
        self.assertEqual(len(tables), 2)
        for table in tables:
            header = table.find(f"./{{{E._WORD_NS}}}tr")
            self.assertIsNotNone(header)
            for cell in header.findall(f"./{{{E._WORD_NS}}}tc"):
                shading = cell.find(
                    f"./{{{E._WORD_NS}}}tcPr/{{{E._WORD_NS}}}shd"
                )
                self.assertIsNotNone(shading)
                self.assertEqual(shading.get(f"{{{E._WORD_NS}}}fill"), "17322A")
                for run in cell.findall(f".//{{{E._WORD_NS}}}r"):
                    color = run.find(
                        f"./{{{E._WORD_NS}}}rPr/{{{E._WORD_NS}}}color"
                    )
                    self.assertIsNotNone(color)
                    self.assertEqual(color.get(f"{{{E._WORD_NS}}}val"), "FFFFFF")
        section = document.find(f".//{{{E._WORD_NS}}}sectPr")
        self.assertIsNotNone(section)
        page_size = section.find(f"./{{{E._WORD_NS}}}pgSz")
        self.assertIsNotNone(page_size)
        self.assertEqual(page_size.get(f"{{{E._WORD_NS}}}w"), "11906")
        self.assertEqual(page_size.get(f"{{{E._WORD_NS}}}h"), "16838")
        margins = section.find(f"./{{{E._WORD_NS}}}pgMar")
        self.assertIsNotNone(margins)
        for edge in ("top", "right", "bottom", "left"):
            self.assertEqual(margins.get(f"{{{E._WORD_NS}}}{edge}"), "720")
        markdown = next(path for path in outputs if path.suffix == ".md").read_text("utf-8")
        self.assertIn("Nigeria", markdown)
        self.assertIn("Draft - no financing decision", markdown)
        self.assertIn("Evidence-backed working paper", markdown)

    def test_pure_manifest_validation_returns_all_canonical_bindings(self):
        validated = E.validate_workflow_manifest(
            self.fixture.manifest,
            self.fixture.contract,
            country="Egypt",
            iso3="egy",
            contract_sha256=self.fixture.contract_sha256,
        )

        self.assertEqual(validated.workflow_version, "1.0.0")
        self.assertEqual(tuple(validated.artifacts), E.EXPECTED_STAGE_IDS[:7])
        self.assertEqual(
            Path(validated.artifacts["country_research"]["country_research_report"].path).suffix,
            ".html",
        )
        self.assertEqual(
            len(
                validated.artifact_records["country_research"][
                    "country_research_report"
                ]
            ),
            2,
        )
        for stage in self.fixture.contract["stages"][:7]:
            expected = set(stage["required_artifacts"])
            expected.update(E.REQUIRED_SUPPLEMENTAL_ARTIFACTS.get(stage["id"], ()))
            self.assertEqual(
                set(validated.artifacts[stage["id"]]), expected
            )

    def test_manifest_validation_rejects_missing_narrative_and_wrong_run_identity(self):
        cases = []
        missing = json.loads(json.dumps(self.fixture.manifest))
        missing["stages"][0]["artifacts"] = [
            item
            for item in missing["stages"][0]["artifacts"]
            if item["key"] != "diagnostic_report"
        ]
        cases.append((missing, "missing required artifacts: diagnostic_report"))
        missing_engine_input = json.loads(json.dumps(self.fixture.manifest))
        missing_engine_input["stages"][0]["artifacts"] = [
            item
            for item in missing_engine_input["stages"][0]["artifacts"]
            if item["key"] != "engine_input"
        ]
        cases.append((missing_engine_input, "missing required artifacts: engine_input"))
        wrong_contract = json.loads(json.dumps(self.fixture.manifest))
        wrong_contract["contract_sha256"] = "0" * 64
        cases.append((wrong_contract, "not bound to the canonical contract bytes"))
        wrong_country = json.loads(json.dumps(self.fixture.manifest))
        wrong_country["country"] = "Jordan"
        cases.append((wrong_country, "country does not match"))
        wrong_stage = json.loads(json.dumps(self.fixture.manifest))
        wrong_stage["stages"][4]["status"] = "running"
        cases.append((wrong_stage, "strategic_foresight is not complete"))
        traversal = json.loads(json.dumps(self.fixture.manifest))
        traversal["stages"][0]["artifacts"][0]["path"] = "../outside.json"
        cases.append((traversal, "not a safe relative path"))

        for value, message in cases:
            with self.subTest(message=message):
                with self.assertRaisesRegex(E.ManifestValidationError, message):
                    E.validate_workflow_manifest(
                        value,
                        self.fixture.contract,
                        country="Egypt",
                        iso3="EGY",
                        contract_sha256=self.fixture.contract_sha256,
                    )

    def test_full_build_exports_every_profile_and_hash_bound_bundle(self):
        result = self.build()

        self.assertEqual(
            result.package_dir.name, "EGY_20260826_dar_package_v1.0.0"
        )
        self.assertEqual(result.zip_path, Path(f"{self.fixture.out}_dar_package.zip"))
        self.assertTrue(result.zip_path.is_file())
        manifest = json.loads(result.package_manifest.read_text(encoding="utf-8"))
        self.assertEqual(manifest["schema_version"], "damm.dar-package/v1")
        self.assertEqual(manifest["lifecycle_state"], "draft")
        self.assertEqual(manifest["created_at"], "2026-08-26T12:00:00Z")
        self.assertEqual(manifest["export_profiles"], self.fixture.contract["export_profiles"])
        self.assertEqual(manifest["file_count"], 58)
        self.assertEqual(len(self.converters.pandoc_calls), 14)
        self.assertEqual(len(self.converters.pdf_calls), 7)

        narrative_files = list((result.package_dir / "narratives").rglob("*"))
        narrative_files = [path for path in narrative_files if path.is_file()]
        self.assertEqual(len(narrative_files), 28)
        for suffix in (".md", ".html", ".docx", ".pdf"):
            self.assertEqual(sum(path.suffix == suffix for path in narrative_files), 7)
        diagnostic_md = (
            result.package_dir
            / "narratives/01_damm_diagnostic/diagnostic_report.md"
        ).read_bytes()
        self.assertNotIn(b"\r", diagnostic_md)
        self.assertTrue(diagnostic_md.endswith(b"\n"))
        country_html = (
            result.package_dir
            / "narratives/02_country_research/country_research_report.html"
        ).read_bytes()
        self.assertEqual(
            country_html,
            self.fixture.artifact_paths[
                ("country_research", "country_research_report")
            ].read_bytes(),
        )
        self.assertIn(b"<h1>Country research</h1>", country_html)

        csv_path = result.package_dir / "source-inventory/source_inventory.csv"
        with csv_path.open(encoding="utf-8", newline="") as handle:
            rows = list(csv.DictReader(handle))
        self.assertEqual(len(rows), 6)
        self.assertEqual(
            {row["stage_id"] for row in rows}, set(E.SOURCE_INVENTORY_STAGES)
        )
        self.assertIn('"tier":"A"', rows[0]["metadata"])
        xlsx_path = result.package_dir / "source-inventory/source_inventory.xlsx"
        workbook = load_workbook(xlsx_path, read_only=True, data_only=False)
        self.assertEqual(workbook["Sources"].max_row, 7)
        workbook.close()
        cost_benefit = (
            result.package_dir
            / "structured/06_investment_options/cost_benefit.xlsx"
        )
        self.assertTrue(cost_benefit.is_file())

        engine_source = self.fixture.artifact_paths[("damm_diagnostic", "engine_input")]
        engine_records = [
            record
            for record in manifest["files"]
            if record.get("stage_id") == "damm_diagnostic"
            and record.get("artifact_id") == "engine_input"
        ]
        self.assertEqual(len(engine_records), 1)
        engine_record = engine_records[0]
        self.assertEqual(engine_record["category"], "structured")
        self.assertEqual(engine_record["sha256"], digest(engine_source))
        self.assertEqual(engine_record["source_sha256"], digest(engine_source))
        packaged_engine = result.package_dir / engine_record["path"]
        self.assertEqual(packaged_engine.read_bytes(), engine_source.read_bytes())

        E.validate_package_files(result.package_dir, manifest)
        for record in manifest["files"]:
            payload = result.package_dir / record["path"]
            self.assertEqual(record["sha256"], digest(payload))
            self.assertEqual(record["bytes"], payload.stat().st_size)
        with zipfile.ZipFile(result.zip_path) as archive:
            self.assertIsNone(archive.testzip())
            root = result.package_dir.name + "/"
            self.assertIn(root + "package-manifest.json", archive.namelist())
            self.assertEqual(len(archive.namelist()), manifest["file_count"] + 1)
            self.assertEqual(
                archive.read(root + engine_record["path"]), engine_source.read_bytes()
            )

    def test_downloadable_source_inventories_neutralize_formulas(self):
        stage_id = "country_research"
        inventory_path = self.fixture.artifact_paths[(stage_id, "source_inventory")]
        dangerous = {
            "ref": "SRC-FORMULA",
            "title": '=HYPERLINK("https://evil.example","open")',
            "publisher": "  +SUM(A1:A2)",
            "source": "\t-10+20",
            "comment": "\n@SUM(A1:A2)",
            "negative_amount": -42.5,
        }
        write_json(inventory_path, [dangerous])
        inventory_digest = digest(inventory_path)
        self.fixture.artifact_record(stage_id, "source_inventory")["sha256"] = (
            inventory_digest
        )
        stage_manifest_path = self.fixture.artifact_paths[(stage_id, "stage_manifest")]
        stage_manifest = json.loads(stage_manifest_path.read_text(encoding="utf-8"))
        stage_manifest["output_hashes"]["source_inventory"] = inventory_digest
        write_json(stage_manifest_path, stage_manifest)
        self.fixture.artifact_record(stage_id, "stage_manifest")["sha256"] = digest(
            stage_manifest_path
        )
        self.fixture.write_manifest()

        result = self.build()

        with (
            result.package_dir / "source-inventory/source_inventory.csv"
        ).open(encoding="utf-8", newline="") as handle:
            csv_row = next(
                candidate
                for candidate in csv.DictReader(handle)
                if candidate["stage_id"] == stage_id
            )
        for field in ("title", "publisher", "source", "comment"):
            with self.subTest(format="csv", field=field):
                self.assertEqual(csv_row[field], "'" + dangerous[field])
        self.assertEqual(csv_row["negative_amount"], "-42.5")

        workbook = load_workbook(
            result.package_dir / "source-inventory/source_inventory.xlsx",
            read_only=True,
            data_only=False,
        )
        sheet = workbook["Sources"]
        headers = {cell.value: index for index, cell in enumerate(sheet[1], 1)}
        row = next(
            candidate
            for candidate in sheet.iter_rows(min_row=2)
            if candidate[headers["stage_id"] - 1].value == stage_id
        )
        for field in ("title", "publisher", "source", "comment"):
            with self.subTest(format="xlsx", field=field):
                cell = row[headers[field] - 1]
                self.assertEqual(cell.value, "'" + dangerous[field])
                self.assertNotEqual(cell.data_type, "f")
        negative = row[headers["negative_amount"] - 1]
        self.assertEqual(negative.value, -42.5)
        self.assertEqual(negative.data_type, "n")
        workbook.close()

    def test_upload_manifest_extracted_text_and_original_are_packaged_with_provenance(self):
        document = self.fixture.add_upload(
            original=b"\x00fixture-original-binary\xff",
            extracted_text="Country evidence with café and soil data.\n",
        )

        result = self.build()
        package_manifest = json.loads(result.package_manifest.read_text(encoding="utf-8"))
        signature = package_manifest["upload_inputs"]
        self.assertEqual(signature["manifest_sha256"], self.fixture.manifest["uploads_manifest"]["sha256"])
        self.assertEqual(signature["document_count"], 1)
        self.assertEqual(signature["documents"][0]["id"], document["id"])
        expected = {
            "inputs/uploads-manifest.json": "uploads_manifest",
            document["content_path"]: "upload_extracted_text",
            document["original_path"]: "upload_original",
        }
        records = {record["path"]: record for record in package_manifest["files"]}
        for relative, artifact_id in expected.items():
            with self.subTest(relative=relative):
                record = records[relative]
                self.assertEqual(record["category"], "input")
                self.assertEqual(record["artifact_id"], artifact_id)
                self.assertEqual(record["sha256"], record["source_sha256"])
                self.assertEqual(
                    (result.package_dir / relative).read_bytes(),
                    (self.root / relative).read_bytes(),
                )
        self.assertEqual(records[document["original_path"]]["input_id"], document["id"])
        self.assertEqual(records[document["original_path"]]["input_kind"], document["kind"])
        with zipfile.ZipFile(result.zip_path) as archive:
            names = set(archive.namelist())
            for relative in expected:
                self.assertIn(f"{result.package_dir.name}/{relative}", names)

    def test_tampered_or_missing_frozen_upload_payload_fails_before_conversion(self):
        document = self.fixture.add_upload()
        original = self.root / document["original_path"]
        original.write_bytes(original.read_bytes() + b"tamper")
        with self.assertRaisesRegex(E.PackagingError, "SHA-256 mismatch"):
            self.build()
        self.assertEqual(self.converters.pandoc_calls, [])

        # A missing original is independently terminal.
        original.write_bytes(b"%PDF-1.4\noriginal fixture\n")
        original.unlink()
        with self.assertRaisesRegex(E.PackagingError, "missing"):
            self.build()
        self.assertEqual(self.converters.pandoc_calls, [])

        # Restore the original, then prove a missing extracted file is also terminal.
        original.write_bytes(b"%PDF-1.4\noriginal fixture\n")
        (self.root / document["content_path"]).unlink()
        with self.assertRaisesRegex(E.PackagingError, "missing"):
            self.build()
        self.assertEqual(self.converters.pandoc_calls, [])

    def test_hash_bound_upload_path_traversal_and_symlink_are_rejected(self):
        document = self.fixture.add_upload()
        uploads_path = self.root / "inputs/uploads-manifest.json"
        envelope = json.loads(uploads_path.read_text(encoding="utf-8"))
        outside = self.root / "outside.pdf"
        outside.write_bytes(b"outside")
        envelope["documents"][0]["original_path"] = "inputs/upload-originals/../../outside.pdf"
        envelope["documents"][0]["original_sha256"] = digest(outside)
        envelope["documents"][0]["original_size_bytes"] = outside.stat().st_size
        write_json(uploads_path, envelope)
        self.fixture.rebind_upload_manifest()
        with self.assertRaisesRegex(E.ManifestValidationError, "not a safe relative path"):
            self.build()

        # A lexically valid path that resolves through a symlink is also forbidden.
        symlink = self.root / "inputs/upload-originals/linked.pdf"
        symlink.symlink_to(outside)
        envelope["documents"][0]["original_path"] = symlink.relative_to(self.root).as_posix()
        write_json(uploads_path, envelope)
        self.fixture.rebind_upload_manifest()
        with self.assertRaisesRegex(E.PackagingError, "symbolic link"):
            self.build()
        self.assertEqual(self.converters.pandoc_calls, [])

    def test_artifact_byte_mutation_fails_before_conversion_and_publishes_nothing(self):
        path = self.fixture.artifact_paths[("country_research", "country_evidence_data")]
        path.write_text('{"tampered":true}\n', encoding="utf-8")

        with self.assertRaisesRegex(E.PackagingError, "SHA-256 mismatch"):
            self.build()

        self.assertEqual(self.converters.pandoc_calls, [])
        self.assertFalse(Path(f"{self.fixture.out}_dar_package.zip").exists())
        self.assertFalse(
            self.fixture.out.parent.joinpath(
                "EGY_20260826_dar_package_v1.0.0"
            ).exists()
        )

    def test_engine_input_mutation_or_stale_stage_binding_is_terminal(self):
        stage_id = "damm_diagnostic"
        key = "engine_input"
        path = self.fixture.artifact_paths[(stage_id, key)]
        path.write_text('{"1.1":{"tampered":true}}\n', encoding="utf-8")

        with self.assertRaisesRegex(E.PackagingError, "SHA-256 mismatch"):
            self.build()
        self.assertEqual(self.converters.pandoc_calls, [])

        self.fixture.artifact_record(stage_id, key)["sha256"] = digest(path)
        self.fixture.write_manifest()
        with self.assertRaisesRegex(
            E.ManifestValidationError, "output_hashes does not bind engine_input"
        ):
            self.build()
        self.assertEqual(self.converters.pandoc_calls, [])

    def test_rehashed_artifact_still_fails_stale_stage_manifest_binding(self):
        stage_id = "country_research"
        key = "country_evidence_data"
        path = self.fixture.artifact_paths[(stage_id, key)]
        write_json(path, {"tampered": True, "country": "Egypt"})
        self.fixture.artifact_record(stage_id, key)["sha256"] = digest(path)
        self.fixture.write_manifest()

        with self.assertRaisesRegex(
            E.ManifestValidationError, "output_hashes does not bind country_evidence_data"
        ):
            self.build()

        self.assertEqual(self.converters.pandoc_calls, [])

    def test_converter_failure_is_terminal_and_leaves_no_publication(self):
        def failing_pandoc(_source: Path, _target: Path):
            raise E.PackagingError("injected pandoc failure")

        with self.assertRaisesRegex(E.PackagingError, "injected pandoc failure"):
            E.build_export_package(
                country="Egypt",
                iso3="EGY",
                out=self.fixture.out,
                workflow_manifest=self.fixture.manifest_path,
                contract_path=CONTRACT_PATH,
                pandoc_converter=failing_pandoc,
                pdf_converter=self.converters.pdf,
            )

        self.assertFalse(Path(f"{self.fixture.out}_dar_package.zip").exists())
        self.assertFalse(
            self.fixture.out.parent.joinpath(
                "EGY_20260826_dar_package_v1.0.0"
            ).exists()
        )

    def test_missing_real_converters_fail_explicitly(self):
        source = self.root / "source.md"
        source.write_text("# Source\n", encoding="utf-8")
        with mock.patch.object(E.shutil, "which", return_value=None):
            with self.assertRaisesRegex(E.PackagingError, "pandoc.*not found"):
                E.default_pandoc_converter(source, self.root / "target.docx")
            with self.assertRaisesRegex(E.PackagingError, "soffice.*not found"):
                E.default_pdf_converter(self.root / "target.docx", self.root / "target.pdf")

    def test_package_manifest_rejects_post_build_tampering(self):
        result = self.build()
        manifest = json.loads(result.package_manifest.read_text(encoding="utf-8"))
        target = result.package_dir / manifest["files"][0]["path"]
        target.write_bytes(target.read_bytes() + b"tamper")

        with self.assertRaisesRegex(E.PackagingError, "byte count does not match"):
            E.validate_package_files(result.package_dir, manifest)

    def test_resume_reuses_verified_complete_package_without_conversion(self):
        first = self.build()
        calls = (list(self.converters.pandoc_calls), list(self.converters.pdf_calls))

        # Coordinator finalization legitimately changes only mutable Stage 8/root state.
        self.fixture.manifest["status"] = "complete"
        self.fixture.manifest["current_stage"] = None
        self.fixture.manifest["stages"][-1]["status"] = "complete"
        self.fixture.write_manifest()
        resumed = self.build(resume=True)

        self.assertEqual(resumed, first)
        self.assertEqual(self.converters.pandoc_calls, calls[0])
        self.assertEqual(self.converters.pdf_calls, calls[1])
        E.validate_package_zip(resumed.package_dir, resumed.zip_path)

    def test_resume_recovers_directory_only_by_atomically_recreating_zip(self):
        first = self.build()
        first.zip_path.unlink()
        calls = (len(self.converters.pandoc_calls), len(self.converters.pdf_calls))

        resumed = self.build(resume=True)

        self.assertEqual(resumed.package_dir, first.package_dir)
        self.assertTrue(resumed.zip_path.is_file())
        self.assertEqual(
            calls,
            (len(self.converters.pandoc_calls), len(self.converters.pdf_calls)),
        )
        self.assertEqual(
            list(resumed.zip_path.parent.glob(f".{resumed.zip_path.name}-recover-*")),
            [],
        )
        E.validate_package_zip(resumed.package_dir, resumed.zip_path)

    def test_resume_rejects_zip_only_and_identity_or_zip_mismatch(self):
        result = self.build()
        package_copy = self.root / "package-copy"
        shutil.copytree(result.package_dir, package_copy)
        shutil.rmtree(result.package_dir)
        with self.assertRaisesRegex(E.PackagingError, "ZIP without.*package directory"):
            self.build(resume=True)

        shutil.copytree(package_copy, result.package_dir)
        package_manifest = json.loads(
            result.package_manifest.read_text(encoding="utf-8")
        )
        package_manifest["country"] = "Jordan"
        write_json(result.package_manifest, package_manifest)
        with self.assertRaisesRegex(E.PackagingError, "identity mismatch for country"):
            self.build(resume=True)

        package_manifest["country"] = "Egypt"
        write_json(result.package_manifest, package_manifest)
        with zipfile.ZipFile(result.zip_path, "a") as archive:
            archive.writestr("unexpected.txt", "tamper")
        with self.assertRaisesRegex(E.PackagingError, "content set does not match"):
            self.build(resume=True)

    def test_resume_rejects_package_without_exact_engine_input_payload(self):
        result = self.build()
        package_manifest = json.loads(
            result.package_manifest.read_text(encoding="utf-8")
        )
        engine_records = [
            record
            for record in package_manifest["files"]
            if record.get("stage_id") == "damm_diagnostic"
            and record.get("artifact_id") == "engine_input"
        ]
        self.assertEqual(len(engine_records), 1)
        (result.package_dir / engine_records[0]["path"]).unlink()
        package_manifest["files"].remove(engine_records[0])
        package_manifest["file_count"] = len(package_manifest["files"])
        write_json(result.package_manifest, package_manifest)

        with self.assertRaisesRegex(E.PackagingError, "exact Stage 1 engine_input"):
            self.build(resume=True)


if __name__ == "__main__":
    unittest.main()
