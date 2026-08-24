#!/usr/bin/env python3
"""DAMM v1.7 — end-to-end build and verification.

Rebuilds every downstream artifact from source and checks each stage. Source of record:
  research/*.json (tiered research)  ·  machine_pass.json (API pass)
  g1_overrides_*.json (assessor)     ·  g2_corrections_*.json (peer review)
  definition_corrections_*.json      ·  definition_notes.json (indicator audit)
Nothing below is hand-edited: every artifact is regenerated, then verified.
"""
import subprocess, json, sys, os, re, html, datetime, time
import openpyxl

HERE = os.path.dirname(os.path.abspath(__file__)); os.chdir(HERE)
sys.path.insert(0, HERE)
SK = ("/Users/randeepsudan/Library/Application Support/Claude/local-agent-mode-sessions/skills-plugin/"
      "7fa72b6b-bf7f-4b6d-9cd3-62aa57ef46d1/b21ce9e4-9f62-4185-8a66-257f1e009678/skills/xlsx")
LOG, FAIL = [], 0

def stage(n, t): LOG.append(f"\n## {n}. {t}\n")
def ok(label, cond, detail=""):
    global FAIL
    if not cond: FAIL += 1
    LOG.append(f"- {'PASS' if cond else '**FAIL**'} — {label}" + (f" · {detail}" if detail else ""))
    return cond

def run(cmd, timeout=300):
    r = subprocess.run(cmd, shell=True, capture_output=True, text=True, timeout=timeout)
    return r.returncode, (r.stdout + r.stderr)

# ── 1 sources present ───────────────────────────────────────────────────────
stage(1, "Sources of record")
for f in (["machine_pass.json", "definition_notes.json"]
          + [f"research/{i}_{b}.json" for i in ("EGY","NGA") for b in "ABCDEFG"]
          + [f"research/{i}_register.json" for i in ("EGY","NGA")]
          + [f"{k}_{i}.json" for k in ("g1_overrides","g2_corrections","definition_corrections") for i in ("egy","nga")]):
    ok(f"`{f}`", os.path.exists(f))

# ── 2 consolidation ─────────────────────────────────────────────────────────
stage(2, "Consolidation — research + machine pass + all correction layers")
rc, out = run("python3 build_inputs.py")
ok("build_inputs.py completes", rc == 0)
for iso in ("EGY","NGA"):
    rows = json.load(open(f"{iso}_v17_input.json"))
    core = {k: v for k, v in rows.items() if not k.startswith("A1-CAND-")}
    ok(f"{iso}: 57 indicator rows + candidates", len(core) == 57 and len(rows) == 59, f"{len(rows)} rows")
    ok(f"{iso}: every row carries a value", all(r.get("value") not in (None, "") for r in core.values()))
    ok(f"{iso}: no level on a gap", all(r["level"] is None for r in core.values() if r["cls"] == "Gap"))
    ok(f"{iso}: every Documented row has a source", all(r.get("src") for r in core.values() if r["cls"] == "Documented"))
    ok(f"{iso}: no T5 row is Documented", not [k for k, r in core.items() if r.get("tier") == "T5" and r["cls"] == "Documented"])

# ── 3 engine ────────────────────────────────────────────────────────────────
stage(3, "Engine")
for iso, c in (("EGY","Egypt"), ("NGA","Nigeria")):
    rc, out = run(f"python3 engine_v17.py {iso}_v17_input.json {iso}_v17.json {c}")
    ok(f"{c}: engine runs", rc == 0)
    d = json.load(open(f"{iso}_v17.json"))
    ok(f"{c}: classes sum to 57", sum(d["counts"].values()) == 57, str(d["counts"]))
    ok(f"{c}: 7 pillars, 4 layers, 6 matrix columns",
       len(d["pillars"]) == 7 and len(d["layers"]) == 4 and len(d["matrix"]) == 6)
    ok(f"{c}: no prerequisite reads Absent on a withheld level",
       not [p for p, v in d["prereq"].items() if v["status"] == "Absent" and d["indicators"][p]["level"] is None])
    ok(f"{c}: every pillar publishes the mean's own denominator",
       all(pl["rated"] == pl["n"] - pl["comp"]["Gap"] - pl["held"] for pl in d["pillars"].values()),
       str({p: (pl["rated"], pl["n"]) for p, pl in d["pillars"].items()}))
    ok(f"{c}: a pillar resting on a minority of its rows cannot read unflagged",
       all(pl["weak"] for pl in d["pillars"].values() if pl["rated"] * 2 <= pl["n"]),
       str([p for p, pl in d["pillars"].items() if pl["rated"] * 2 <= pl["n"] and not pl["weak"]]))
    ok(f"{c}: verify-first carries every Judged row and every gap",
       len(d["verify"]) == d["counts"]["Judged"] + d["counts"]["Gap"],
       f'{len(d["verify"])} listed vs {d["counts"]["Judged"]}J + {d["counts"]["Gap"]}G')
    ok(f"{c}: a mean-driven use-case status publishes its enabler-only mean",
       all(m.get("mean_enabler") is not None for m in d["matrix"].values() if m.get("mean_driven")),
       str([u for u, m in d["matrix"].items() if m.get("mean_driven")]))

# ── 4 report render (QC gate is emit-blocking) ──────────────────────────────
stage(4, "Report render — automated QC blocks emission on failure")
for iso, c in (("EGY","Egypt"), ("NGA","Nigeria")):
    rc, out = run(f"python3 render_v17.py {iso}")
    ok(f"{c}: renders (QC passed)", rc == 0, out.strip().splitlines()[-1][:90] if out.strip() else "")
    t = re.sub(r"\s+", " ", html.unescape(re.sub(r"<[^>]+>", " ", open(f"{c}-DAR-Diagnostic.html").read())))
    ok(f"{c}: QC line states all checks passed", re.search(r"Automated quality control: (\d+)/\1 renderer checks passed", t) is not None)
    leaks = [w for w in ["gauntlet","loop 1","loop-1","prior pass","clean-slate","§13","§9","Protocol.md","Schema.md"]
             if w.lower() in t.lower()]
    ok(f"{c}: standalone (no process history, no internal cross-references)", not leaks, str(leaks))

# ── 5 workbooks ─────────────────────────────────────────────────────────────
stage(5, "Workbooks — rebuild and recalculate")
rc, out = run("python3 build_workbook_v17.py")
ok("builder completes (3 files)", rc == 0)
def lo_is_busy():
    r = subprocess.run("pgrep -f 'soffice.*recalc-lo-profile'", shell=True, capture_output=True, text=True)
    return bool(r.stdout.strip())

def lo_cleanup():
    """Clear leftovers from a previous run. Never kill a LIVE conversion — killing soffice
    mid-write leaves a zombie holding a document lock, which then breaks every later recalc."""
    if lo_is_busy():
        for _ in range(60):
            time.sleep(2)
            if not lo_is_busy(): break
        else:
            subprocess.run("pkill -9 -f 'soffice.*recalc-lo-profile'", shell=True, capture_output=True)
            time.sleep(3)
    subprocess.run("rm -rf ${TMPDIR}recalc-lo-profile-* 2>/dev/null", shell=True, capture_output=True)
    subprocess.run('find workbooks-v1.7 -name ".~lock*" -delete 2>/dev/null', shell=True, capture_output=True)

def recalc_once(path):
    lo_cleanup()
    rc, out = run(f'python3 "{SK}/scripts/recalc.py" "{path}" 280', timeout=380)
    try:
        return json.loads(out[out.find("{"):out.rfind("}")+1])
    except Exception:
        return None

for f in ("Egypt", "Nigeria", "Blank-Template"):
    path = f"workbooks-v1.7/DAMM-v1.7-Scoring-Workbook-{f}.xlsx"
    j = recalc_once(path)
    if not j or j.get("status") != "success":
        j = recalc_once(path)                      # one retry after a full cleanup
    ok(f"{f}: recalculated, zero formula errors",
       bool(j) and j.get("status") == "success" and j.get("total_errors") == 0,
       f"{j.get('total_formulas')} formulas" if j else "recalc returned no result after retry")
lo_cleanup()

# ── 6 parity: workbook formulas vs engine ───────────────────────────────────
stage(6, "Parity — every workbook formula against the engine")
from engine_v17 import MODEL
PIL = ["A1","C1","C2","C3","C4","E1","O1"]; LAY = ["Foundation","Enablers","Transformation","Outcomes"]
PRE = ["2.1","2.9","4.1","3.3","3.11","4.5","4.7","4.9","5.5","5.7","6.14","7.12"]
UCS = ["ADV","SMF","MKT","SCM","FIN","AGI"]; ids = list(MODEL.keys())
for iso, c in (("EGY","Egypt"), ("NGA","Nigeria")):
    e = json.load(open(f"{iso}_v17.json"))
    ws = openpyxl.load_workbook(f"workbooks-v1.7/DAMM-v1.7-Scoring-Workbook-{c}.xlsx", data_only=True)[c]
    n = bad = 0; first = []
    def cmp(label, a, b):
        global n, bad
        n += 1; a = "" if a is None else a; b = "" if b is None else b
        good = (abs(float(a)-float(b)) < 0.005) if (isinstance(a,(int,float)) and isinstance(b,(int,float))) \
               else str(a).strip() == str(b).strip()
        if not good:
            bad += 1
            if len(first) < 5: first.append(f"{label}: wb={a!r} engine={b!r}")
    for i, iid in enumerate(ids):
        r = 5+i; x = e["indicators"][iid]
        cmp(f"{iid} class", ws[f"S{r}"].value, x["cls"])
        cmp(f"{iid} level", ws[f"T{r}"].value, x["level"] or "")
        cmp(f"{iid} stale", ws[f"U{r}"].value, "STALE" if x["stale"] else "")
    for i, p in enumerate(PIL):
        d = e["pillars"][p]
        cmp(f"{p} n", ws[f"B{66+i}"].value, d["n"])
        cmp(f"{p} rated", ws[f"C{66+i}"].value, d["rated"])
        cmp(f"{p} mean", ws[f"D{66+i}"].value, d["mean"])
        cmp(f"{p} band", ws[f"E{66+i}"].value, f"({d['band']})" if d["weak"] else d["band"])
        for j, cl in enumerate(["Measured","Documented","Judged","Gap"]):
            cmp(f"{p} {cl}", ws[f"{chr(70+j)}{66+i}"].value, d["comp"][cl])
        cmp(f"{p} held", ws[f"J{66+i}"].value, d["held"])
        cmp(f"{p} stale", ws[f"K{66+i}"].value, d["stale"])
    for i, L in enumerate(LAY): cmp(f"layer {L}", ws[f"B{76+i}"].value, e["layers"][L])
    cmp("leapfrog", ws["B81"].value, e["leapfrog"]["gap"])
    for i, pid in enumerate(PRE): cmp(f"prereq {pid}", ws[f"D{86+i}"].value, e["prereq"][pid]["status"])
    for i, uc in enumerate(UCS):
        cmp(f"matrix {uc}", ws[f"C{102+i}"].value, e["matrix"][uc]["status"])
        # Ruling 13.12: the workbook column now averages ENABLING rows only, so it is
        # checked against the readiness mean rather than the old mixed one.
        cmp(f"matrix {uc} readiness mean", ws[f"B{102+i}"].value,
            e["matrix"][uc]["mean_readiness"])
    ok(f"{c}: {n-bad}/{n} checks match", bad == 0, "; ".join(first))

# ── 7 ratification apparatus visible in the instrument ──────────────────────
stage(7, "Ratification apparatus carried by the instrument")
notes = json.load(open("definition_notes.json"))
for c in ("Egypt", "Nigeria"):
    w = openpyxl.load_workbook(f"workbooks-v1.7/DAMM-v1.7-Scoring-Workbook-{c}.xlsx")[c]
    ncount = sum(1 for r in range(5, 62) if w.cell(row=r, column=28).value)
    holds = [w.cell(row=r, column=1).value for r in range(5, 62) if w.cell(row=r, column=29).value]
    cands = [w.cell(row=r, column=1).value for r in range(108, 118)
             if str(w.cell(row=r, column=1).value or "").startswith("A1-CAND")]
    core_notes = len([k for k in notes if not k.startswith("A1-CAND")])
    ok(f"{c}: open definition question on every audited row", ncount == core_notes, f"{ncount}/{core_notes}")
    ok(f"{c}: ratification holds recorded", len(holds) > 0, ", ".join(map(str, holds)))
    ok(f"{c}: both candidate rows visible and unscored", len(cands) == 2, ", ".join(map(str, cands)))
    wb = openpyxl.load_workbook(f"workbooks-v1.7/DAMM-v1.7-Scoring-Workbook-{c}.xlsx")
    ok(f"{c}: 7 sheets incl. Tiers and Issues",
       len(wb.sheetnames) == 7 and {"Tiers","Issues"} <= set(wb.sheetnames), ", ".join(wb.sheetnames))

# ── 8 canonical model export ────────────────────────────────────────────────
stage(8, "Canonical model — the export DAR Studio consumes")
rc, out = run("python3 ../../model/export_model.py")
ok("model exports from the engine", rc == 0, out.strip().splitlines()[-1][:90] if out.strip() else "")
rc, out = run("python3 ../../model/test_model_parity.py")
last = out.strip().splitlines()[-1][:110] if out.strip() else ""
ok("model file alone reproduces every engine figure", rc == 0, last)
mj = json.load(open("../../model/DAMM-v1.7-model.json"))
ok("model is versioned and flagged unratified",
   mj["version"] == "1.7" and mj["ratified"] is False and mj["revision"] >= 1,
   f'v{mj["version"]} rev{mj["revision"]} ratified={mj["ratified"]}')
ok("every open decision names the fields it governs",
   len(mj["open_decisions"]) == 12 and all(d.get("governs") is not None for d in mj["open_decisions"]),
   f'{len(mj["open_decisions"])} decisions')

stamp = datetime.datetime.now().strftime("%d %B %Y, %H:%M")
head = (f"# DAMM v1.7 — end-to-end build and verification record\n\n"
        f"Run {stamp}. Every artifact below was regenerated from the sources of record and then checked; "
        f"nothing was hand-edited between generation and verification.\n\n"
        f"**Result: {'ALL CHECKS PASS' if FAIL == 0 else str(FAIL) + ' CHECK(S) FAILED'}**\n")
open("VERIFICATION-RECORD.md", "w").write(head + "\n".join(LOG) + "\n")
print(head)
print("\n".join(LOG[-14:]))
sys.exit(1 if FAIL else 0)
