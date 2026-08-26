#!/usr/bin/env python3
"""Focused, non-mutating tests for verify_workbook_parity.py."""

from __future__ import annotations

import copy
import contextlib
import hashlib
import io
import json
from pathlib import Path
import shutil
import tempfile
import unittest

import openpyxl

import verify_workbook_parity as W


def _fixture_model():
    indicator_ids = [f"I{index:02d}" for index in range(1, 58)]
    prerequisite_ids = indicator_ids[:12]
    indicators = [
        {
            "id": indicator_id,
            "method": "ladder",
            "pillar": W.PILLARS[index % len(W.PILLARS)],
            "layer": W.LAYERS[index % len(W.LAYERS)],
        }
        for index, indicator_id in enumerate(indicator_ids)
    ]
    edges = [
        {
            "id": f"edge:{prerequisite_id}:{use_case}",
            "prerequisite_id": prerequisite_id,
            "use_case_id": use_case,
            "effect": "none",
            "applicability": {"mode": "never"},
        }
        for prerequisite_id in prerequisite_ids for use_case in W.USE_CASES
    ]
    return {
        "version": "test", "revision": 1, "ratified": True, "status": "ratified",
        "indicators": indicators,
        "pillars": list(W.PILLARS), "layers": list(W.LAYERS),
        "use_cases": {use_case: {} for use_case in W.USE_CASES},
        "config": {"readiness_threshold": 2.5},
        "prerequisite_mapping": {
            "ratified": True, "status": "ratified", "revision": 1,
            "prerequisite_ids": prerequisite_ids,
            "use_case_ids": list(W.USE_CASES),
            "status_precedence": ["Blocked", "Unverified", "Partial", "Ready"],
            "edges": edges,
        },
    }


def _fixture_observations(model):
    return {
        row["id"]: {
            "value": f"Reviewed evidence for {row['id']}",
            "src": f"Source {row['id']}",
            "url": f"https://example.test/{row['id']}",
            "tier": "T3", "year": 2026, "level": 3,
            "definition_metadata": {"definition_match": True},
        }
        for row in model["indicators"]
    }


def _fixture_score(model):
    indicator_ids, prerequisite_ids = W._model_inventory(model)
    score = {
        "indicators": {
            indicator_id: {"cls": "Documented", "level": 3, "stale": False}
            for indicator_id in indicator_ids
        },
        "pillars": {}, "layers": {layer: 3.0 for layer in W.LAYERS},
        "leapfrog": {"gap": 0.0, "flag": False, "reading": "No structural flag"},
        "prereq": {
            prerequisite_id: {"status": "Present"}
            for prerequisite_id in prerequisite_ids
        },
        "matrix": {},
    }
    for pillar in W.PILLARS:
        score["pillars"][pillar] = {
            "n": 8, "rated": 8, "held": 0, "mean": 3.0,
            "band": "Established", "margin": 0.0, "weak": False,
            "comp": {"Measured": 0, "Documented": 8, "Judged": 0, "Gap": 0},
            "stale": 0,
        }
    for use_case in W.USE_CASES:
        score["matrix"][use_case] = {
            "n_bearing": 9,
            "basis": {"need": 2, "enabler": 5, "outcome": 2},
            "mean_need": 3.0, "mean_readiness": 3.0, "mean_outcome": 3.0,
            "status": "Ready", "why": "", "mean_driven": False,
            "active_gates": [], "conditional_constraints": [],
            "delivery_risks": [],
            "status_reason": {
                "status": "Ready",
                "status_precedence": ["Blocked", "Unverified", "Partial", "Ready"],
                "drivers": [],
            },
        }
    return score


def _fixture_workbook(path, model, observations, score):
    workbook = openpyxl.Workbook()
    scoring = workbook.active
    scoring.title = "Scoring"
    config = workbook.create_sheet("Config")
    indicator_ids, prerequisite_ids = W._model_inventory(model)

    for offset, indicator_id in enumerate(indicator_ids):
        row = W.FIRST_INDICATOR_ROW + offset
        observation = observations[indicator_id]
        scoring[f"A{row}"] = indicator_id
        for column, field in (("M", "value"), ("N", "src"), ("O", "url"),
                              ("P", "tier"), ("Q", "year"), ("R", "level")):
            scoring[f"{column}{row}"] = observation[field]
        scoring[f"AB{row}"] = "match"
        scoring[f"S{row}"] = score["indicators"][indicator_id]["cls"]
        scoring[f"T{row}"] = score["indicators"][indicator_id]["level"]
        scoring[f"U{row}"] = ""

    for index, pillar in enumerate(W.PILLARS):
        row = W.PILLAR_ROW + index
        payload = score["pillars"][pillar]
        scoring[f"A{row}"] = pillar
        for column, value in zip(
                ("B", "C", "D", "E", "F", "G", "H"),
                (payload["n"], payload["rated"], payload["held"], payload["mean"],
                 payload["band"], payload["margin"], payload["weak"])):
            scoring[f"{column}{row}"] = value
        for column, evidence_class in zip(
                ("I", "J", "K", "L"), ("Measured", "Documented", "Judged", "Gap")):
            scoring[f"{column}{row}"] = payload["comp"][evidence_class]
        scoring[f"M{row}"] = payload["stale"]

    for index, layer in enumerate(W.LAYERS):
        row = W.LAYER_ROW + index
        scoring[f"A{row}"] = layer
        scoring[f"B{row}"] = score["layers"][layer]
    scoring[f"B{W.LEAPFROG_ROW}"] = score["leapfrog"]["gap"]
    scoring[f"C{W.LEAPFROG_ROW}"] = score["leapfrog"]["flag"]
    scoring[f"D{W.LEAPFROG_ROW}"] = score["leapfrog"]["reading"]

    for index, prerequisite_id in enumerate(prerequisite_ids):
        row = W.PREREQUISITE_ROW + index
        scoring[f"A{row}"] = prerequisite_id
        scoring[f"D{row}"] = score["prereq"][prerequisite_id]["status"]

    for index, use_case in enumerate(W.USE_CASES):
        row = W.MATRIX_ROW + index
        payload = score["matrix"][use_case]
        scoring[f"A{row}"] = use_case
        values = (
            payload["n_bearing"], payload["basis"]["need"],
            payload["basis"]["enabler"], payload["basis"]["outcome"],
            payload["mean_need"], payload["mean_readiness"], payload["mean_outcome"],
            payload["status"], payload["why"], payload["mean_driven"], 0, 0,
        )
        for column, value in zip("BCDEFGHIJKLM", values):
            scoring[f"{column}{row}"] = value

    for row, use_case in enumerate(W.USE_CASES, start=2):
        config[f"N{row}"] = use_case
    for index, edge in enumerate(sorted(
            model["prerequisite_mapping"]["edges"],
            key=lambda item: (item["prerequisite_id"], item["use_case_id"]))):
        row = W.MAPPING_ROW + index
        config[f"A{row}"] = edge["id"]
        config[f"B{row}"] = edge["prerequisite_id"]
        config[f"C{row}"] = edge["use_case_id"]
        config[f"D{row}"] = edge["effect"]
        config[f"K{row}"] = "Present"
        config[f"L{row}"] = "inactive"
        for column in "MNOP":
            config[f"{column}{row}"] = ""
    workbook.save(path)


def _summary(_model):
    return {
        "formula_manifest_sha256": "a" * 64,
        "semantic_formula_count": 1226,
        "verification_mode": "static_exact_formula_manifest",
        "runtime_recalculation": "external_release_boundary",
    }


class RuntimeParityTests(unittest.TestCase):
    def setUp(self):
        self.temporary = tempfile.TemporaryDirectory()
        self.root = Path(self.temporary.name)
        self.model = _fixture_model()
        self.observations = _fixture_observations(self.model)
        self.score = _fixture_score(self.model)
        self.workbook = self.root / "release.xlsx"
        _fixture_workbook(self.workbook, self.model, self.observations, self.score)
        self.model_raw = json.dumps(self.model, sort_keys=True).encode()
        self.observations_raw = json.dumps(self.observations, sort_keys=True).encode()
        self.profiles_raw = b"{}"

    def tearDown(self):
        self.temporary.cleanup()

    def _verify(self, recalculator):
        return W.verify_runtime_parity(
            workbook_path=self.workbook,
            model=self.model, model_raw=self.model_raw,
            observations=self.observations, observations_raw=self.observations_raw,
            profiles={}, profiles_raw=self.profiles_raw,
            country="Fixtureland", recalculator=recalculator,
            semantic_check=lambda _raw, _model: True,
            formula_summary_provider=_summary,
            score_provider=lambda _model, _observations, _profiles, _country: (
                copy.deepcopy(self.score), copy.deepcopy(self.score)),
        )

    def test_full_surface_pass_is_content_addressed_and_source_is_unchanged(self):
        before = hashlib.sha256(self.workbook.read_bytes()).hexdigest()

        def recalculate(source, temp_root):
            output = temp_root / "recalculated.xlsx"
            shutil.copyfile(source, output)
            return output, {"implementation": "test-copy", "exit_code": 0}

        evidence = self._verify(recalculate)
        self.assertEqual(evidence["status"], "passed")
        self.assertEqual(evidence["mismatch_count"], 0)
        self.assertEqual(evidence["input_binding_mismatch_count"], 0)
        self.assertEqual(evidence["formula_error_count"], 0)
        self.assertGreater(evidence["comparison_count"], 1000)
        self.assertEqual(evidence["engine_projection_sha256"],
                         evidence["reference_projection_sha256"])
        self.assertEqual(evidence["engine_projection_sha256"],
                         evidence["workbook_projection_sha256"])
        self.assertEqual(hashlib.sha256(self.workbook.read_bytes()).hexdigest(), before)
        self.assertTrue(evidence["source_workbook_unchanged"])
        expected = dict(
            expected_country="Fixtureland",
            expected_workbook_sha256=evidence["workbook_sha256"],
            expected_model_file_sha256=evidence["model_file_sha256"],
            expected_model_payload_sha256=evidence["model_payload_sha256"],
            expected_observations_file_sha256=evidence["observations_file_sha256"],
            expected_observations_payload_sha256=(
                evidence["observations_payload_sha256"]),
            expected_profiles_file_sha256=evidence["profiles_file_sha256"],
            expected_profiles_payload_sha256=evidence["profiles_payload_sha256"],
            expected_formula_manifest_sha256=evidence["formula_manifest_sha256"],
            expected_semantic_formula_count=evidence["semantic_formula_count"],
        )
        self.assertTrue(W.runtime_evidence_is_complete(evidence, **expected))

        mutations = {
            "unequal projection hashes": lambda item: item.__setitem__(
                "workbook_projection_sha256", "b" * 64),
            "output mismatches": lambda item: item.__setitem__("mismatch_count", 1),
            "formula errors": lambda item: item.__setitem__("formula_error_count", 1),
            "input mismatches": lambda item: item.__setitem__(
                "input_binding_mismatch_count", 1),
            "wrong scope": lambda item: item["scope"].__setitem__("mapping_edges", 71),
            "nonzero recalculator exit": lambda item: item["recalculator"].__setitem__(
                "exit_code", 1),
            "wrong workbook binding": lambda item: item.__setitem__(
                "workbook_sha256", "b" * 64),
            "wrong model file binding": lambda item: item.__setitem__(
                "model_file_sha256", "b" * 64),
            "wrong model payload binding": lambda item: item.__setitem__(
                "model_payload_sha256", "b" * 64),
            "wrong observations file binding": lambda item: item.__setitem__(
                "observations_file_sha256", "b" * 64),
            "wrong observations payload binding": lambda item: item.__setitem__(
                "observations_payload_sha256", "b" * 64),
            "wrong profiles file binding": lambda item: item.__setitem__(
                "profiles_file_sha256", "b" * 64),
            "wrong profiles payload binding": lambda item: item.__setitem__(
                "profiles_payload_sha256", "b" * 64),
            "wrong formula binding": lambda item: item.__setitem__(
                "formula_manifest_sha256", "b" * 64),
            "invalid digest": lambda item: item.__setitem__(
                "recalculated_workbook_sha256", "not-a-digest"),
            "missing digest": lambda item: item.pop(
                "reference_projection_sha256"),
            "wrong country": lambda item: item.__setitem__("country", "Elsewhere"),
        }
        for label, mutate in mutations.items():
            with self.subTest(label=label):
                forged = copy.deepcopy(evidence)
                mutate(forged)
                self.assertFalse(W.runtime_evidence_is_complete(
                    forged, **expected))

    def test_recalculated_output_tamper_is_rejected_without_touching_source(self):
        before = self.workbook.read_bytes()

        def recalculate(source, temp_root):
            output = temp_root / "recalculated.xlsx"
            shutil.copyfile(source, output)
            workbook = openpyxl.load_workbook(output)
            workbook["Scoring"]["S2"] = "Judged"
            workbook.save(output)
            return output, {"implementation": "test-tamper", "exit_code": 0}

        evidence = self._verify(recalculate)
        self.assertEqual(evidence["status"], "failed")
        self.assertGreater(evidence["mismatch_count"], 0)
        self.assertTrue(any(item["path"].endswith("indicators.I01.class")
                            for item in evidence["mismatches"]))
        self.assertEqual(self.workbook.read_bytes(), before)

    def test_recalculator_cannot_return_an_external_path(self):
        def recalculate(_source, _temp_root):
            return self.workbook, {"implementation": "unsafe", "exit_code": 0}

        with self.assertRaisesRegex(W.VerificationError, "outside its temporary"):
            self._verify(recalculate)

    def test_input_binding_detects_a_different_observation(self):
        changed = copy.deepcopy(self.observations)
        changed["I01"]["src"] = "Different source"
        original = self.observations
        self.observations = changed
        try:
            def recalculate(source, temp_root):
                output = temp_root / "recalculated.xlsx"
                shutil.copyfile(source, output)
                return output, {"implementation": "test-copy", "exit_code": 0}

            evidence = self._verify(recalculate)
        finally:
            self.observations = original
        self.assertEqual(evidence["status"], "failed")
        self.assertGreater(evidence["input_binding_mismatch_count"], 0)

    def test_nonzero_recalculator_exit_cannot_attest_a_pass(self):
        def recalculate(source, temp_root):
            output = temp_root / "recalculated.xlsx"
            shutil.copyfile(source, output)
            return output, {"implementation": "test-copy", "exit_code": 9}

        evidence = self._verify(recalculate)
        self.assertEqual(evidence["status"], "failed")
        self.assertEqual(evidence["mismatch_count"], 0)
        self.assertEqual(evidence["recalculator"]["exit_code"], 9)

    def test_cli_output_cannot_alias_and_overwrite_an_input(self):
        before = self.workbook.read_bytes()
        with contextlib.redirect_stdout(io.StringIO()):
            exit_code = W.main([
                "--workbook", str(self.workbook),
                "--model", str(self.root / "missing-model.json"),
                "--observations", str(self.root / "missing-observations.json"),
                "--profiles", str(self.root / "missing-profiles.json"),
                "--country", "Fixtureland",
                "--output", str(self.workbook),
            ])
        self.assertEqual(exit_code, 2)
        self.assertEqual(self.workbook.read_bytes(), before)


if __name__ == "__main__":
    unittest.main()
