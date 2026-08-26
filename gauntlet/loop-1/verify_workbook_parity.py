#!/usr/bin/env python3
"""Safe runtime parity verification for the ratified DAMM workbook.

The release workbook has two distinct verification surfaces:

* ``generate_dar._workbook_content_is_semantic`` verifies the exact OOXML formula and
  literal manifest without executing spreadsheet formulas.
* This program copies that already-validated workbook into a temporary directory,
  recalculates the copy with LibreOffice, and compares its cached results with both the
  production engine and the independent reference scorer.

The source workbook is never opened for writing.  The verifier does not import or invoke
``verify_end_to_end.py`` (that script rebuilds tracked artifacts and is intentionally an
unsafe release-time boundary).

CLI contract::

    python3 verify_workbook_parity.py \
      --workbook release.xlsx \
      --model ../../model/DAMM-v1.7-model.json \
      --observations reviewed-observations.json \
      --profiles intervention-profiles.json \
      --country Egypt \
      --output workbook-runtime-parity.json

The model must be ratified and carry the complete 12 x 6 prerequisite mapping.  The
observations payload is the exact 57-row object consumed by both scorers.  A held ladder
row may additionally carry ``workbook_assessor_level`` and
``workbook_definition_match``; the scorers ignore these audit-only literals, while the
verifier binds them to columns R and AB.  Profiles is a use-case keyed object; an omitted
field is deliberately unknown and is represented by a blank Config cell.  Exit 0 means
all comparisons passed, exit 1 means parity failed, and exit 2 means the invocation or
recalculation environment was invalid/unavailable.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import math
import os
from pathlib import Path
import shutil
import subprocess
import sys
import tempfile
from typing import Any, Callable, Iterable, Mapping

import openpyxl


HERE = Path(__file__).resolve().parent
REPO = HERE.parent.parent
MODEL_DIR = REPO / "model"
PIPELINE_DIR = HERE / "research_pipeline"

USE_CASES = ("ADV", "SMF", "MKT", "SCM", "FIN", "AGI")
PILLARS = ("A1", "C1", "C2", "C3", "C4", "E1", "O1")
LAYERS = ("Foundation", "Enablers", "Transformation", "Outcomes")
PROFILE_FIELDS = (
    "targeted_farmer_level_delivery",
    "cross_organization_agricultural_data_sharing",
    "cross_ministerial_delivery",
    "moag_led_or_owned",
    "uses_personal_data",
    "uses_farm_level_data",
    "ai_enabled",
)
PROFILE_COLUMNS = dict(zip(PROFILE_FIELDS, ("O", "P", "Q", "R", "S", "T", "U")))

FIRST_INDICATOR_ROW = 2
PILLAR_ROW = 62
LAYER_ROW = 72
LEAPFROG_ROW = 78
PREREQUISITE_ROW = 82
MATRIX_ROW = 97
MAPPING_ROW = 21

_MISSING = object()
_FORMULA_ERRORS = frozenset({
    "#NULL!", "#DIV/0!", "#VALUE!", "#REF!", "#NAME?", "#NUM!", "#N/A",
    "#GETTING_DATA", "#SPILL!", "#CALC!",
})


class VerificationError(RuntimeError):
    """The supplied release evidence is malformed or unsafe to verify."""


class RecalculationUnavailable(VerificationError):
    """No usable LibreOffice runtime is available."""


def _canonical_json(value: Any) -> bytes:
    return json.dumps(
        value, sort_keys=True, separators=(",", ":"), ensure_ascii=False,
        allow_nan=False,
    ).encode("utf-8")


def _canonical_sha256(value: Any) -> str:
    return hashlib.sha256(_canonical_json(value)).hexdigest()


def _sha256_bytes(raw: bytes) -> str:
    return hashlib.sha256(raw).hexdigest()


def _sha256_path(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        while True:
            chunk = source.read(1024 * 1024)
            if not chunk:
                return digest.hexdigest()
            digest.update(chunk)


def _read_json(path: Path) -> tuple[Any, bytes]:
    raw = path.read_bytes()
    try:
        return json.loads(raw.decode("utf-8")), raw
    except (UnicodeDecodeError, json.JSONDecodeError) as exc:
        raise VerificationError(f"invalid JSON in {path}: {exc}") from exc


def _blank(value: Any) -> bool:
    return value is None or value == ""


def _text(value: Any) -> str:
    return "" if _blank(value) else str(value)


def _integer(value: Any, label: str) -> int | None:
    if _blank(value):
        return None
    if isinstance(value, bool):
        raise VerificationError(f"{label} is boolean, not an integer")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise VerificationError(f"{label} is not numeric: {value!r}") from exc
    if not math.isfinite(number) or not number.is_integer():
        raise VerificationError(f"{label} is not a finite integer: {value!r}")
    return int(number)


def _number(value: Any, label: str) -> float | None:
    if _blank(value):
        return None
    if isinstance(value, bool):
        raise VerificationError(f"{label} is boolean, not numeric")
    try:
        number = float(value)
    except (TypeError, ValueError) as exc:
        raise VerificationError(f"{label} is not numeric: {value!r}") from exc
    if not math.isfinite(number):
        raise VerificationError(f"{label} is not finite: {value!r}")
    return number


def _boolean(value: Any, label: str) -> bool:
    if isinstance(value, bool):
        return value
    if value in (0, 0.0, "0", "FALSE", "False", "false"):
        return False
    if value in (1, 1.0, "1", "TRUE", "True", "true"):
        return True
    raise VerificationError(f"{label} is not boolean: {value!r}")


def _stale(value: Any, label: str) -> bool:
    if _blank(value):
        return False
    if value == "STALE":
        return True
    raise VerificationError(f"{label} is neither blank nor STALE: {value!r}")


def _model_inventory(model: Mapping[str, Any]) -> tuple[list[str], list[str]]:
    if not isinstance(model, dict):
        raise VerificationError("model must be a JSON object")
    if model.get("ratified") is not True or model.get("status") != "ratified":
        raise VerificationError("runtime workbook parity requires a ratified model")
    indicators = model.get("indicators")
    mapping = model.get("prerequisite_mapping")
    if not isinstance(indicators, list) or len(indicators) != 57:
        raise VerificationError("model must contain exactly 57 indicators")
    indicator_ids = [row.get("id") for row in indicators if isinstance(row, dict)]
    if (len(indicator_ids) != 57 or any(not isinstance(item, str) for item in indicator_ids)
            or len(set(indicator_ids)) != 57):
        raise VerificationError("model indicator ids are missing or duplicated")
    if (not isinstance(mapping, dict) or mapping.get("ratified") is not True
            or mapping.get("status") != "ratified"):
        raise VerificationError("model lacks a ratified prerequisite mapping")
    prerequisite_ids = mapping.get("prerequisite_ids")
    edges = mapping.get("edges")
    if (not isinstance(prerequisite_ids, list) or len(prerequisite_ids) != 12
            or len(set(prerequisite_ids)) != 12
            or not isinstance(edges, list) or len(edges) != 72):
        raise VerificationError("mapping must contain 12 prerequisites and 72 edges")
    expected_pairs = {(prerequisite, use_case)
                      for prerequisite in prerequisite_ids for use_case in USE_CASES}
    actual_pairs = {(edge.get("prerequisite_id"), edge.get("use_case_id"))
                    for edge in edges if isinstance(edge, dict)}
    if actual_pairs != expected_pairs or len(actual_pairs) != 72:
        raise VerificationError("mapping is not the complete 12 x 6 graph")
    if set(model.get("pillars") or []) != set(PILLARS):
        raise VerificationError("model pillar inventory differs from the workbook contract")
    if set(model.get("layers") or []) != set(LAYERS):
        raise VerificationError("model layer inventory differs from the workbook contract")
    use_cases = model.get("use_cases")
    if not isinstance(use_cases, dict) or set(use_cases) != set(USE_CASES):
        raise VerificationError("model use-case inventory differs from the workbook contract")
    return sorted(indicator_ids), sorted(prerequisite_ids)


def _normalise_observations(
        observations: Mapping[str, Any], indicator_ids: Iterable[str]) -> dict[str, dict[str, Any]]:
    if not isinstance(observations, dict):
        raise VerificationError("observations must be an indicator-keyed object")
    expected = set(indicator_ids)
    if set(observations) != expected:
        missing = sorted(expected - set(observations))
        extra = sorted(set(observations) - expected)
        raise VerificationError(
            f"observations must contain exactly the 57 model rows; missing={missing}, extra={extra}")
    if any(not isinstance(row, dict) for row in observations.values()):
        raise VerificationError("every observation must be an object")
    return {indicator_id: dict(observations[indicator_id])
            for indicator_id in sorted(expected)}


def _normalise_profiles(profiles: Mapping[str, Any] | None) -> dict[str, dict[str, bool]]:
    if profiles is None:
        profiles = {}
    if not isinstance(profiles, dict) or set(profiles) - set(USE_CASES):
        raise VerificationError("profiles must be keyed only by the six model use cases")
    result: dict[str, dict[str, bool]] = {}
    for use_case, profile in profiles.items():
        if (not isinstance(profile, dict) or set(profile) - set(PROFILE_FIELDS)
                or any(type(value) is not bool for value in profile.values())):
            raise VerificationError(f"profile {use_case} contains invalid intervention facts")
        result[use_case] = dict(profile)
    return result


def _default_semantic_check(raw: bytes, model: dict[str, Any]) -> bool:
    if str(PIPELINE_DIR) not in sys.path:
        sys.path.insert(0, str(PIPELINE_DIR))
    import generate_dar  # Imported lazily: this verifier never imports end-to-end tooling.
    return bool(generate_dar._workbook_content_is_semantic(raw, model))


def _default_formula_summary(model: dict[str, Any]) -> dict[str, Any] | None:
    if str(PIPELINE_DIR) not in sys.path:
        sys.path.insert(0, str(PIPELINE_DIR))
    import generate_dar
    return generate_dar._workbook_formula_manifest_summary(model)


def _default_score_provider(
        model: dict[str, Any], observations: dict[str, dict[str, Any]],
        profiles: dict[str, dict[str, bool]], country: str) -> tuple[dict[str, Any], dict[str, Any]]:
    if str(HERE) not in sys.path:
        sys.path.insert(0, str(HERE))
    if str(MODEL_DIR) not in sys.path:
        sys.path.insert(0, str(MODEL_DIR))
    from engine_v17 import run as engine_run
    from reference_scorer import Scorer as ReferenceScorer

    engine = engine_run(
        country, observations, model_spec=model, intervention_profiles=profiles)
    reference = ReferenceScorer(model).run(observations, intervention_profiles=profiles)
    return engine, reference


def _profile_cell_value(value: Any) -> bool | None | object:
    if _blank(value):
        return None
    if isinstance(value, bool):
        return value
    if isinstance(value, str) and value.lower() in ("true", "false"):
        return value.lower() == "true"
    return _MISSING


def _same_input_value(left: Any, right: Any) -> bool:
    if _blank(left) and _blank(right):
        return True
    if (type(left) in (int, float) and not isinstance(left, bool)
            and type(right) in (int, float) and not isinstance(right, bool)):
        return math.isfinite(float(left)) and math.isfinite(float(right)) and float(left) == float(right)
    return left == right


def _input_binding_mismatches(
        workbook_path: Path, model: dict[str, Any],
        observations: dict[str, dict[str, Any]],
        profiles: dict[str, dict[str, bool]], indicator_ids: list[str]) -> tuple[int, list[dict[str, Any]]]:
    """Bind the external scorer inputs to the literal cells the workbook executes."""
    try:
        workbook = openpyxl.load_workbook(workbook_path, data_only=False, read_only=True)
    except Exception as exc:
        raise VerificationError(f"cannot open source workbook: {exc}") from exc
    try:
        if "Scoring" not in workbook.sheetnames or "Config" not in workbook.sheetnames:
            raise VerificationError("workbook lacks the Scoring or Config sheet")
        scoring = workbook["Scoring"]
        config = workbook["Config"]
        definitions = {row["id"]: row for row in model["indicators"]}
        mismatches: list[dict[str, Any]] = []
        comparisons = 0

        def compare(path: str, expected: Any, actual: Any) -> None:
            nonlocal comparisons
            comparisons += 1
            if not _same_input_value(expected, actual):
                mismatches.append({"path": path, "expected": expected, "actual": actual})

        for offset, indicator_id in enumerate(indicator_ids):
            row_number = FIRST_INDICATOR_ROW + offset
            compare(f"observations.{indicator_id}.indicator_id", indicator_id,
                    scoring[f"A{row_number}"].value)
            observation = observations[indicator_id]
            for field, column in (("value", "M"), ("src", "N"), ("url", "O"),
                                  ("tier", "P"), ("year", "Q")):
                compare(f"observations.{indicator_id}.{field}", observation.get(field, ""),
                        scoring[f"{column}{row_number}"].value)
            method = definitions[indicator_id].get("method")
            # A held ladder row can preserve the assessor's proposed rung in R while the
            # scorer-facing ``level`` is null.  The optional audit-only field binds that
            # pre-hold value without changing either scorer's input semantics.
            expected_assessor_level = (
                observation.get("workbook_assessor_level", observation.get("level"))
                if method == "ladder" else "")
            compare(f"observations.{indicator_id}.assessor_level", expected_assessor_level,
                    scoring[f"R{row_number}"].value)
            metadata = observation.get("definition_metadata")
            definition_match = observation.get(
                "workbook_definition_match",
                metadata.get("definition_match") if isinstance(metadata, dict) else None)
            actual_match = scoring[f"AB{row_number}"].value == "match"
            compare(f"observations.{indicator_id}.definition_match", definition_match, actual_match)

        for use_case_index, use_case in enumerate(USE_CASES, start=2):
            compare(f"profiles.{use_case}.use_case_id", use_case,
                    config[f"N{use_case_index}"].value)
            profile = profiles.get(use_case, {})
            for field, column in PROFILE_COLUMNS.items():
                actual = _profile_cell_value(config[f"{column}{use_case_index}"].value)
                compare(f"profiles.{use_case}.{field}", profile.get(field), actual)
        return comparisons, mismatches
    finally:
        workbook.close()


def _edge_runtime_from_score(
        score: Mapping[str, Any], model: Mapping[str, Any]) -> dict[str, dict[str, str]]:
    result: dict[str, dict[str, str]] = {}
    mapping = model["prerequisite_mapping"]
    conditions: dict[tuple[str, str, str], str] = {}
    for use_case, matrix in score["matrix"].items():
        for condition in matrix.get("conditional_constraints", []):
            conditions[(use_case, condition["prerequisite_id"], condition["effect"])] = (
                condition["evaluation"])
    for edge in sorted(mapping["edges"], key=lambda item: (
            item["prerequisite_id"], item["use_case_id"])):
        edge_id = edge["id"]
        prerequisite_id = edge["prerequisite_id"]
        status = score["prereq"][prerequisite_id]["status"]
        effect = edge["effect"]
        mode = edge["applicability"]["mode"]
        if mode == "never":
            state = "inactive"
        elif mode == "always":
            state = "active"
        else:
            state = conditions.get((edge["use_case_id"], prerequisite_id, effect), "")
        selected = ((edge.get("on_prerequisite_status") or {}).get(status, ""))
        active = selected if state == "active" else ""
        result[edge_id] = {
            "prerequisite_status": status,
            "predicate_state": state,
            "selected_action": selected,
            "active_action": active,
            "delivery_risk": (active if effect == "delivery_risk"
                              and active != "no_change" else ""),
            "gate_candidate": (active if effect == "gate"
                               and active != "no_change" else ""),
        }
    return result


def _readiness_projection(matrix: Mapping[str, Any]) -> dict[str, Any]:
    conditions = list(matrix.get("conditional_constraints") or [])
    risks = list(matrix.get("delivery_risks") or [])
    return {
        "n_bearing": int(matrix["n_bearing"]),
        "basis": {name: int(matrix["basis"][name])
                  for name in ("need", "enabler", "outcome")},
        "mean_need": matrix.get("mean_need"),
        "mean_readiness": matrix.get("mean_readiness"),
        "mean_outcome": matrix.get("mean_outcome"),
        "status": matrix["status"],
        "why": matrix.get("why", ""),
        "mean_driven": bool(matrix.get("mean_driven")),
        "active_gates": list(matrix.get("active_gates") or []),
        "conditional_constraints": conditions,
        "delivery_risks": risks,
        "status_reason": matrix.get("status_reason"),
        "conditional_unresolved": sum(
            item.get("evaluation") == "unresolved" for item in conditions),
        "delivery_risk_count": len(risks),
    }


def projection_from_score(score: Mapping[str, Any], model: Mapping[str, Any]) -> dict[str, Any]:
    """Select the complete runtime surface implemented by the release workbook."""
    indicator_ids, prerequisite_ids = _model_inventory(model)
    try:
        projection = {
            "indicators": {
                indicator_id: {
                    "class": score["indicators"][indicator_id]["cls"],
                    "level": score["indicators"][indicator_id].get("level"),
                    "stale": bool(score["indicators"][indicator_id]["stale"]),
                } for indicator_id in indicator_ids
            },
            "pillars": {
                pillar: {
                    "n": int(score["pillars"][pillar]["n"]),
                    "rated": int(score["pillars"][pillar]["rated"]),
                    "held": int(score["pillars"][pillar]["held"]),
                    "mean": score["pillars"][pillar].get("mean"),
                    "band": score["pillars"][pillar]["band"],
                    "margin": score["pillars"][pillar].get("margin"),
                    "weak": bool(score["pillars"][pillar]["weak"]),
                    "evidence": {
                        evidence_class: int(score["pillars"][pillar]["comp"][evidence_class])
                        for evidence_class in ("Measured", "Documented", "Judged", "Gap")
                    },
                    "stale": int(score["pillars"][pillar]["stale"]),
                } for pillar in PILLARS
            },
            "layers": {layer: score["layers"].get(layer) for layer in LAYERS},
            "leapfrog": {
                "gap": score["leapfrog"].get("gap"),
                "flag": bool(score["leapfrog"].get("flag")),
                "reading": score["leapfrog"]["reading"],
            },
            "prerequisites": {
                prerequisite_id: score["prereq"][prerequisite_id]["status"]
                for prerequisite_id in prerequisite_ids
            },
            "readiness": {
                use_case: _readiness_projection(score["matrix"][use_case])
                for use_case in USE_CASES
            },
            "edge_runtime": _edge_runtime_from_score(score, model),
        }
    except (KeyError, TypeError, ValueError) as exc:
        raise VerificationError(f"scorer output is incomplete: {exc}") from exc
    return projection


def _data_cell(sheet: Any, reference: str) -> Any:
    return sheet[reference].value


def projection_from_workbook(
        recalculated_workbook: Path, model: Mapping[str, Any]) -> dict[str, Any]:
    """Read the recalculated cache at the exact canonical manifest locations."""
    indicator_ids, prerequisite_ids = _model_inventory(model)
    try:
        workbook = openpyxl.load_workbook(
            recalculated_workbook, data_only=True, read_only=True)
    except Exception as exc:
        raise VerificationError(f"cannot open recalculated workbook: {exc}") from exc
    try:
        if "Scoring" not in workbook.sheetnames or "Config" not in workbook.sheetnames:
            raise VerificationError("recalculated workbook lacks Scoring or Config")
        scoring = workbook["Scoring"]
        config = workbook["Config"]
        indicators: dict[str, Any] = {}
        for offset, indicator_id in enumerate(indicator_ids):
            row = FIRST_INDICATOR_ROW + offset
            if _data_cell(scoring, f"A{row}") != indicator_id:
                raise VerificationError(f"Scoring row {row} does not identify {indicator_id}")
            indicators[indicator_id] = {
                "class": _text(_data_cell(scoring, f"S{row}")),
                "level": _integer(_data_cell(scoring, f"T{row}"), f"Scoring!T{row}"),
                "stale": _stale(_data_cell(scoring, f"U{row}"), f"Scoring!U{row}"),
            }

        pillars: dict[str, Any] = {}
        for index, pillar in enumerate(PILLARS):
            row = PILLAR_ROW + index
            if _data_cell(scoring, f"A{row}") != pillar:
                raise VerificationError(f"Scoring row {row} does not identify pillar {pillar}")
            pillars[pillar] = {
                "n": _integer(_data_cell(scoring, f"B{row}"), f"Scoring!B{row}"),
                "rated": _integer(_data_cell(scoring, f"C{row}"), f"Scoring!C{row}"),
                "held": _integer(_data_cell(scoring, f"D{row}"), f"Scoring!D{row}"),
                "mean": _number(_data_cell(scoring, f"E{row}"), f"Scoring!E{row}"),
                "band": _text(_data_cell(scoring, f"F{row}")),
                "margin": _number(_data_cell(scoring, f"G{row}"), f"Scoring!G{row}"),
                "weak": _boolean(_data_cell(scoring, f"H{row}"), f"Scoring!H{row}"),
                "evidence": {
                    evidence_class: _integer(
                        _data_cell(scoring, f"{column}{row}"), f"Scoring!{column}{row}")
                    for column, evidence_class in zip(
                        ("I", "J", "K", "L"),
                        ("Measured", "Documented", "Judged", "Gap"))
                },
                "stale": _integer(_data_cell(scoring, f"M{row}"), f"Scoring!M{row}"),
            }

        layers: dict[str, Any] = {}
        for index, layer in enumerate(LAYERS):
            row = LAYER_ROW + index
            if _data_cell(scoring, f"A{row}") != layer:
                raise VerificationError(f"Scoring row {row} does not identify layer {layer}")
            layers[layer] = _number(_data_cell(scoring, f"B{row}"), f"Scoring!B{row}")

        leapfrog = {
            "gap": _number(_data_cell(scoring, f"B{LEAPFROG_ROW}"),
                           f"Scoring!B{LEAPFROG_ROW}"),
            "flag": _boolean(_data_cell(scoring, f"C{LEAPFROG_ROW}"),
                             f"Scoring!C{LEAPFROG_ROW}"),
            "reading": _text(_data_cell(scoring, f"D{LEAPFROG_ROW}")),
        }

        prerequisites: dict[str, str] = {}
        for index, prerequisite_id in enumerate(prerequisite_ids):
            row = PREREQUISITE_ROW + index
            if _data_cell(scoring, f"A{row}") != prerequisite_id:
                raise VerificationError(
                    f"Scoring row {row} does not identify prerequisite {prerequisite_id}")
            prerequisites[prerequisite_id] = _text(_data_cell(scoring, f"D{row}"))

        edges = sorted(model["prerequisite_mapping"]["edges"], key=lambda item: (
            item["prerequisite_id"], item["use_case_id"]))
        edge_runtime: dict[str, dict[str, str]] = {}
        for index, edge in enumerate(edges):
            row = MAPPING_ROW + index
            if (_data_cell(config, f"A{row}") != edge["id"]
                    or _data_cell(config, f"B{row}") != edge["prerequisite_id"]
                    or _data_cell(config, f"C{row}") != edge["use_case_id"]):
                raise VerificationError(f"Config row {row} does not identify mapping edge {edge['id']}")
            edge_runtime[edge["id"]] = {
                "prerequisite_status": _text(_data_cell(config, f"K{row}")),
                "predicate_state": _text(_data_cell(config, f"L{row}")),
                # Calc evaluates a reference to an empty mapping outcome cell as
                # numeric zero.  On a ``none`` edge the selected action is
                # semantically not applicable, so canonicalise that spreadsheet
                # representation back to the model's empty action.
                "selected_action": (
                    "" if edge["effect"] == "none"
                    and _data_cell(config, f"M{row}") in (None, "", 0, 0.0, "0")
                    else _text(_data_cell(config, f"M{row}"))),
                "active_action": _text(_data_cell(config, f"N{row}")),
                "delivery_risk": _text(_data_cell(config, f"O{row}")),
                "gate_candidate": _text(_data_cell(config, f"P{row}")),
            }

        readiness: dict[str, Any] = {}
        threshold = model["config"]["readiness_threshold"]
        precedence = list(model["prerequisite_mapping"]["status_precedence"])
        for index, use_case in enumerate(USE_CASES):
            row = MATRIX_ROW + index
            if _data_cell(scoring, f"A{row}") != use_case:
                raise VerificationError(f"Scoring row {row} does not identify use case {use_case}")
            status = _text(_data_cell(scoring, f"I{row}"))
            mean_readiness = _number(_data_cell(scoring, f"G{row}"), f"Scoring!G{row}")
            active_gates, conditions, risks, drivers = [], [], [], []
            for edge in (item for item in edges if item["use_case_id"] == use_case):
                runtime = edge_runtime[edge["id"]]
                mode = edge["applicability"]["mode"]
                if edge["effect"] == "gate" and runtime["predicate_state"] == "active":
                    active_gates.append({
                        "prerequisite_id": edge["prerequisite_id"],
                        "prerequisite_status": runtime["prerequisite_status"],
                        "outcome": runtime["selected_action"],
                        "applicability": mode,
                    })
                if mode == "conditional" and edge["effect"] != "none":
                    conditions.append({
                        "prerequisite_id": edge["prerequisite_id"],
                        "effect": edge["effect"],
                        "predicate": edge["applicability"]["predicate"],
                        "prerequisite_status": runtime["prerequisite_status"],
                        "outcome_if_active": runtime["selected_action"],
                        "evaluation": runtime["predicate_state"],
                    })
                if runtime["delivery_risk"]:
                    risks.append({
                        "prerequisite_id": edge["prerequisite_id"],
                        "prerequisite_status": runtime["prerequisite_status"],
                        "action": runtime["delivery_risk"],
                        "applicability": mode,
                    })
                if runtime["gate_candidate"] == status:
                    drivers.append({
                        "type": "gate",
                        "prerequisite_id": edge["prerequisite_id"],
                        "prerequisite_status": runtime["prerequisite_status"],
                    })
            if status == "Partial" and mean_readiness is not None and mean_readiness < threshold:
                drivers.append({
                    "type": "readiness_mean", "mean_readiness": mean_readiness,
                    "threshold": threshold,
                })
            active_gates.sort(key=lambda item: item["prerequisite_id"])
            conditions.sort(key=lambda item: (item["prerequisite_id"], item["effect"]))
            risks.sort(key=lambda item: item["prerequisite_id"])
            drivers.sort(key=lambda item: (
                0 if item["type"] == "gate" else 1, item.get("prerequisite_id", "")))
            readiness[use_case] = {
                "n_bearing": _integer(_data_cell(scoring, f"B{row}"), f"Scoring!B{row}"),
                "basis": {
                    "need": _integer(_data_cell(scoring, f"C{row}"), f"Scoring!C{row}"),
                    "enabler": _integer(_data_cell(scoring, f"D{row}"), f"Scoring!D{row}"),
                    "outcome": _integer(_data_cell(scoring, f"E{row}"), f"Scoring!E{row}"),
                },
                "mean_need": _number(_data_cell(scoring, f"F{row}"), f"Scoring!F{row}"),
                "mean_readiness": mean_readiness,
                "mean_outcome": _number(_data_cell(scoring, f"H{row}"), f"Scoring!H{row}"),
                "status": status,
                "why": _text(_data_cell(scoring, f"J{row}")),
                "mean_driven": _boolean(_data_cell(scoring, f"K{row}"), f"Scoring!K{row}"),
                "active_gates": active_gates,
                "conditional_constraints": conditions,
                "delivery_risks": risks,
                "status_reason": {
                    "status": status, "status_precedence": precedence, "drivers": drivers,
                },
                "conditional_unresolved": _integer(
                    _data_cell(scoring, f"L{row}"), f"Scoring!L{row}"),
                "delivery_risk_count": _integer(
                    _data_cell(scoring, f"M{row}"), f"Scoring!M{row}"),
            }

        return {
            "indicators": indicators, "pillars": pillars, "layers": layers,
            "leapfrog": leapfrog, "prerequisites": prerequisites,
            "readiness": readiness, "edge_runtime": edge_runtime,
        }
    finally:
        workbook.close()


def _formula_error_count(workbook_path: Path) -> int:
    try:
        workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
    except Exception as exc:
        raise VerificationError(f"cannot inspect recalculated formula caches: {exc}") from exc
    try:
        count = 0
        for worksheet in workbook.worksheets:
            for row in worksheet.iter_rows():
                for cell in row:
                    if cell.data_type == "e" or (isinstance(cell.value, str)
                                                  and cell.value in _FORMULA_ERRORS):
                        count += 1
        return count
    finally:
        workbook.close()


def _flatten(value: Any, path: str = "") -> dict[str, Any]:
    if isinstance(value, dict):
        result: dict[str, Any] = {}
        for key in sorted(value):
            child = f"{path}.{key}" if path else str(key)
            result.update(_flatten(value[key], child))
        return result
    if isinstance(value, list):
        result = {}
        for index, item in enumerate(value):
            result.update(_flatten(item, f"{path}[{index}]"))
        if not value:
            result[path] = []
        return result
    return {path: value}


def _compare_projections(
        expected: Mapping[str, Any], actual: Mapping[str, Any], pair: str) -> tuple[int, list[dict[str, Any]]]:
    expected_flat, actual_flat = _flatten(expected), _flatten(actual)
    paths = sorted(set(expected_flat) | set(actual_flat))
    mismatches = []
    for path in paths:
        left, right = expected_flat.get(path, _MISSING), actual_flat.get(path, _MISSING)
        if left != right:
            mismatches.append({
                "pair": pair, "path": path,
                "expected": "<missing>" if left is _MISSING else left,
                "actual": "<missing>" if right is _MISSING else right,
            })
    return len(paths), mismatches


def _find_libreoffice(explicit: str | None = None) -> str:
    # Codex ships a genuinely headless LibreOffice build.  Prefer it on desktop:
    # the ordinary macOS app binary may abort inside the app sandbox even though
    # ``soffice --version`` succeeds.
    bundled = (Path.home() / ".cache" / "codex-runtimes" /
               "codex-primary-runtime" / "dependencies" / "bin" /
               "override" / "soffice")
    candidates = ([explicit] if explicit else
                  [str(bundled), "libreoffice", "soffice"])
    for candidate in candidates:
        if not candidate:
            continue
        found = shutil.which(candidate)
        if found:
            return found
    raise RecalculationUnavailable(
        "LibreOffice was not found (looked for libreoffice/soffice)")


def _libreoffice_recalculate(
        source_copy: Path, temp_root: Path, executable: str | None = None,
        timeout: int = 240) -> tuple[Path, dict[str, Any]]:
    """Recalculate a temporary input copy into a separate temporary output directory."""
    binary = _find_libreoffice(executable)
    output_dir = temp_root / "recalculated"
    profile_dir = temp_root / "libreoffice-profile"
    output_dir.mkdir()
    profile_dir.mkdir()
    xdg_config = profile_dir / "xdg-config"
    xdg_cache = profile_dir / "xdg-cache"
    xdg_data = profile_dir / "xdg-data"
    for directory in (xdg_config, xdg_cache, xdg_data):
        directory.mkdir()
    command = [
        binary,
        f"-env:UserInstallation={profile_dir.resolve().as_uri()}",
        "--invisible", "--headless", "--nologo", "--nodefault", "--nolockcheck",
        "--norestore", "--nofirststartwizard", "--convert-to",
        "xlsx:Calc MS Excel 2007 XML",
        "--outdir", str(output_dir), str(source_copy),
    ]
    environment = os.environ.copy()
    environment.update({
        "HOME": str(profile_dir),
        "XDG_CONFIG_HOME": str(xdg_config),
        "XDG_CACHE_HOME": str(xdg_cache),
        "XDG_DATA_HOME": str(xdg_data),
    })
    # Codex desktop can inherit a per-app macOS TMPDIR that makes headless
    # LibreOffice abort before it opens a document.  Keep all LO scratch state in the
    # system's stable writable temp root while retaining the per-run isolated profile.
    if sys.platform == "darwin" and Path("/private/tmp").is_dir():
        environment.update({
            "TMPDIR": "/private/tmp", "TEMP": "/private/tmp", "TMP": "/private/tmp",
        })
    try:
        process = subprocess.run(
            command, capture_output=True, text=True, timeout=timeout, check=False,
            env=environment)
    except (OSError, subprocess.TimeoutExpired) as exc:
        raise RecalculationUnavailable(f"LibreOffice recalculation failed: {exc}") from exc
    output_path = output_dir / source_copy.name
    if process.returncode != 0 or not output_path.is_file():
        detail = (process.stdout + process.stderr).strip()[-1000:]
        raise RecalculationUnavailable(
            f"LibreOffice did not emit a recalculated workbook (exit {process.returncode}): {detail}")
    return output_path, {
        "implementation": "LibreOffice",
        "executable": str(Path(binary).resolve()),
        "exit_code": process.returncode,
        "stdout_sha256": _sha256_bytes(process.stdout.encode("utf-8")),
        "stderr_sha256": _sha256_bytes(process.stderr.encode("utf-8")),
    }


Recalculator = Callable[[Path, Path], tuple[Path, dict[str, Any]]]


def verify_runtime_parity(
        *, workbook_path: Path, model: dict[str, Any], model_raw: bytes,
        observations: Mapping[str, Any], observations_raw: bytes,
        profiles: Mapping[str, Any] | None, profiles_raw: bytes,
        country: str,
        recalculator: Recalculator | None = None,
        semantic_check: Callable[[bytes, dict[str, Any]], bool] | None = None,
        formula_summary_provider: Callable[[dict[str, Any]], dict[str, Any] | None] | None = None,
        score_provider: Callable[[dict[str, Any], dict[str, dict[str, Any]],
                                 dict[str, dict[str, bool]], str],
                                tuple[dict[str, Any], dict[str, Any]]] | None = None,
        libreoffice: str | None = None) -> dict[str, Any]:
    """Verify one workbook and return a content-addressed evidence record."""
    workbook_path = workbook_path.resolve()
    if not workbook_path.is_file():
        raise VerificationError(f"workbook does not exist: {workbook_path}")
    if not isinstance(country, str) or not country.strip():
        raise VerificationError("country must be a non-empty string")
    indicator_ids, _ = _model_inventory(model)
    clean_observations = _normalise_observations(observations, indicator_ids)
    clean_profiles = _normalise_profiles(profiles)
    source_raw = workbook_path.read_bytes()
    source_sha256 = _sha256_bytes(source_raw)

    semantic_check = semantic_check or _default_semantic_check
    formula_summary_provider = formula_summary_provider or _default_formula_summary
    if not semantic_check(source_raw, model):
        raise VerificationError("source workbook does not match the canonical static manifest")
    formula_summary = formula_summary_provider(model)
    if not isinstance(formula_summary, dict):
        raise VerificationError("canonical workbook formula manifest is unavailable")

    score_provider = score_provider or _default_score_provider
    engine_score, reference_score = score_provider(
        model, clean_observations, clean_profiles, country)
    engine_projection = projection_from_score(engine_score, model)
    reference_projection = projection_from_score(reference_score, model)

    stable_temp_root = ("/private/tmp" if sys.platform == "darwin"
                        and Path("/private/tmp").is_dir() else None)
    with tempfile.TemporaryDirectory(
            prefix="damm-workbook-parity-", dir=stable_temp_root) as temporary:
        temp_root = Path(temporary).resolve()
        input_dir = temp_root / "input"
        input_dir.mkdir()
        source_copy = input_dir / "workbook.xlsx"
        shutil.copyfile(workbook_path, source_copy)
        if _sha256_path(source_copy) != source_sha256:
            raise VerificationError("source workbook changed while it was being copied")
        input_comparisons, input_mismatches = _input_binding_mismatches(
            source_copy, model, clean_observations, clean_profiles, indicator_ids)
        if recalculator is None:
            recalculated, recalculator_evidence = _libreoffice_recalculate(
                source_copy, temp_root, executable=libreoffice)
        else:
            recalculated, recalculator_evidence = recalculator(source_copy, temp_root)
        recalculated = Path(recalculated).resolve()
        try:
            recalculated.relative_to(temp_root)
        except ValueError as exc:
            raise VerificationError(
                "recalculator returned a path outside its temporary workspace") from exc
        if not recalculated.is_file():
            raise VerificationError("recalculator did not return a workbook file")
        recalculated_sha256 = _sha256_path(recalculated)
        formula_errors = _formula_error_count(recalculated)
        workbook_projection = projection_from_workbook(recalculated, model)

    source_unchanged = _sha256_path(workbook_path) == source_sha256
    engine_reference_count, engine_reference_mismatches = _compare_projections(
        engine_projection, reference_projection, "engine-reference")
    engine_workbook_count, engine_workbook_mismatches = _compare_projections(
        engine_projection, workbook_projection, "engine-workbook")
    output_mismatches = engine_reference_mismatches + engine_workbook_mismatches
    mismatch_count = len(output_mismatches)
    recalculator_exit_code = recalculator_evidence.get("exit_code")
    passed = (source_unchanged and formula_errors == 0 and not input_mismatches
              and mismatch_count == 0 and type(recalculator_exit_code) is int
              and recalculator_exit_code == 0)
    evidence = {
        "schema": "damm.workbook-runtime-parity/v1",
        "status": "passed" if passed else "failed",
        "country": country,
        "workbook_sha256": source_sha256,
        "recalculated_workbook_sha256": recalculated_sha256,
        "model_file_sha256": _sha256_bytes(model_raw),
        "model_payload_sha256": _canonical_sha256(model),
        "observations_file_sha256": _sha256_bytes(observations_raw),
        "observations_payload_sha256": _canonical_sha256(clean_observations),
        "profiles_file_sha256": _sha256_bytes(profiles_raw),
        "profiles_payload_sha256": _canonical_sha256(clean_profiles),
        "formula_manifest_sha256": formula_summary.get("formula_manifest_sha256"),
        "semantic_formula_count": formula_summary.get("semantic_formula_count"),
        "static_verification_mode": formula_summary.get("verification_mode"),
        "runtime_recalculation_boundary": formula_summary.get("runtime_recalculation"),
        "engine_projection_sha256": _canonical_sha256(engine_projection),
        "reference_projection_sha256": _canonical_sha256(reference_projection),
        "workbook_projection_sha256": _canonical_sha256(workbook_projection),
        "comparison_count": engine_reference_count + engine_workbook_count,
        "mismatch_count": mismatch_count,
        "mismatches": output_mismatches[:200],
        "mismatches_truncated": mismatch_count > 200,
        "input_binding_comparison_count": input_comparisons,
        "input_binding_mismatch_count": len(input_mismatches),
        "input_binding_mismatches": input_mismatches[:200],
        "formula_error_count": formula_errors,
        "source_workbook_unchanged": source_unchanged,
        "scope": {
            "indicators": 57,
            "indicator_outputs": ["class", "level", "stale"],
            "pillars": 7,
            "pillar_outputs": [
                "n", "rated", "held", "mean", "band", "margin", "weak",
                "evidence.Measured", "evidence.Documented", "evidence.Judged",
                "evidence.Gap", "stale",
            ],
            "layers": 4,
            "leapfrog_outputs": ["gap", "flag", "reading"],
            "prerequisites": 12,
            "mapped_readiness_outputs": 6,
            "mapping_edges": 72,
        },
        "recalculator": recalculator_evidence,
    }
    return evidence


def _is_sha256(value: Any) -> bool:
    return (isinstance(value, str) and len(value) == 64
            and all(character in "0123456789abcdef" for character in value))


def runtime_evidence_is_complete(
        record: Any, *, expected_country: str,
        expected_workbook_sha256: str,
        expected_model_file_sha256: str,
        expected_model_payload_sha256: str,
        expected_observations_file_sha256: str,
        expected_observations_payload_sha256: str,
        expected_profiles_file_sha256: str,
        expected_profiles_payload_sha256: str,
        expected_formula_manifest_sha256: str,
        expected_semantic_formula_count: int) -> bool:
    """Fail-closed validator for embedding this record in release evidence.

    ``generate_dar`` mirrors this predicate for its ``single_source_parity`` release
    record and adds authoritative scorer replay plus strict LibreOffice provenance.
    File and payload hashes are arguments because the release manifest, not a
    self-assertion inside the runtime record, is their source of authority.
    """
    if not isinstance(record, dict):
        return False
    scope = record.get("scope")
    recalculator = record.get("recalculator")
    digest_fields = (
        "workbook_sha256", "recalculated_workbook_sha256", "model_file_sha256",
        "model_payload_sha256", "observations_file_sha256",
        "observations_payload_sha256", "profiles_file_sha256",
        "profiles_payload_sha256", "formula_manifest_sha256",
        "engine_projection_sha256", "reference_projection_sha256",
        "workbook_projection_sha256",
    )
    if (record.get("schema") != "damm.workbook-runtime-parity/v1"
            or record.get("status") != "passed"
            or record.get("country") != expected_country
            or any(not _is_sha256(record.get(field)) for field in digest_fields)
            or record.get("workbook_sha256") != expected_workbook_sha256
            or record.get("model_file_sha256") != expected_model_file_sha256
            or record.get("model_payload_sha256") != expected_model_payload_sha256
            or record.get("observations_file_sha256")
            != expected_observations_file_sha256
            or record.get("observations_payload_sha256")
            != expected_observations_payload_sha256
            or record.get("profiles_file_sha256") != expected_profiles_file_sha256
            or record.get("profiles_payload_sha256")
            != expected_profiles_payload_sha256
            or record.get("formula_manifest_sha256")
            != expected_formula_manifest_sha256
            or record.get("semantic_formula_count")
            != expected_semantic_formula_count
            or record.get("static_verification_mode")
            != "static_exact_formula_manifest"
            or record.get("runtime_recalculation_boundary")
            != "external_release_boundary"
            or record.get("engine_projection_sha256")
            != record.get("reference_projection_sha256")
            or record.get("engine_projection_sha256")
            != record.get("workbook_projection_sha256")
            or type(record.get("comparison_count")) is not int
            or record["comparison_count"] <= 0
            or type(record.get("mismatch_count")) is not int
            or record["mismatch_count"] != 0
            or record.get("mismatches") != []
            or record.get("mismatches_truncated") is not False
            or type(record.get("input_binding_comparison_count")) is not int
            or record["input_binding_comparison_count"] <= 0
            or type(record.get("input_binding_mismatch_count")) is not int
            or record["input_binding_mismatch_count"] != 0
            or record.get("input_binding_mismatches") != []
            or type(record.get("formula_error_count")) is not int
            or record["formula_error_count"] != 0
            or record.get("source_workbook_unchanged") is not True
            or not isinstance(recalculator, dict)
            or not isinstance(recalculator.get("implementation"), str)
            or not recalculator["implementation"].strip()
            or type(recalculator.get("exit_code")) is not int
            or recalculator["exit_code"] != 0
            or not isinstance(scope, dict)
            or scope.get("indicators") != 57
            or scope.get("pillars") != 7
            or scope.get("layers") != 4
            or scope.get("prerequisites") != 12
            or scope.get("mapped_readiness_outputs") != 6
            or scope.get("mapping_edges") != 72
            or scope.get("indicator_outputs") != ["class", "level", "stale"]
            or scope.get("pillar_outputs") != [
                "n", "rated", "held", "mean", "band", "margin", "weak",
                "evidence.Measured", "evidence.Documented", "evidence.Judged",
                "evidence.Gap", "stale",
            ]
            or scope.get("leapfrog_outputs") != ["gap", "flag", "reading"]):
        return False
    return True


def _atomic_write_json(path: Path, payload: Mapping[str, Any]) -> None:
    path = path.resolve()
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary = tempfile.mkstemp(
        prefix=f".{path.name}.", suffix=".tmp", dir=path.parent)
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8") as target:
            json.dump(payload, target, indent=2, sort_keys=True, ensure_ascii=False,
                      allow_nan=False)
            target.write("\n")
            target.flush()
            os.fsync(target.fileno())
        os.replace(temporary, path)
    except Exception:
        try:
            os.unlink(temporary)
        except FileNotFoundError:
            pass
        raise


def _paths_alias(left: Path, right: Path) -> bool:
    """Catch lexical, symlink, and existing-hard-link aliases."""
    if left.resolve() == right.resolve():
        return True
    try:
        return left.exists() and right.exists() and os.path.samefile(left, right)
    except OSError:
        return False


def _parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--model", required=True, type=Path)
    parser.add_argument("--observations", required=True, type=Path)
    parser.add_argument("--profiles", required=True, type=Path)
    parser.add_argument("--country", required=True)
    parser.add_argument("--output", type=Path,
                        help="write the evidence JSON atomically (stdout when omitted)")
    parser.add_argument("--libreoffice",
                        help="explicit LibreOffice/soffice executable")
    return parser


def main(argv: list[str] | None = None) -> int:
    arguments = _parser().parse_args(argv)
    if arguments.output:
        aliased = next((
            input_path for input_path in (
                arguments.workbook, arguments.model, arguments.observations,
                arguments.profiles)
            if _paths_alias(arguments.output, input_path)
        ), None)
        if aliased is not None:
            # Do not write even an error record: the requested destination is an input.
            print(json.dumps({
                "schema": "damm.workbook-runtime-parity/v1",
                "status": "invalid",
                "error": f"--output aliases input path {aliased}",
            }, indent=2, sort_keys=True, ensure_ascii=False))
            return 2
    try:
        model, model_raw = _read_json(arguments.model)
        observations, observations_raw = _read_json(arguments.observations)
        profiles, profiles_raw = _read_json(arguments.profiles)
        evidence = verify_runtime_parity(
            workbook_path=arguments.workbook,
            model=model, model_raw=model_raw,
            observations=observations, observations_raw=observations_raw,
            profiles=profiles, profiles_raw=profiles_raw,
            country=arguments.country, libreoffice=arguments.libreoffice)
        if arguments.output:
            _atomic_write_json(arguments.output, evidence)
        else:
            print(json.dumps(evidence, indent=2, sort_keys=True, ensure_ascii=False))
        return 0 if evidence["status"] == "passed" else 1
    except RecalculationUnavailable as exc:
        payload = {
            "schema": "damm.workbook-runtime-parity/v1",
            "status": "unavailable", "error": str(exc),
        }
    except (VerificationError, OSError, ValueError, KeyError, TypeError) as exc:
        payload = {
            "schema": "damm.workbook-runtime-parity/v1",
            "status": "invalid", "error": str(exc),
        }
    if arguments.output:
        _atomic_write_json(arguments.output, payload)
    else:
        print(json.dumps(payload, indent=2, sort_keys=True, ensure_ascii=False))
    return 2


if __name__ == "__main__":
    raise SystemExit(main())
